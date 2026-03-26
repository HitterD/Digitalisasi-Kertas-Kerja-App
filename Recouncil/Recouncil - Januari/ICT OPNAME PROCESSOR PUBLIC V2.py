
"""
================================================================================
ICT OPNAME PROCESSOR V4 - Enhanced Edition
================================================================================

DESKRIPSI:
Aplikasi untuk memproses file opname ICT dengan interface modern menggunakan 
ttkbootstrap. Aplikasi ini terdiri dari 4 modul utama untuk berbagai keperluan
processing data opname.

MODUL APLIKASI:
1. DAFT ICT FILE PROCESSOR - Memproses file allocation user aktiva
2. EXTRACT OPNAME - Ekstrak dan proses data opname  
3. EXTRACT MAT - Ekstrak data MAT dari file opname
4. EXTRACT RECOUNCIL - Proses data recouncil dengan filtering

FITUR UTAMA:
- Modern dark theme interface dengan ttkbootstrap
- Sheet selection untuk setiap file Excel yang diupload
- Comprehensive error handling dan logging
- Progress tracking untuk semua operasi
- Template-based processing untuk konsistensi data

DEVELOPER NOTES:
- Semua processing menggunakan template/sheet yang sudah ditentukan
- Setiap file upload memiliki opsi pemilihan sheet
- Error handling komprehensif di setiap level
- UI menggunakan ModernBaseWindow untuk konsistensi

VERSION: 4.1 Enhanced Edition - BAT Filter & Year Processing Fixed
AUTHOR: ICT Development Team
DATE: 2024
UPDATE: Enhanced BAT Filter yang benar-benar berfungsi + Tahun dari kolom K/P
================================================================================
"""

# =========================
# IMPORTS
# =========================
# Built-in modules
import sys
import os
import re
import traceback
from datetime import date, timedelta, datetime
from queue import Queue
import threading
import locale
import json
import hashlib
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
import pickle

# Third-party modules
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, font
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from ttkbootstrap.tooltip import ToolTip
from ttkbootstrap.dialogs import Messagebox
from ttkbootstrap.scrolled import ScrolledText
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
try:
    import pyarrow as pa
    import pyarrow.parquet as pq
    PARQUET_AVAILABLE = True
except ImportError:
    PARQUET_AVAILABLE = False

# =========================
# KONSTANTA & STYLE GLOBAL
# =========================
BLUE_FILL = PatternFill(start_color="5C97FA", end_color="5C97FA", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
PURPLE_FILL = PatternFill(start_color="AF52FF", end_color="AF52FF", fill_type="solid")
DEFAULT_FONT_EXCEL = Font(name='Calibri', size=12)
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Modern ttkbootstrap Configuration
BOOTSTRAP_THEME = 'darkly'  # Pure dark theme
FONT_CONFIG = {
    "title": ("Segoe UI", 16, "bold"),
    "subtitle": ("Segoe UI", 11),
    "button": ("Segoe UI", 10, "bold"),
    "text": ("Segoe UI", 9),
    "small": ("Segoe UI", 8)
}

STYLE = {
    "background": "#2E2E2E",
    "widget_bg": "#3C3C3C",
    "text": "#FFFFFF",
    "accent": "#0078D7",
    "accent_hover": "#005A9E",
    "success": "#2ECC71",
    "error": "#E74C3C",
    "font_title": ("Segoe UI", 12, "bold"),
    "font_main": ("Segoe UI", 10),
    "font_button": ("Segoe UI", 11, "bold")
}

# =========================
# BASELINE & DIFF CONFIGURATION
# =========================

# Highlight Colors (Hex format for Excel)
HIGHLIGHT_NEW = "FFFF00"        # Yellow for new records
HIGHLIGHT_CHANGED = "5C97FA"    # Blue for changed records
HIGHLIGHT_UNCHANGED = None      # No highlight for unchanged

# Sheet Names
SHEET_RECOUNCIL = "Recouncil"
SHEET_SUMMARY_QA = "Summary & QA"
SHEET_ERROR_WARNING = "Error/Warning"

# Baseline Configuration
CONFIG_DIR = "config"
BASELINE_CONFIG_FILE = "baseline.json"
CLEAN_BASELINE_SUFFIX = "(CLEAN BASELINE)"

# Comparison Columns for Diff Detection
DIFF_COLUMNS = ['ASSET ORACLE', 'LOKASI', 'JENIS HARTA', 'KONDISI', 'Tahun Perolehan']
JOIN_KEY = 'NO BARCODE'

# Error/Warning Severity Levels
SEVERITY_ERROR = "ERROR"
SEVERITY_WARN = "WARN"

# Enhanced Configuration
CONFIG_DIR = os.path.join(os.path.dirname(__file__), "config")
CACHE_DIR = os.path.join(os.path.dirname(__file__), "cache")
RULES_FILE = os.path.join(CONFIG_DIR, "column_mapping_rules.json")
LOCATION_DICT_FILE = os.path.join(CONFIG_DIR, "location_dictionary.json")
JENIS_HARTA_DICT_FILE = os.path.join(CONFIG_DIR, "jenis_harta_dictionary.json")
CHECKPOINT_FILE = os.path.join(CONFIG_DIR, "processing_checkpoint.json")

# Ensure directories exist
for directory in [CONFIG_DIR, CACHE_DIR]:
    if not os.path.exists(directory):
        os.makedirs(directory)

#########################################################
# HELPER FUNCTIONS FOR SHEET SELECTION
#########################################################

def get_excel_sheets(file_path):
    """
    Membaca daftar sheet dari file Excel.
    
    Args:
        file_path (str): Path ke file Excel
        
    Returns:
        list: Daftar nama sheet dalam file Excel
        
    Raises:
        Exception: Jika file tidak dapat dibaca
    """
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        sheets = workbook.sheetnames
        workbook.close()
        return sheets
    except Exception as e:
        raise Exception(f"Error membaca file Excel: {str(e)}")

def filter_relevant_sheets(sheets, app_type):
    """
    Filter sheet berdasarkan tipe aplikasi dan keyword yang relevan.
    
    Args:
        sheets (list): Daftar nama sheet
        app_type (str): Tipe aplikasi ('app1', 'app2', 'app3', 'app4')
        
    Returns:
        list: Daftar sheet yang relevan untuk aplikasi tersebut
    """
    # Keyword mapping untuk setiap aplikasi
    keywords_map = {
        'app1': ['daft', 'ict', 'master', 'allocation', 'user', 'aktiva'],
        'app2': ['opname', 'target', 'scan', 'terscan', 'tidak'],
        'app3': ['mat', 'pic', 'ruangan', 'asset', 'management'],
        'app4': ['recouncil', 'reconcile', 'master', 'daft', 'ict']
    }
    
    keywords = keywords_map.get(app_type, [])
    relevant_sheets = []
    
    for sheet in sheets:
        sheet_lower = sheet.lower()
        if any(keyword in sheet_lower for keyword in keywords):
            relevant_sheets.append(sheet)
    
    # Jika tidak ada sheet yang cocok, return semua sheet
    return relevant_sheets if relevant_sheets else sheets

def create_sheet_selection_dialog(parent, file_path, app_type, file_description, allow_multiple=True):
    """
    Membuat dialog untuk pemilihan sheet dari file Excel dengan checkbox.
    
    Args:
        parent: Parent window
        file_path (str): Path ke file Excel
        app_type (str): Tipe aplikasi
        file_description (str): Deskripsi file untuk user
        allow_multiple (bool): Jika True, menggunakan checkbox; False untuk listbox
        
    Returns:
        list: Daftar nama sheet yang dipilih user, atau None jika dibatalkan
    """
    try:
        # Baca daftar sheet
        all_sheets = get_excel_sheets(file_path)
        relevant_sheets = filter_relevant_sheets(all_sheets, app_type)
        
        # Buat dialog window
        dialog = ttk.Toplevel(parent)
        dialog.title(f"Pilih Sheet - {file_description}")
        dialog.transient(parent)
        dialog.grab_set()
        dialog.resizable(True, True)
        
        # Position dialog di tengah dengan memanjang dari atas ke bawah
        dialog.update_idletasks()
        screen_width = dialog.winfo_screenwidth()
        screen_height = dialog.winfo_screenheight()
        
        # Window width tetap 700, height memanjang hampir penuh layar
        window_width = 700
        window_height = screen_height - 100  # Sisakan 100px untuk taskbar
        
        # Position di tengah horizontal, mulai dari atas
        x = (screen_width // 2) - (window_width // 2)
        y = 50  # Mulai dari 50px dari atas
        
        dialog.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        selected_sheets = []
        
        # Main frame
        main_frame = ttk.Frame(dialog)
        main_frame.pack(fill=BOTH, expand=True, padx=20, pady=20)
        
        # Title
        title_label = ttk.Label(
            main_frame,
            text=f"Pilih Sheet untuk {file_description}",
            font=("Segoe UI", 14, "bold")
        )
        title_label.pack(pady=(0, 10))
        
        # Info label
        info_label = ttk.Label(
            main_frame,
            text=f"File: {os.path.basename(file_path)}\nCentang sheet yang ingin diproses:",
            font=("Segoe UI", 10)
        )
        info_label.pack(pady=(0, 15))
        
        # Sheet selection frame dengan scrollable area
        sheet_selection_frame = ttk.LabelFrame(
            main_frame,
            text="  📋 Daftar Sheet yang Tersedia  ",
            padding=15
        )
        sheet_selection_frame.pack(fill=BOTH, expand=True, pady=(0, 15))
        
        # Create canvas dan scrollbar untuk scrolling
        canvas = tk.Canvas(sheet_selection_frame, bg='#2b2b2b', highlightthickness=0)
        scrollbar = ttk.Scrollbar(sheet_selection_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        # Configure scrolling
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Pack canvas dan scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Dictionary untuk menyimpan checkbox variables
        checkbox_vars = {}
        
        # Create checkboxes untuk setiap sheet
        for i, sheet in enumerate(all_sheets):
            # Frame untuk setiap checkbox
            checkbox_frame = ttk.Frame(scrollable_frame)
            checkbox_frame.pack(fill=X, pady=3, padx=10)
            
            # Variable untuk checkbox
            var = tk.BooleanVar()
            checkbox_vars[sheet] = var
            
            # Pre-select recommended sheets
            if sheet in relevant_sheets:
                var.set(True)
            
            # Checkbox dengan styling yang lebih besar
            checkbox = ttk.Checkbutton(
                checkbox_frame,
                text=f"📄 {sheet}",
                variable=var,
                bootstyle="primary-round-toggle"
            )
            checkbox.pack(side=LEFT, anchor=W, pady=2)
            
            # Indicator untuk recommended sheets
            if sheet in relevant_sheets:
                rec_label = ttk.Label(
                    checkbox_frame,
                    text="⭐ DEFAULT",
                    font=("Segoe UI", 9),
                    bootstyle="success"
                )
                rec_label.pack(side=RIGHT, anchor=E)
        
        # Mouse wheel scrolling support with error handling
        def _on_mousewheel(event):
            try:
                # Check if canvas still exists and is valid
                if canvas.winfo_exists():
                    canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            except (tk.TclError, AttributeError):
                # Canvas has been destroyed or is no longer valid, ignore the event
                pass
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        # MAIN ACTION BUTTONS - Pindah ke atas untuk visibility
        main_button_frame = ttk.Frame(main_frame)
        main_button_frame.pack(fill=X, pady=(0, 20))
        
        # Add background frame untuk highlight
        button_bg_frame = ttk.LabelFrame(
            main_button_frame,
            text="  🎯 Aksi Utama  ",
            bootstyle="primary",
            padding=15
        )
        button_bg_frame.pack(fill=X)
        
        button_container = ttk.Frame(button_bg_frame)
        button_container.pack(fill=X)
        
        def on_ok():
            nonlocal selected_sheets
            # Collect selected sheets from checkboxes
            selected_sheets = [sheet for sheet, var in checkbox_vars.items() if var.get()]
            if not selected_sheets:
                messagebox.showwarning("Peringatan", "Pilih minimal satu sheet!")
                return
            dialog.destroy()
        
        def on_cancel():
            nonlocal selected_sheets
            selected_sheets = None
            dialog.destroy()
        
        # Main action buttons - prominent dan di atas
        ok_btn = ttk.Button(
            button_container, 
            text="✅ OK - Proses Sheet Terpilih", 
            command=on_ok, 
            bootstyle="success",
            width=30
        )
        ok_btn.pack(side=RIGHT, padx=(10, 0), ipady=8)
        
        cancel_btn = ttk.Button(
            button_container, 
            text="❌ Cancel", 
            command=on_cancel, 
            bootstyle="danger",
            width=15
        )
        cancel_btn.pack(side=RIGHT, ipady=8)
        
        # Keyboard shortcuts
        dialog.bind('<Return>', lambda e: on_ok())
        dialog.bind('<Escape>', lambda e: on_cancel())
        
        # Focus pada OK button
        ok_btn.focus_set()
        
        # Helper buttons untuk select/deselect all
        helper_frame = ttk.Frame(main_frame)
        helper_frame.pack(fill=X, pady=(0, 15))
        
        helper_bg_frame = ttk.LabelFrame(
            helper_frame,
            text="  🛠️ Bantuan Pemilihan  ",
            bootstyle="secondary",
            padding=10
        )
        helper_bg_frame.pack(fill=X)
        
        def select_all():
            for var in checkbox_vars.values():
                var.set(True)
        
        def deselect_all():
            for var in checkbox_vars.values():
                var.set(False)
        
        def select_recommended():
            for sheet, var in checkbox_vars.items():
                var.set(sheet in relevant_sheets)
        
        # Helper buttons styling
        ttk.Button(
            helper_bg_frame, 
            text="✅ Pilih Semua", 
            command=select_all, 
            bootstyle="info-outline",
            width=15
        ).pack(side=LEFT, padx=(0, 10))
        
        ttk.Button(
            helper_bg_frame, 
            text="❌ Hapus Semua", 
            command=deselect_all, 
            bootstyle="warning-outline",
            width=15
        ).pack(side=LEFT, padx=(0, 10))
        
        if relevant_sheets:
            ttk.Button(
                helper_bg_frame, 
                text="⭐ Pilih Default", 
                command=select_recommended, 
                bootstyle="success-outline",
                width=15
            ).pack(side=LEFT)
        
        # Info tentang recommended sheets - MOVED DOWN
        if relevant_sheets:
            rec_info = ttk.Label(
                main_frame,
                text=f"Sheet Default: {', '.join(relevant_sheets[:5])}{'...' if len(relevant_sheets) > 5 else ''}",
                font=("Segoe UI", 9),
                bootstyle="info"
            )
            rec_info.pack(pady=(0, 15))
        
        # Wait for dialog to close
        dialog.wait_window()
        
        return selected_sheets
        
    except Exception as e:
        messagebox.showerror("Error", f"Error membaca file Excel:\n{str(e)}")
        return None

def create_multiple_sheet_selection_dialog(parent, file_path, app_type, file_description):
    """
    Convenience function untuk multiple sheet selection.
    
    Args:
        parent: Parent window
        file_path (str): Path ke file Excel
        app_type (str): Tipe aplikasi
        file_description (str): Deskripsi file untuk user
        
    Returns:
        list: Daftar nama sheet yang dipilih user, atau None jika dibatalkan
    """
    return create_sheet_selection_dialog(parent, file_path, app_type, file_description, allow_multiple=True)

#########################################################
# ENHANCED HELPER CLASSES
#########################################################

class RuleBuilder:
    """
    Rule Builder untuk mapping kolom dengan GUI dan JSON storage.
    """
    
    def __init__(self):
        self.rules = self.load_rules()
    
    def load_rules(self):
        """Load rules dari JSON file."""
        if os.path.exists(RULES_FILE):
            try:
                with open(RULES_FILE, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception as e:
                print(f"Error loading rules: {e}")
        
        # Default rules
        return {
            "column_mappings": {
                "NO ERS": "ASSET ORACLE",
                "NO ASSET": "ASSET ORACLE", 
                "ASSET ORACLE": "ASSET ORACLE",
                "NO PO": "ASSET ORACLE",
                "BARCODE": "NO BARCODE",
                "NO BARCODE": "NO BARCODE",
                "KODE": "NO BARCODE",
                "LOKASI": "LOKASI",
                "LOCATION": "LOKASI",
                "JENIS HARTA": "JENIS HARTA",
                "TYPE": "JENIS HARTA",
                "KONDISI": "KONDISI",
                "CONDITION": "KONDISI",
                "STATUS": "KONDISI"
            },
            "date_patterns": [
                r"TAHUN",
                r"YEAR", 
                r"PERO/AN\.1",
                r"BULAN",
                r"MONTH",
                r"PERO/AN"
            ]
        }
    
    def save_rules(self):
        """Save rules ke JSON file."""
        try:
            with open(RULES_FILE, 'w', encoding='utf-8') as f:
                json.dump(self.rules, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"Error saving rules: {e}")
    
    def add_mapping_rule(self, source_pattern, target_column):
        """Tambah rule mapping kolom."""
        self.rules["column_mappings"][source_pattern] = target_column
        self.save_rules()
    
    def get_column_mapping(self, header_row):
        """Apply rules untuk mendapatkan column mapping."""
        mapping = {}
        
        # Handle both Series and list inputs
        if hasattr(header_row, 'iloc'):
            header_list = [str(col).strip().upper() for col in header_row]
            original_headers = header_row
        else:
            header_list = [str(col).strip().upper() for col in header_row]
            original_headers = header_row
        
        for i, header_val in enumerate(header_list):
            for pattern, target in self.rules["column_mappings"].items():
                if pattern.upper() in header_val:
                    if hasattr(original_headers, 'iloc'):
                        mapping[target] = original_headers.iloc[i]
                    else:
                        mapping[target] = original_headers[i]
                    break
        
        return mapping

class ConflictResolver:
    """
    Conflict Resolver untuk review duplikat data secara visual.
    """
    
    def __init__(self, parent):
        self.parent = parent
        self.conflicts = []
        self.resolutions = {}
    
    def detect_conflicts(self, df_combined):
        """Deteksi konflik data berdasarkan barcode."""
        conflicts = []
        
        # Group by barcode untuk cari duplikat
        barcode_groups = df_combined.groupby('NO BARCODE')
        
        for barcode, group in barcode_groups:
            if len(group) > 1 and barcode and str(barcode).strip() not in ('', 'nan'):
                # Ada konflik untuk barcode ini
                conflict_data = {
                    'barcode': barcode,
                    'records': group.to_dict('records'),
                    'sources': group['__source'].tolist() if '__source' in group.columns else ['Unknown'] * len(group)
                }
                conflicts.append(conflict_data)
        
        self.conflicts = conflicts
        return len(conflicts)
    
    def show_conflict_dialog(self):
        """Tampilkan dialog untuk resolve conflicts."""
        if not self.conflicts:
            return {}
        
        # Create conflict resolution window
        conflict_window = ttk.Toplevel(self.parent)
        conflict_window.title("🔍 Conflict Resolver - Review Data Conflicts")
        conflict_window.geometry("900x600")
        
        # Main frame
        main_frame = ttk.Frame(conflict_window)
        main_frame.pack(fill=BOTH, expand=True, padx=10, pady=10)
        
        # Title
        title_label = ttk.Label(main_frame, text=f"Found {len(self.conflicts)} barcode conflicts", 
                               font=("Segoe UI", 14, "bold"))
        title_label.pack(pady=(0, 10))
        
        # Scrollable frame for conflicts
        canvas = tk.Canvas(main_frame)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Add conflicts to scrollable frame
        for i, conflict in enumerate(self.conflicts[:10]):  # Limit to first 10 conflicts
            self.create_conflict_item(scrollable_frame, conflict, i)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Buttons
        button_frame = ttk.Frame(conflict_window)
        button_frame.pack(fill=X, pady=(10, 0))
        
        ttk.Button(button_frame, text="Auto-Resolve (Use INV Priority)", 
                  command=lambda: self.auto_resolve_conflicts()).pack(side=LEFT, padx=(0, 5))
        ttk.Button(button_frame, text="Apply Selections", 
                  command=conflict_window.destroy).pack(side=LEFT, padx=(0, 5))
        ttk.Button(button_frame, text="Cancel", 
                  command=lambda: [self.resolutions.clear(), conflict_window.destroy()]).pack(side=RIGHT)
        
        # Center window
        conflict_window.transient(self.parent)
        conflict_window.grab_set()
        conflict_window.wait_window()
        
        return self.resolutions
    
    def create_conflict_item(self, parent, conflict, index):
        """Create UI item untuk satu conflict."""
        # Conflict frame
        conflict_frame = ttk.LabelFrame(parent, text=f"Barcode: {conflict['barcode']}", padding=10)
        conflict_frame.pack(fill=X, pady=5)
        
        # Variable untuk selection
        selection_var = tk.StringVar()
        self.resolutions[conflict['barcode']] = selection_var
        
        # Records
        for i, (record, source) in enumerate(zip(conflict['records'], conflict['sources'])):
            record_frame = ttk.Frame(conflict_frame)
            record_frame.pack(fill=X, pady=2)
            
            # Radio button
            radio = ttk.Radiobutton(record_frame, text=f"Source: {source}", 
                                   variable=selection_var, value=str(i))
            radio.pack(side=LEFT)
            
            # Record details
            details = f"Asset: {record.get('ASSET ORACLE', 'N/A')} | Lokasi: {record.get('LOKASI', 'N/A')} | Kondisi: {record.get('KONDISI', 'N/A')}"
            ttk.Label(record_frame, text=details, font=("Segoe UI", 9)).pack(side=LEFT, padx=(10, 0))
        
        # Set default selection (prioritize INV)
        if 'INV' in conflict['sources']:
            selection_var.set(str(conflict['sources'].index('INV')))
        else:
            selection_var.set("0")
    
    def auto_resolve_conflicts(self):
        """Auto-resolve conflicts menggunakan priority: INV > EXA > MASTER."""
        priority = ['INV', 'EXA', 'MASTER']
        
        for conflict in self.conflicts:
            sources = conflict['sources']
            best_index = 0
            best_priority = 999
            
            for i, source in enumerate(sources):
                if source in priority:
                    source_priority = priority.index(source)
                    if source_priority < best_priority:
                        best_priority = source_priority
                        best_index = i
            
            if conflict['barcode'] in self.resolutions:
                self.resolutions[conflict['barcode']].set(str(best_index))

class DateNormalizer:
    """
    Enhanced date normalizer untuk berbagai format tanggal.
    """
    
    @staticmethod
    def normalize_date(date_value):
        """
        Normalize berbagai format tanggal ke datetime object.
        
        Supports:
        - 'Jan-24', 'Feb-2024'
        - '2024/01', '01/2024'
        - '01.2024', '2024.01'
        - Excel serial dates
        - Standard datetime formats
        """
        if pd.isna(date_value) or not date_value:
            return pd.NaT
        
        date_str = str(date_value).strip()
        
        # Handle Excel serial dates (numbers)
        try:
            if date_str.replace('.', '').isdigit():
                serial_date = float(date_str)
                if 25000 < serial_date < 50000:  # Reasonable Excel date range
                    # Excel epoch starts at 1900-01-01, but has a leap year bug
                    excel_epoch = datetime(1899, 12, 30)
                    return excel_epoch + timedelta(days=serial_date)
        except (ValueError, OverflowError):
            pass
        
        # Month-Year patterns
        month_patterns = [
            (r'^([A-Za-z]{3})-?(\d{2,4})$', '%b-%Y'),  # Jan-24, Jan-2024
            (r'^(\d{1,2})[/](\d{4})$', '%m/%Y'),       # 01/2024
            (r'^(\d{1,2})[.](\d{4})$', '%m.%Y'),       # 01.2024
            (r'^(\d{1,2})[-](\d{4})$', '%m-%Y'),       # 01-2024
            (r'^(\d{4})[/](\d{1,2})$', '%Y/%m'),       # 2024/01
            (r'^(\d{4})[.](\d{1,2})$', '%Y.%m'),       # 2024.01
            (r'^(\d{4})[-](\d{1,2})$', '%Y-%m'),       # 2024-01
        ]
        
        for pattern, format_str in month_patterns:
            match = re.match(pattern, date_str)
            if match:
                try:
                    if format_str in ['%b-%Y']:
                        # Handle abbreviated month names
                        month_str, year_str = match.groups()
                        if len(year_str) == 2:
                            year_str = '20' + year_str
                        date_obj = datetime.strptime(f"{month_str}-{year_str}", format_str)
                        return pd.Timestamp(date_obj)
                    else:
                        return pd.to_datetime(date_str, format=format_str, errors='coerce')
                except ValueError:
                    continue
        
        # Standard datetime parsing
        try:
            return pd.to_datetime(date_str, errors='coerce')
        except:
            return pd.NaT

class DataDictionary:
    """
    Kamus data untuk standardisasi LOKASI dan JENIS HARTA.
    """
    
    def __init__(self):
        self.location_dict = self.load_location_dict()
        self.jenis_harta_dict = self.load_jenis_harta_dict()
    
    def load_location_dict(self):
        """Load location dictionary dari JSON."""
        if os.path.exists(LOCATION_DICT_FILE):
            try:
                with open(LOCATION_DICT_FILE, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception as e:
                print(f"Error loading location dict: {e}")
        
        # Default location mappings
        return {
            "aliases": {
                "KERAWANG": "KARAWANG",
                "KRW": "KARAWANG", 
                "KARAWANG": "KARAWANG",
                "JTB": "JATIBARU",
                "JATIBARU": "JATIBARU",
                "SMG": "SEMARANG",
                "SEMARANG": "SEMARANG",
                "JKT": "JAKARTA",
                "JAKARTA": "JAKARTA",
                "PJN": "PULOGADUNG",
                "PULOGADUNG": "PULOGADUNG"
            },
            "standard_values": [
                "KARAWANG", "JATIBARU", "SEMARANG", "JAKARTA", "PULOGADUNG"
            ]
        }
    
    def load_jenis_harta_dict(self):
        """Load jenis harta dictionary dari JSON."""
        if os.path.exists(JENIS_HARTA_DICT_FILE):
            try:
                with open(JENIS_HARTA_DICT_FILE, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception as e:
                print(f"Error loading jenis harta dict: {e}")
        
        # Default jenis harta mappings
        return {
            "aliases": {
                "PC": "PERSONAL COMPUTER",
                "DESKTOP": "PERSONAL COMPUTER",
                "COMPUTER": "PERSONAL COMPUTER",
                "LAPTOP": "NOTEBOOK",
                "NOTEBOOK": "NOTEBOOK",
                "MONITOR": "MONITOR",
                "LCD": "MONITOR",
                "LED": "MONITOR",
                "PRINTER": "PRINTER",
                "SCANNER": "SCANNER",
                "UPS": "UPS",
                "SERVER": "SERVER",
                "SWITCH": "NETWORK EQUIPMENT",
                "ROUTER": "NETWORK EQUIPMENT",
                "PROJECTOR": "PROJECTOR"
            },
            "standard_values": [
                "PERSONAL COMPUTER", "NOTEBOOK", "MONITOR", "PRINTER", 
                "SCANNER", "UPS", "SERVER", "NETWORK EQUIPMENT", "PROJECTOR"
            ]
        }
    
    def normalize_location(self, location_value):
        """Normalize location value menggunakan dictionary."""
        if pd.isna(location_value) or not location_value:
            return ""
        
        location_str = str(location_value).strip().upper()
        
        # Check aliases
        for alias, standard in self.location_dict["aliases"].items():
            if alias.upper() in location_str:
                return standard
        
        return location_str
    
    def normalize_jenis_harta(self, jenis_value):
        """Normalize jenis harta value menggunakan dictionary.""" 
        if pd.isna(jenis_value) or not jenis_value:
            return ""
        
        jenis_str = str(jenis_value).strip().upper()
        
        # Check aliases
        for alias, standard in self.jenis_harta_dict["aliases"].items():
            if alias.upper() in jenis_str:
                return standard
        
        return jenis_str
    
    def save_dictionaries(self):
        """Save dictionaries ke JSON files."""
        try:
            with open(LOCATION_DICT_FILE, 'w', encoding='utf-8') as f:
                json.dump(self.location_dict, f, indent=2, ensure_ascii=False)
            
            with open(JENIS_HARTA_DICT_FILE, 'w', encoding='utf-8') as f:
                json.dump(self.jenis_harta_dict, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"Error saving dictionaries: {e}")

class CacheManager:
    """
    Cache manager untuk menyimpan hasil baca sheet ke format parquet.
    """
    
    @staticmethod
    def get_file_hash(file_path):
        """Generate hash untuk file berdasarkan path dan modification time."""
        try:
            stat = os.stat(file_path)
            content = f"{file_path}_{stat.st_mtime}_{stat.st_size}"
            return hashlib.md5(content.encode()).hexdigest()
        except Exception:
            return None
    
    @staticmethod
    def get_cache_path(file_path, sheet_name):
        """Get cache file path untuk file dan sheet tertentu."""
        file_hash = CacheManager.get_file_hash(file_path)
        if not file_hash:
            return None
        
        safe_sheet_name = re.sub(r'[^\w\-_\.]', '_', sheet_name)
        cache_filename = f"{file_hash}_{safe_sheet_name}.parquet"
        return os.path.join(CACHE_DIR, cache_filename)
    
    @staticmethod
    def load_from_cache(file_path, sheet_name):
        """Load DataFrame dari cache jika tersedia dan valid."""
        if not PARQUET_AVAILABLE:
            return None
        
        cache_path = CacheManager.get_cache_path(file_path, sheet_name)
        if not cache_path or not os.path.exists(cache_path):
            return None
        
        try:
            return pd.read_parquet(cache_path)
        except Exception as e:
            print(f"Error loading cache: {e}")
            return None
    
    @staticmethod
    def save_to_cache(df, file_path, sheet_name):
        """Save DataFrame ke cache dalam format parquet."""
        if not PARQUET_AVAILABLE:
            return False
        
        cache_path = CacheManager.get_cache_path(file_path, sheet_name)
        if not cache_path:
            return False
        
        try:
            df.to_parquet(cache_path, index=False)
            return True
        except Exception as e:
            print(f"Error saving cache: {e}")
            return False
    
    @staticmethod
    def clear_cache():
        """Clear semua cache files."""
        try:
            for filename in os.listdir(CACHE_DIR):
                if filename.endswith('.parquet'):
                    os.remove(os.path.join(CACHE_DIR, filename))
            return True
        except Exception as e:
            print(f"Error clearing cache: {e}")
            return False

class CheckpointManager:
    """
    Checkpoint manager untuk auto-save dan recovery.
    """
    
    @staticmethod
    def save_checkpoint(stage, data):
        """Save checkpoint untuk stage tertentu."""
        checkpoint_data = {
            'timestamp': datetime.now().isoformat(),
            'stage': stage,
            'data': data
        }
        
        try:
            with open(CHECKPOINT_FILE, 'w', encoding='utf-8') as f:
                json.dump(checkpoint_data, f, indent=2, ensure_ascii=False, default=str)
            return True
        except Exception as e:
            print(f"Error saving checkpoint: {e}")
            return False
    
    @staticmethod
    def load_checkpoint():
        """Load checkpoint terakhir jika ada."""
        if not os.path.exists(CHECKPOINT_FILE):
            return None
        
        try:
            with open(CHECKPOINT_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"Error loading checkpoint: {e}")
            return None
    
    @staticmethod
    def clear_checkpoint():
        """Clear checkpoint file."""
        try:
            if os.path.exists(CHECKPOINT_FILE):
                os.remove(CHECKPOINT_FILE)
            return True
        except Exception as e:
            print(f"Error clearing checkpoint: {e}")
            return False

#########################################################
# MODERN BASE WINDOW CLASS (ttkbootstrap)
#########################################################
class ModernBaseWindow:
    """
    Base class untuk semua window aplikasi dengan ttkbootstrap modern theme.
    
    Menyediakan:
    - Dark theme configuration
    - Consistent window styling
    - Card-based layout helpers
    - Status bar dan navigation
    - Error handling utilities
    
    Usage:
        class MyApp(ModernBaseWindow):
            def __init__(self, root, title):
                super().__init__(root, title, "800x600")
                # Add your UI components here
    """
    
    def __init__(self, root, title, geometry="800x600"):
        """
        Initialize modern base window.
        
        Args:
            root: Parent window (ttk.Toplevel)
            title (str): Window title
            geometry (str): Window size in format "WIDTHxHEIGHT"
        """
        self.root = root
        # Always use the provided window (which is now a Toplevel)
        self.window = root
        
        self.window.title(title)
        
        # Configure window to prevent white flash
        self.window.configure(bg='#2b2b2b')  # Dark background immediately
        
        # Set initial geometry kemudian adjust untuk full height
        self.window.geometry(geometry)
        
        # Center window dengan full height
        self.center_window_full_height()
        
        # Create main container with modern styling
        self.create_main_container(title)
        
        # Status bar
        self.create_status_bar()
    
    def center_window(self):
        self.window.update_idletasks()
        screen_width = self.window.winfo_screenwidth()
        screen_height = self.window.winfo_screenheight()
        
        # Get current window width, tapi height memanjang dari atas ke bawah
        width = self.window.winfo_width()
        height = screen_height - 100  # Sisakan 100px untuk taskbar
        
        # Position di tengah horizontal, mulai dari atas
        x = (screen_width // 2) - (width // 2)
        y = 50  # Mulai dari 50px dari atas
        
        self.window.geometry(f"{width}x{height}+{x}+{y}")
    
    def center_window_full_height(self):
        """Position window di tengah horizontal dengan height memanjang dari atas ke bawah"""
        self.window.update_idletasks()
        screen_width = self.window.winfo_screenwidth()
        screen_height = self.window.winfo_screenheight()
        
        # Get current window width, tapi height memanjang dari atas ke bawah
        width = self.window.winfo_width()
        height = screen_height - 100  # Sisakan 100px untuk taskbar
        
        # Position di tengah horizontal, mulai dari atas
        x = (screen_width // 2) - (width // 2)
        y = 50  # Mulai dari 50px dari atas
        
        self.window.geometry(f"{width}x{height}+{x}+{y}")
    
    def create_main_container(self, title):
        # Main container with dark theme
        self.main_frame = ttk.Frame(self.window)
        self.main_frame.pack(fill=BOTH, expand=True, padx=20, pady=20)
        
        # Simple header with title only
        self.header_frame = ttk.Frame(self.main_frame)
        self.header_frame.pack(fill=X, pady=(0, 20))
        
        # Title with consistent styling
        self.title_label = ttk.Label(
            self.header_frame,
            text=title,
            font=("Segoe UI", 18, "bold"),
            bootstyle="light"
        )
        self.title_label.pack()
    
    def create_status_bar(self):
        self.status_frame = ttk.Frame(self.window)
        self.status_frame.pack(fill=X, side=BOTTOM, padx=15, pady=(0, 15))
        
        self.status_label = ttk.Label(
            self.status_frame,
            text="Ready",
            font=FONT_CONFIG["text"],
            bootstyle="secondary"
        )
        self.status_label.pack(side=LEFT)
    
    def create_card_frame(self, parent, title, subtitle=""):
        """Create a modern card-style frame with dark theme"""
        card = ttk.LabelFrame(
            parent,
            text=f"  {title}  ",
            bootstyle="secondary",
            padding=15
        )
        
        if subtitle:
            subtitle_lbl = ttk.Label(
                card,
                text=subtitle,
                font=FONT_CONFIG["text"],
                bootstyle="secondary"
            )
            subtitle_lbl.pack(anchor=W, pady=(0, 10))
        
        return card
    
    def add_back_button(self, callback):
        back_frame = ttk.Frame(self.main_frame)
        back_frame.pack(fill=X, pady=(0, 20))
        
        # Create a more prominent back button
        back_btn = ttk.Button(
            back_frame,
            text="🏠 Kembali ke Menu Utama",
            bootstyle="info",
            command=lambda: callback(self.window),
            width=25
        )
        back_btn.pack(side=LEFT, ipady=5)
        
        # Add separator line
        ttk.Separator(back_frame, orient=HORIZONTAL).pack(fill=X, pady=(15, 0))
        
        # Add tooltip
        ToolTip(back_btn, text="Klik untuk kembali ke menu utama aplikasi")

#########################################################
# LEGACY BASE WINDOW CLASS (untuk backward compatibility)
#########################################################
class BaseAppWindow:
    def __init__(self, root, title, geometry="500x550"):
        self.root = root
        self.root.overrideredirect(True) # Menghilangkan title bar default
        self.root.geometry(geometry)
        self.root.config(bg=STYLE["background"], bd=2, relief=tk.SOLID)

        self._offsetx = 0
        self._offsety = 0

        # Title Bar Custom
        self.title_bar = tk.Frame(root, bg=STYLE["widget_bg"], relief='raised', bd=0)
        self.title_bar.pack(fill=tk.X)

        self.lbl_title = tk.Label(self.title_bar, text=title, bg=STYLE["widget_bg"], fg=STYLE["text"], font=STYLE["font_title"])
        self.lbl_title.pack(side=tk.LEFT, padx=10)

        self.close_button = tk.Button(self.title_bar, text='✕', bg=STYLE["widget_bg"], fg=STYLE["text"], command=self.root.destroy, relief=tk.FLAT, font=("Segoe UI", 10, "bold"), width=4)
        self.close_button.pack(side=tk.RIGHT)
        
        self.minimize_button = tk.Button(self.title_bar, text='—', bg=STYLE["widget_bg"], fg=STYLE["text"], command=self.root.iconify, relief=tk.FLAT, font=("Segoe UI", 10, "bold"), width=4)
        self.minimize_button.pack(side=tk.RIGHT)

        # Bind event untuk drag window
        self.title_bar.bind('<Button-1>', self.clickwin)
        self.title_bar.bind('<B1-Motion>', self.dragwin)
        self.lbl_title.bind('<Button-1>', self.clickwin)
        self.lbl_title.bind('<B1-Motion>', self.dragwin)
        
        # Main content frame
        self.main_frame = tk.Frame(root, bg=STYLE["background"], padx=20, pady=20)
        self.main_frame.pack(expand=True, fill=tk.BOTH)

        # Center window on first open
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')

    def clickwin(self, event):
        self._offsetx = event.x
        self._offsety = event.y

    def dragwin(self, event):
        x = self.root.winfo_pointerx() - self._offsetx
        y = self.root.winfo_pointery() - self._offsety
        self.root.geometry(f"+{x}+{y}")
        
    def add_back_button(self, command):
        back_button = tk.Button(self.title_bar, text='← Kembali', bg=STYLE["widget_bg"], fg=STYLE["text"], command=command, relief=tk.FLAT, font=("Segoe UI", 10, "bold"))
        back_button.pack(side=tk.RIGHT, padx=5)


#########################################################
# APLIKASI 1: DAFT ICT FILE PROCESSOR
#########################################################
class DaftProcessorApp(ModernBaseWindow):
    """
    📊 DAFT ICT FILE PROCESSOR - Versi Terbaru
    
    FUNGSI:
    Menggabungkan 3 file Excel menjadi 1 file Kamus Data Master:
    - Kamus Data Master (data existing)
    - File EXA (asset existing)
    - File INV (asset baru)
    
    FITUR UTAMA:
    1. Pilih sheet dari setiap file
    2. Filter data berdasarkan BAT (ICT/ENG/HRGA/BAT/BLANKS)
    3. Semua lokasi diproses (Kerawang/Karawang/Jatibaru/Semarang)
    4. Barcode duplikat otomatis dihapus (ambil data terbaru)
    5. Barcode range otomatis dipecah (123456-123460 jadi 5 baris)
    
    CARA PAKAI:
    1. Pilih file Kamus Data Master → pilih sheet
    2. Pilih file EXA → pilih sheet → set filter BAT
    3. Pilih file INV → pilih sheet → set filter BAT
    4. Klik "PROSES & BUAT KAMUS"
    5. File hasil otomatis tersimpan
    
    HASIL:
    - File Excel sheet 'Recouncil'
    - Kolom: NO BARCODE, ASSET ORACLE, LOKASI, JENIS HARTA, KONDISI, Tahun Perolehan
    - Format rapi dengan zebra banding
    
    PERBAIKAN TERBARU (V2.1):
    - BAT Filter benar-benar berfungsi
    - Tahun perolehan dari kolom K (EXA) dan P (INV)
    - Asset Oracle dari kolom E
    - Proses lebih cepat dan akurat
    """
    
    def __init__(self, root, on_close):
        super().__init__(root, "📊 DAFT ICT FILE PROCESSOR", "800x750")
        self.add_back_button(on_close)

        # File paths dan sheet names
        self.master_file = ""
        self.master_sheet = ""
        self.exa_file = ""
        self.exa_sheets = []  # List of selected sheets for EXA
        self.inv_file = ""
        self.inv_sheets = []  # List of selected sheets for INV

        # BAT filters per file (akan muncul setelah file dipilih)
        self.bat_opts_exa = {name: tk.BooleanVar(value=(name in ["ICT", "BLANKS"])) for name in ["ICT","ENG","HRGA","BAT","BLANKS"]}
        self.bat_opts_inv = {name: tk.BooleanVar(value=(name in ["ICT", "BLANKS"])) for name in ["ICT","ENG","HRGA","BAT","BLANKS"]}
        self.frm_filters_exa = None
        self.frm_filters_inv = None

        self.processing_result = None
        self.log_queue = Queue()

        # === BASELINE & DIFF FEATURES ===
        self.promote_baseline = tk.BooleanVar(value=False)
        self.error_log = []  # Store errors/warnings during processing
        
        # === ENHANCED FEATURES ===
        self.rule_builder = RuleBuilder()
        self.conflict_resolver = ConflictResolver(self.root)
        self.date_normalizer = DateNormalizer()
        self.data_dictionary = DataDictionary()
        self.cache_manager = CacheManager()
        self.checkpoint_manager = CheckpointManager()
        
        # Progress tracking
        self.current_stage = ""
        self.stage_progress = {}
        self.total_stages = 5  # Master, EXA, INV, Merge, Save
        
        # Performance tracking
        self.start_time = None
        self.stage_times = {}

        # === MODERN UI ===
        self.create_modern_ui()

    def create_modern_ui(self):
        # Description card
        desc_card = self.create_card_frame(
            self.main_frame, 
            "📋 Process Description",
            "Combine Kamus Data Master with EXA & INV files to create updated master data. All locations included (KERAWANG/KARAWANG/JATIBARU/SEMARANG)."
        )
        desc_card.pack(fill=X, pady=(0, 20))
        
        # File selection section
        file_section = self.create_card_frame(self.main_frame, "📁 File Selection")
        file_section.pack(fill=X, pady=(0, 20))
        
        # Master file selection
        master_frame = ttk.Frame(file_section)
        master_frame.pack(fill=X, pady=(0, 15))
        
        btn_master = ttk.Button(
            master_frame, 
            text="📊 1. Pilih File Kamus Data Master",
            bootstyle="secondary",
            command=self.select_master_file
        )
        btn_master.pack(fill=X, pady=(0, 5))
        
        self.lbl_master_status = ttk.Label(
            master_frame, 
            text="❌ Belum ada file yang dipilih",
            font=FONT_CONFIG["text"],
            bootstyle="danger"
        )
        self.lbl_master_status.pack(anchor=W)
        
        # EXA file selection
        exa_frame = ttk.Frame(file_section)
        exa_frame.pack(fill=X, pady=(0, 15))
        
        btn_exa = ttk.Button(
            exa_frame, 
            text="📈 2. Pilih File EXA",
            bootstyle="secondary",
            command=self.select_exa_file
        )
        btn_exa.pack(fill=X, pady=(0, 5))
        
        self.lbl_exa_status = ttk.Label(
            exa_frame, 
            text="❌ Belum ada file yang dipilih",
            font=FONT_CONFIG["text"],
            bootstyle="danger"
        )
        self.lbl_exa_status.pack(anchor=W)
        
        # EXA filter container
        self.container_exa_filter = ttk.Frame(file_section)
        self.container_exa_filter.pack(fill=X, pady=(5, 0))
        
        # INV file selection
        inv_frame = ttk.Frame(file_section)
        inv_frame.pack(fill=X, pady=(15, 0))
        
        btn_inv = ttk.Button(
            inv_frame, 
            text="📋 3. Pilih File INV",
            bootstyle="secondary",
            command=self.select_inv_file
        )
        btn_inv.pack(fill=X, pady=(0, 5))
        
        self.lbl_inv_status = ttk.Label(
            inv_frame, 
            text="❌ Belum ada file yang dipilih",
            font=FONT_CONFIG["text"],
            bootstyle="danger"
        )
        self.lbl_inv_status.pack(anchor=W)
        
        # INV filter container
        self.container_inv_filter = ttk.Frame(file_section)
        self.container_inv_filter.pack(fill=X, pady=(5, 0))
        
        # Process button section with additional features
        process_section = ttk.Frame(self.main_frame)
        process_section.pack(fill=X, pady=(20, 0))
        
        # Baseline options section
        baseline_frame = ttk.Frame(process_section)
        baseline_frame.pack(fill=X, pady=(0, 10))
        
        # Promote baseline checkbox
        self.chk_promote_baseline = ttk.Checkbutton(
            baseline_frame,
            text="🔄 Gunakan hasil ini sebagai Master Baru (reset highlight)",
            variable=self.promote_baseline,
            bootstyle="info"
        )
        self.chk_promote_baseline.pack(anchor=W)
        
        # Tooltip for promote baseline
        ToolTip(self.chk_promote_baseline, 
                "Jika dicentang: Hasil akan disimpan dengan highlight + salinan bersih tanpa highlight untuk baseline berikutnya")
        
        # Main process button
        self.btn_process = ttk.Button(
            process_section, 
            text="🚀 PROSES & BUAT KAMUS (Updated)",
            bootstyle="success",
            command=self.start_processing,
            state=DISABLED
        )
        self.btn_process.pack(fill=X, pady=(0, 10), ipady=10)
        
        # Additional feature buttons
        feature_frame = ttk.Frame(process_section)
        feature_frame.pack(fill=X, pady=(5, 15))
        
        # Preview button
        self.btn_preview = ttk.Button(
            feature_frame,
            text="👁️ Preview Data",
            bootstyle="info-outline",
            command=self.preview_data,
            state=DISABLED,
            width=15
        )
        self.btn_preview.pack(side=LEFT, padx=(0, 5))
        
        # Backup button
        self.btn_backup = ttk.Button(
            feature_frame,
            text="💾 Backup Files",
            bootstyle="warning-outline",
            command=self.backup_files,
            state=DISABLED,
            width=15
        )
        self.btn_backup.pack(side=LEFT, padx=(0, 5))
        
        # Validate button
        self.btn_validate = ttk.Button(
            feature_frame,
            text="✅ Validate Data",
            bootstyle="secondary-outline",
            command=self.validate_data,
            state=DISABLED,
            width=15
        )
        self.btn_validate.pack(side=LEFT, padx=(0, 5))
        
        # Export settings button
        self.btn_export_settings = ttk.Button(
            feature_frame,
            text="📤 Export Settings",
            bootstyle="primary-outline",
            command=self.export_settings,
            width=18
        )
        self.btn_export_settings.pack(side=RIGHT)
        
        # Enhanced features section
        enhanced_section = self.create_card_frame(self.main_frame, "🚀 Enhanced Features")
        enhanced_section.pack(fill=X, pady=(20, 0))
        
        # Enhanced features buttons
        enhanced_frame = ttk.Frame(enhanced_section)
        enhanced_frame.pack(fill=X, pady=5)
        
        # Rule Builder button
        self.btn_rule_builder = ttk.Button(
            enhanced_frame,
            text="⚙️ Column Mapping Rules",
            bootstyle="info-outline",
            command=self.show_rule_builder,
            width=20
        )
        self.btn_rule_builder.pack(side=LEFT, padx=(0, 5))
        
        # Conflict Resolver toggle
        self.show_conflicts = tk.BooleanVar(value=False)
        self.chk_conflicts = ttk.Checkbutton(
            enhanced_frame,
            text="🔍 Show Conflict Resolver",
            variable=self.show_conflicts,
            bootstyle="info"
        )
        self.chk_conflicts.pack(side=LEFT, padx=(0, 5))
        
        # Cache management button
        self.btn_cache_mgmt = ttk.Button(
            enhanced_frame,
            text="💾 Cache Management",
            bootstyle="secondary-outline",
            command=self.show_cache_management,
            width=18
        )
        self.btn_cache_mgmt.pack(side=LEFT, padx=(0, 5))
        
        # Recovery button
        self.btn_recovery = ttk.Button(
            enhanced_frame,
            text="🔄 Recovery Options",
            bootstyle="warning-outline",
            command=self.show_recovery_options,
            width=18
        )
        self.btn_recovery.pack(side=RIGHT)
        
        # Add tooltips
        ToolTip(self.btn_process, text="Process files to create updated master data")
        ToolTip(self.btn_preview, text="Preview data before processing")
        ToolTip(self.btn_backup, text="Create backup of original files")
        ToolTip(self.btn_validate, text="Validate data integrity and format")
        ToolTip(self.btn_export_settings, text="Export current filter settings")
        
        # Enhanced features tooltips
        ToolTip(self.btn_rule_builder, text="Configure column mapping rules for automatic detection")
        ToolTip(self.chk_conflicts, text="Enable visual conflict resolution for duplicate barcodes")
        ToolTip(self.btn_cache_mgmt, text="Manage parquet cache for faster processing")
        ToolTip(self.btn_recovery, text="Recovery options and checkpoint management")

    # ---------- UI helpers ----------
    def button_style(self):
        return {"bg": STYLE["accent"], "fg": STYLE["text"], "font": STYLE["font_button"], "relief": tk.FLAT, "pady": 10}
    def label_style(self, type="default"):
        color = STYLE["text"] if type == "default" else (STYLE["success"] if type == "success" else STYLE["error"])
        return {"bg": STYLE["background"], "fg": color}

    def _render_bat_filter(self, container, title, opts_dict):
        for child in container.winfo_children():
            child.destroy()
        
        # Modern filter frame
        filter_frame = ttk.LabelFrame(
            container, 
            text=f"  🔍 {title}  ",
            bootstyle="secondary",
            padding=15
        )
        filter_frame.pack(fill=X, pady=(10, 5))
        
        # Checkboxes container
        checkbox_frame = ttk.Frame(filter_frame)
        checkbox_frame.pack(fill=X)
        
        # Create modern checkboxes
        for i, name in enumerate(["ICT", "ENG", "HRGA", "BAT", "BLANKS"]):
            cb = ttk.Checkbutton(
                checkbox_frame, 
                text=f"📁 {name}",
                variable=opts_dict[name],
                bootstyle="primary-round-toggle"
            )
            cb.grid(row=0, column=i, padx=10, pady=5, sticky=W)
            
            # Add tooltip
            ToolTip(cb, text=f"Include {name} data in processing")

    # ---------- File selectors ----------
    def select_master_file(self):
        """
        Pilih file Kamus Data Master dengan sheet selection dialog.
        
        File master berisi data referensi utama yang akan digabungkan dengan
        data EXA dan INV. User dapat memilih sheet yang sesuai dari file Excel.
        
        Template kolom yang dibutuhkan:
        - NO BARCODE: Nomor barcode asset
        - ASSET ORACLE: Nomor asset dari Oracle
        - LOKASI: Lokasi fisik asset
        - JENIS HARTA: Kategori/jenis asset
        - KONDISI: Status kondisi asset
        - Tahun Perolehan: Tanggal perolehan asset
        """
        file = filedialog.askopenfilename(
            title="Pilih file Kamus Data Master", 
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if file:
            # Show sheet selection dialog dengan checkbox
            selected_sheets = create_sheet_selection_dialog(
                self.window, 
                file, 
                'app1', 
                "File Kamus Data Master"
            )
            
            if selected_sheets:
                self.master_file = file
                # Untuk master file, ambil sheet pertama yang dipilih
                self.master_sheet = selected_sheets[0] if isinstance(selected_sheets, list) else selected_sheets
                sheet_text = f"{len(selected_sheets)} sheet dipilih, menggunakan: {self.master_sheet}" if isinstance(selected_sheets, list) and len(selected_sheets) > 1 else self.master_sheet
                self.lbl_master_status.config(
                    text=f"✅ {os.path.basename(file)} ({sheet_text})", 
                    bootstyle="success"
                )
            else:
                # User cancelled sheet selection
                return
        self.check_files_selected()

    def select_exa_file(self):
        """
        Pilih file EXA (Allocation Data) dengan sheet selection dialog.
        
        File EXA berisi data allocation user aktiva yang akan diproses.
        User dapat memilih multiple sheets untuk diproses sekaligus.
        
        Template kolom yang dibutuhkan:
        - NO BARCODE: Nomor barcode asset
        - ASSET ORACLE/NO ERS/NO ASSET: Nomor asset
        - LOKASI: Lokasi fisik asset
        - JENIS HARTA: Kategori/jenis asset
        - KONDISI: Status kondisi asset
        - BAT: Kategori BAT (ICT/ENG/HRGA/BAT)
        - PERO/AN & PERO/AN.1: Bulan dan tahun perolehan
        
        Fitur:
        - Multiple sheet selection: Pilih beberapa sheet sekaligus
        - BAT filtering: Filter data berdasarkan kategori BAT
        - Auto-expand barcodes: Range barcode otomatis di-expand
        """
        file = filedialog.askopenfilename(
            title="Pilih file EXA", 
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if file:
            # Show sheet selection dialog dengan checkbox
            selected_sheets = create_sheet_selection_dialog(
                self.window, 
                file, 
                'app1', 
                "File EXA (Allocation Data)"
            )
            
            if selected_sheets:
                self.exa_file = file
                # Store all selected sheets for EXA
                self.exa_sheets = selected_sheets if isinstance(selected_sheets, list) else [selected_sheets]
                sheet_text = f"{len(self.exa_sheets)} sheet dipilih: {', '.join(self.exa_sheets[:3])}{'...' if len(self.exa_sheets) > 3 else ''}"
                self.lbl_exa_status.config(
                    text=f"✅ {os.path.basename(file)} ({sheet_text})", 
                    bootstyle="success"
                )
                self._render_bat_filter(self.container_exa_filter, "Filter BAT untuk EXA", self.bat_opts_exa)
            else:
                return
        self.check_files_selected()

    def select_inv_file(self):
        """
        Pilih file INV (Inventory Data) dengan sheet selection dialog.
        
        File INV berisi data inventory aktiva yang akan diproses.
        User dapat memilih multiple sheets untuk diproses sekaligus.
        
        Template kolom yang dibutuhkan:
        - NO BARCODE: Nomor barcode asset
        - ASSET ORACLE/NO ERS/NO ASSET: Nomor asset
        - LOKASI: Lokasi fisik asset
        - JENIS HARTA: Kategori/jenis asset
        - KONDISI: Status kondisi asset
        - BAT: Kategori BAT (ICT/ENG/HRGA/BAT)
        - PERO/AN & PERO/AN.1: Bulan dan tahun perolehan
        
        Fitur:
        - Multiple sheet selection: Pilih beberapa sheet sekaligus
        - BAT filtering: Filter data berdasarkan kategori BAT
        - Auto-expand barcodes: Range barcode otomatis di-expand
        """
        file = filedialog.askopenfilename(
            title="Pilih file INV", 
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if file:
            # Show sheet selection dialog dengan checkbox
            selected_sheets = create_sheet_selection_dialog(
                self.window, 
                file, 
                'app1', 
                "File INV (Inventory Data)"
            )
            
            if selected_sheets:
                self.inv_file = file
                # Store all selected sheets for INV
                self.inv_sheets = selected_sheets if isinstance(selected_sheets, list) else [selected_sheets]
                sheet_text = f"{len(self.inv_sheets)} sheet dipilih: {', '.join(self.inv_sheets[:3])}{'...' if len(self.inv_sheets) > 3 else ''}"
                self.lbl_inv_status.config(
                    text=f"✅ {os.path.basename(file)} ({sheet_text})", 
                    bootstyle="success"
                )
                self._render_bat_filter(self.container_inv_filter, "Filter BAT untuk INV", self.bat_opts_inv)
            else:
                return
        self.check_files_selected()

    def check_files_selected(self):
        if self.master_file and self.exa_file and self.inv_file:
            self.btn_process.config(state=NORMAL)
            self.btn_process.config(bootstyle="success")
            # Enable additional feature buttons
            self.btn_preview.config(state=NORMAL)
            self.btn_backup.config(state=NORMAL)
            self.btn_validate.config(state=NORMAL)
        else:
            self.btn_process.config(state=DISABLED)
            self.btn_process.config(bootstyle="secondary-outline")
            # Disable additional feature buttons
            self.btn_preview.config(state=DISABLED)
            self.btn_backup.config(state=DISABLED)
            self.btn_validate.config(state=DISABLED)

    def preview_data(self):
        """Preview data from selected files before processing"""
        try:
            preview_window = tk.Toplevel(self.window)
            preview_window.title("👁️ Data Preview")
            preview_window.geometry("800x600")
            preview_window.transient(self.window)
            
            # Create notebook for tabs
            notebook = ttk.Notebook(preview_window)
            notebook.pack(fill=BOTH, expand=True, padx=10, pady=10)
            
            # Preview master file
            if self.master_file:
                master_frame = ttk.Frame(notebook)
                notebook.add(master_frame, text="📊 Master Data")
                
                df_master = pd.read_excel(self.master_file, sheet_name=self.master_sheet if self.master_sheet else 0)
                preview_text = scrolledtext.ScrolledText(master_frame, wrap=tk.WORD)
                preview_text.pack(fill=BOTH, expand=True, padx=5, pady=5)
                preview_text.insert(tk.END, f"Master File: {os.path.basename(self.master_file)}\n")
                preview_text.insert(tk.END, f"Rows: {len(df_master)}, Columns: {len(df_master.columns)}\n\n")
                preview_text.insert(tk.END, "Columns:\n" + "\n".join(df_master.columns.tolist()) + "\n\n")
                preview_text.insert(tk.END, "First 10 rows:\n" + df_master.head(10).to_string())
            
            messagebox.showinfo("Preview", "Data preview window opened!")
            
        except Exception as e:
            messagebox.showerror("Preview Error", f"Error previewing data:\n{str(e)}")

    def backup_files(self):
        """Create backup of original files"""
        try:
            backup_folder = filedialog.askdirectory(title="Select Backup Folder")
            if backup_folder:
                import shutil
                from datetime import datetime
                
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                backup_subfolder = os.path.join(backup_folder, f"ICT_Backup_{timestamp}")
                os.makedirs(backup_subfolder, exist_ok=True)
                
                files_to_backup = [
                    (self.master_file, "master"),
                    (self.exa_file, "exa"),
                    (self.inv_file, "inv")
                ]
                
                for file_path, file_type in files_to_backup:
                    if file_path:
                        backup_name = f"{file_type}_{timestamp}_{os.path.basename(file_path)}"
                        shutil.copy2(file_path, os.path.join(backup_subfolder, backup_name))
                
                messagebox.showinfo("Backup Complete", f"Files backed up to:\n{backup_subfolder}")
        except Exception as e:
            messagebox.showerror("Backup Error", f"Error creating backup:\n{str(e)}")

    def validate_data(self):
        """Validate data integrity and format"""
        try:
            validation_results = []
            
            # Validate master file
            if self.master_file:
                df_master = pd.read_excel(self.master_file, sheet_name=self.master_sheet if self.master_sheet else 0)
                required_cols = ['NO BARCODE', 'ASSET ORACLE', 'LOKASI', 'JENIS HARTA', 'KONDISI']
                missing_cols = [col for col in required_cols if col not in df_master.columns]
                
                if missing_cols:
                    validation_results.append(f"❌ Master file missing columns: {', '.join(missing_cols)}")
                else:
                    validation_results.append("✅ Master file structure is valid")
                
                # Check for empty barcodes
                empty_barcodes = df_master['NO BARCODE'].isna().sum()
                if empty_barcodes > 0:
                    validation_results.append(f"⚠️ Master file has {empty_barcodes} empty barcodes")
                else:
                    validation_results.append("✅ All barcodes in master file are present")
            
            # Show validation results
            result_text = "\n".join(validation_results)
            messagebox.showinfo("Validation Results", result_text)
            
        except Exception as e:
            messagebox.showerror("Validation Error", f"Error validating data:\n{str(e)}")

    def export_settings(self):
        """Export current filter settings to file"""
        try:
            settings_file = filedialog.asksaveasfilename(
                title="Export Settings",
                defaultextension=".json",
                filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
            )
            
            if settings_file:
                import json
                settings = {
                    "exa_filters": {name: var.get() for name, var in self.bat_opts_exa.items()},
                    "inv_filters": {name: var.get() for name, var in self.bat_opts_inv.items()},
                    "master_file": self.master_file,
                    "exa_file": self.exa_file,
                    "inv_file": self.inv_file,
                    "export_timestamp": str(pd.Timestamp.now())
                }
                
                with open(settings_file, 'w') as f:
                    json.dump(settings, f, indent=2)
                
                messagebox.showinfo("Export Complete", f"Settings exported to:\n{settings_file}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Error exporting settings:\n{str(e)}")

    # ---------- Loading / Log ----------
    def show_loading_screen(self):
        self.loading_window = tk.Toplevel(self.root)
        self.loading_window.overrideredirect(True)
        self.loading_window.geometry("760x360")
        self.loading_window.config(bg=STYLE["background"], bd=1, relief=tk.SOLID)

        self.root.update_idletasks()
        main_x = self.root.winfo_x()
        main_y = self.root.winfo_y()
        main_width = self.root.winfo_width()
        self.loading_window.geometry(f"+{main_x + main_width}+{main_y}")

        title_bar = tk.Frame(self.loading_window, bg=STYLE["widget_bg"], relief='raised', bd=0)
        title_bar.pack(fill=tk.X)

        lbl_title = tk.Label(title_bar, text="Memproses... (log)",
                             bg=STYLE["widget_bg"], fg=STYLE["text"], font=STYLE["font_title"])
        lbl_title.pack(side=tk.LEFT, padx=10)

        def clickwin(event):
            self._offsetx_log = event.x
            self._offsety_log = event.y
        def dragwin(event):
            x = self.loading_window.winfo_pointerx() - self._offsetx_log
            y = self.loading_window.winfo_pointery() - self._offsety_log
            self.loading_window.geometry(f"+{x}+{y}")
        title_bar.bind('<Button-1>', clickwin)
        title_bar.bind('<B1-Motion>', dragwin)
        lbl_title.bind('<Button-1>', clickwin)
        lbl_title.bind('<B1-Motion>', dragwin)

        self.log_text = scrolledtext.ScrolledText(self.loading_window, wrap=tk.WORD, bg=STYLE["widget_bg"],
                                                  fg=STYLE["text"], font=STYLE["font_main"])
        self.log_text.pack(padx=10, pady=10, expand=True, fill=tk.BOTH)
        self.log_text.insert(tk.END, "Inisialisasi...\n")
        self.log_text.config(state=tk.DISABLED)

    def close_loading_screen(self):
        if hasattr(self, 'loading_window') and self.loading_window.winfo_exists():
            self.loading_window.destroy()

    # ---------- Processor ----------
    def start_processing(self):
        self.output_file = filedialog.asksaveasfilename(
            title="Simpan Hasil Sebagai...",
            defaultextension=".xlsx",
            initialfile="Kamus Data Master Updated.xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not self.output_file:
            messagebox.showwarning("Batal", "Proses dibatalkan.")
            return

        self.show_loading_screen()
        self.processing_result = None
        self.processing_thread = threading.Thread(target=self.run_processing)
        self.processing_thread.daemon = True
        self.processing_thread.start()
        self.check_thread()

    def check_thread(self):
        while not self.log_queue.empty():
            msg = self.log_queue.get_nowait()
            if hasattr(self, 'log_text') and self.loading_window and self.loading_window.winfo_exists():
                self.log_text.config(state=tk.NORMAL)
                self.log_text.insert(tk.END, msg + "\n")
                self.log_text.see(tk.END)
                self.log_text.config(state=tk.DISABLED)
        if self.processing_thread.is_alive():
            self.root.after(100, self.check_thread)
        else:
            self.close_loading_screen()
            if self.processing_result:
                if self.processing_result['status'] == 'success':
                    messagebox.showinfo("Berhasil", self.processing_result['message'])
                else:
                    messagebox.showerror("Error", self.processing_result['message'])

    def run_processing(self):
        """
        Fungsi utama untuk proses gabungan data.
        
        Langkah:
        1. Baca file Kamus Data Master
        2. Baca file EXA dengan filter BAT
        3. Baca file INV dengan filter BAT
        4. Gabungkan semua data
        5. Simpan hasil ke Excel
        
        Detail:
        - Barcode range otomatis dipecah (123456-123460 jadi 5 baris)
        - Barcode duplikat dihapus (ambil yang terbaru)
        - Filter BAT berfungsi dengan benar
        - Semua lokasi diproses
        - Tahun perolehan dari kolom K (EXA) dan P (INV)
        - Asset Oracle dari kolom E
        """
        try:
            # 1) MASTER - Baca dari sheet yang dipilih user
            self.log_queue.put("[1] Membaca Kamus Data Master...")
            df_master = self.read_master_file(self.master_file)
            df_master = self.map_master_to_template(df_master)
            self.log_queue.put(f"    ✅ Master data: {len(df_master)} baris")

            # 2) Proses EXA - File asset existing
            self.log_queue.put("[2] 📊 Proses file EXA...")
            self.log_queue.put(f"    • Sheet: {', '.join(self.exa_sheets)}")
            selected_bats_exa = self.get_selected_bats(self.bat_opts_exa)
            self.log_queue.put(f"    • Filter BAT: {', '.join(selected_bats_exa) if selected_bats_exa else 'Semua'}")
            df_exa = self.process_exa_file(self.exa_file, self.exa_sheets, selected_bats_exa)
            self.log_queue.put(f"    ✅ Data EXA: {len(df_exa)} baris")

            # 3) Proses INV - File asset baru
            self.log_queue.put("[3] 📊 Proses file INV...")
            self.log_queue.put(f"    • Sheet: {', '.join(self.inv_sheets)}")
            selected_bats_inv = self.get_selected_bats(self.bat_opts_inv)
            self.log_queue.put(f"    • Filter BAT: {', '.join(selected_bats_inv) if selected_bats_inv else 'Semua'}")
            df_inv = self.process_inv_file(self.inv_file, self.inv_sheets, selected_bats_inv)
            self.log_queue.put(f"    ✅ Data INV: {len(df_inv)} baris")

            # 4) Gabungkan dan normalize dengan source tracking
            self.log_queue.put("[4] Menggabungkan & menormalkan data (vertical expand)...")
            
            # Add source tracking for priority deduplication
            df_master['__source'] = 'MASTER'
            df_exa['__source'] = 'EXA' 
            df_inv['__source'] = 'INV'
            
            df_all = pd.concat([df_master, df_exa, df_inv], ignore_index=True)
            self.log_queue.put(f"    • Total baris sebelum normalisasi: {len(df_all)}")
            df_final, qa_stats = self.finalize_dataframe_vertical_enhanced(df_all)
            self.log_queue.put(f"    ✅ Total baris setelah normalisasi: {len(df_final)}")
            
            # Log QA statistics
            self.log_queue.put("[QA] Statistik Processing:")
            for key, value in qa_stats.items():
                self.log_queue.put(f"    • {key}: {value:,}")

            # 5) Simpan hasil
            self.log_queue.put("[5] Menyimpan ke Excel...")
            self.save_with_formatting(df_final, self.output_file)
            self.log_queue.put(f"    ✅ File berhasil disimpan: {os.path.basename(self.output_file)}")

            self.processing_result = {
                'status': 'success',
                'message': f"✅ Proses berhasil!\n\nFile: {os.path.basename(self.output_file)}\nTotal baris: {len(df_final):,}\n\nBreakdown:\n- Master: {len(df_master):,} baris\n- EXA: {len(df_exa):,} baris\n- INV: {len(df_inv):,} baris"
            }
        except Exception as e:
            self.processing_result = {'status': 'error', 'message': f"❌ Terjadi kesalahan:\n\n{e}\n\n{traceback.format_exc()}"}

    # ---------- Helpers ----------
    def get_selected_bats(self, opts_dict):
        return {name for name, var in opts_dict.items() if var.get()}

    def read_master_file(self, path):
        """
        Membaca file master menggunakan sheet yang dipilih user.
        Jika tidak ada sheet yang dipilih, gunakan 'cleanData' atau sheet pertama.
        """
        try:
            # Gunakan sheet yang dipilih user
            if self.master_sheet:
                df = pd.read_excel(path, sheet_name=self.master_sheet)
                self.log_queue.put(f"    • Menggunakan sheet: {self.master_sheet}")
            else:
                # Fallback ke cleanData atau sheet pertama
                xl = pd.ExcelFile(path)
                sheet_to_use = 'cleanData' if 'cleanData' in xl.sheet_names else xl.sheet_names[0]
                df = pd.read_excel(path, sheet_name=sheet_to_use)
                self.log_queue.put(f"    • Menggunakan sheet default: {sheet_to_use}")
        except Exception as e:
            self.log_queue.put(f"    ⚠ Error membaca master file: {e}")
            df = pd.read_excel(path)
        return df

    def map_master_to_template(self, df):
        target_cols = ['NO BARCODE', 'ASSET ORACLE', 'LOKASI', 'JENIS HARTA', 'KONDISI', 'Tahun Perolehan']
        out = pd.DataFrame()
        for c in target_cols:
            out[c] = df[c] if c in df.columns else ''
        out['NO BARCODE'] = out['NO BARCODE'].apply(self.normalize_single_barcode)
        return out[target_cols]

    def smart_header_detection(self, df):
        """
        Smart header detection with keyword-based approach and column fallback.
        
        Strategy:
        1. Search for rows containing key column keywords
        2. Fallback to specific column positions if keywords not found
        3. Return both header row index and column mapping
        
        Returns:
            tuple: (header_row_index, column_mapping_dict)
        """
        # Define column keywords for detection
        column_keywords = {
            'NO BARCODE': ['BARCODE', 'NO BARCODE', 'KODE', 'NO. BARCODE'],
            'ASSET ORACLE': ['ASSET', 'ORACLE', 'NO ERS', 'NO ASSET', 'ASSET ORACLE'],
            'LOKASI': ['LOKASI', 'LOCATION', 'LOC', 'TEMPAT'],
            'JENIS HARTA': ['JENIS', 'HARTA', 'TYPE', 'KATEGORI'],
            'KONDISI': ['KONDISI', 'CONDITION', 'STATUS'],
            'BAT': ['BAT', 'KATEGORI BAT', 'DIVISI'],
            'PERO/AN': ['PERO', 'BULAN', 'MONTH', 'PEROLEHAN'],
            'PERO/AN.1': ['TAHUN', 'YEAR', 'PERO/AN.1']
        }
        
        # Search for header row
        for i, row in df.head(25).iterrows():
            row_str = ' '.join(str(s).strip().upper() for s in row.dropna() if str(s).strip())
            
            # Count how many required keywords are found
            found_keywords = 0
            for target_col, keywords in column_keywords.items():
                if any(kw.upper() in row_str for kw in keywords):
                    found_keywords += 1
                    
            # If we find at least 3 key columns, this is likely the header
            if found_keywords >= 3:
                return i, self._build_column_mapping(df.iloc[i], column_keywords)
                
        # Fallback: assume row 0 is header and use column position fallback
        return 0, self._build_column_mapping_fallback(df.iloc[0])
    
    def _build_column_mapping(self, header_row, column_keywords):
        """
        Build column mapping based on header row and keywords.
        """
        mapping = {}
        header_list = [str(col).strip().upper() for col in header_row]
        
        for target_col, keywords in column_keywords.items():
            for i, header_val in enumerate(header_list):
                if any(kw.upper() in header_val for kw in keywords):
                    mapping[target_col] = header_row.iloc[i]
                    break
                    
        return mapping
    
    def _build_column_mapping_fallback(self, header_row):
        """
        Fallback column mapping using traditional column positions.
        """
        mapping = {}
        columns = list(header_row)
        
        # Traditional fallback positions
        fallback_positions = {
            'NO BARCODE': ['K', 10],  # Column K or index 10
            'ASSET ORACLE': ['E', 4], # Column E or index 4  
            'LOKASI': ['P', 15],      # Column P or index 15
        }
        
        for target_col, (col_letter, col_index) in fallback_positions.items():
            if col_index < len(columns):
                mapping[target_col] = columns[col_index]
                
        return mapping
    
    def find_header_row(self, df, required_keywords):
        """
        Legacy function - now uses smart_header_detection.
        """
        header_idx, _ = self.smart_header_detection(df)
        return header_idx

    def get_col_data(self, row, primary_name, fallbacks=[]):
        """
        Mendapatkan data dari kolom dengan fallback names yang lebih fleksibel.
        Mendukung partial matching untuk nama kolom.
        """
        all_names = [primary_name] + list(fallbacks)
        
        # Exact match first
        for name in all_names:
            if name in row.index and pd.notna(row[name]):
                val = row[name]
                val_str = str(val).strip()
                if val_str not in ('', 'nan', 'NaN', 'None', 'none', '<NA>'):
                    return val
        
        # Partial match (case insensitive) - untuk menangani variasi nama kolom
        # Coba semua nama (primary + fallbacks)
        for search_name in all_names:
            search_upper = search_name.upper()
            for col in row.index:
                col_upper = str(col).upper()
                # Cek apakah nama kolom mengandung keyword
                if search_upper in col_upper or col_upper in search_upper:
                    if pd.notna(row[col]):
                        val = row[col]
                        val_str = str(val).strip()
                        if val_str not in ('', 'nan', 'NaN', 'None', 'none', '<NA>'):
                            return val
        
        return ''

    # ---------- Enhanced Barcode Parser ----------
    def normalize_and_expand_barcodes(self, val):
        """
        Production-ready barcode parser dengan comprehensive multi-value support.
        
        New Features:
        - Multi-value barcode handling (comma-separated)
        - Enhanced description removal
        - Better range expansion validation
        - Comprehensive logging
        """
        if not isinstance(val, str):
            val = str(val or '').strip()
        
        # Handle special cases
        if not val or val.lower() in ('nan', 'none', '', '(tidak ada)', 'tidak ada'):
            return []
        
        original_val = val
        barcodes = []
        
        # Step 1: Handle multi-value barcodes (comma/semicolon separated)
        if any(delimiter in val for delimiter in [',', ';']):
            tokens = re.split(r'[,;]+', val)
            if len(tokens) > 1:
                self.log_queue.put(f"🔍 Multi-value barcode detected: {len(tokens)} tokens")
        else:
            tokens = [val]
        
        # Step 2: Process each token
        for token in tokens:
            token = token.strip()
            if not token:
                continue
            
            # Remove descriptions in parentheses and brackets
            clean_token = re.sub(r'[\(\[][^\)\]]*[\)\]]', '', token).strip()
            
            # Check for range pattern (e.g., 2408005143-2408005156)
            range_match = re.match(r'(\d{6,})\s*[-–]\s*(\d{6,})', clean_token)
            if range_match:
                start_num = int(range_match.group(1))
                end_num = int(range_match.group(2))
                
                # Safety validation for ranges
                if start_num <= end_num and (end_num - start_num) <= 1000:
                    expanded_range = [str(x) for x in range(start_num, end_num + 1)]
                    barcodes.extend(expanded_range)
                    self.log_queue.put(f"📈 Expanded range {start_num}-{end_num}: {len(expanded_range)} barcodes")
                else:
                    self.log_queue.put(f"⚠️ Invalid range {start_num}-{end_num}: skipped")
                continue
            
            # Extract individual barcode numbers (6+ digits)
            barcode_matches = re.findall(r'\b(\d{6,})\b', clean_token)
            if barcode_matches:
                barcodes.extend(barcode_matches)
            else:
                # Fallback: try to extract any number sequence
                number_match = re.search(r'(\d{4,})', clean_token)
                if number_match:
                    barcodes.append(number_match.group(1))
        
        # Step 3: Remove duplicates while preserving order
        unique_barcodes = []
        seen = set()
        for bc in barcodes:
            if bc not in seen:
                unique_barcodes.append(bc)
                seen.add(bc)
        
        # Log expansion results for complex cases
        if len(unique_barcodes) > 1 or len(original_val) > 20:
            preview = original_val[:50] + "..." if len(original_val) > 50 else original_val
            self.log_queue.put(f"🎯 Barcode expansion: '{preview}' → {len(unique_barcodes)} codes")
        
        return unique_barcodes
    
    def normalize_single_barcode(self, raw):
        """
        Legacy function for backward compatibility.
        Now uses the enhanced parser but returns single string.
        """
        barcodes = self.normalize_and_expand_barcodes(raw)
        return barcodes[0] if barcodes else ''

    def parse_barcodes(self, barcode_str):
        """
        Enhanced barcode parsing dengan support untuk berbagai format.
        Mengadopsi logic dari script yang diberikan dengan perbaikan.
        
        Args:
            barcode_str: String barcode yang akan di-parse
            
        Returns:
            List[str]: List barcode yang sudah dibersihkan
        """
        if pd.isna(barcode_str) or not isinstance(barcode_str, str) or barcode_str.strip() == '':
            return ['']

        # Normalisasi string dan delimiter
        cleaned_str = str(barcode_str).replace(';', ',').replace('\n', ',').replace('|', ',')
        
        # Handle kasus khusus
        special_cases = ['N/A', 'NULL', 'NONE', 'TIDAK ADA', 'KOSONG', '-']
        if cleaned_str.strip().upper() in special_cases:
            return [cleaned_str.strip()]

        # Enhanced regex pattern untuk barcode
        # Menangkap barcode numerik (10+ digit) dan alfanumerik legacy
        pattern = re.compile(r'\b\d{10,}\b|\b[A-Z0-9/.-]{6,}\b', re.IGNORECASE)
        barcodes = pattern.findall(cleaned_str)

        if not barcodes:
            # Fallback: coba ekstrak angka panjang
            number_pattern = re.compile(r'\d{6,}')
            numbers = number_pattern.findall(cleaned_str)
            if numbers:
                barcodes = numbers
            else:
                return ['']
                
        # Clean dan deduplicate
        cleaned_barcodes = []
        seen = set()
        
        for bc in barcodes:
            bc_clean = bc.strip()
            if bc_clean and bc_clean not in seen:
                seen.add(bc_clean)
                cleaned_barcodes.append(bc_clean)
        
        return cleaned_barcodes if cleaned_barcodes else ['']
    
    def expand_and_clean_barcodes(self, raw_barcode_text):
        """
        Legacy method untuk backward compatibility.
        Menggunakan parse_barcodes yang sudah ditingkatkan.
        """
        barcodes = self.parse_barcodes(raw_barcode_text)
        
        # Handle range expansion (e.g., "123456-123460")
        expanded_barcodes = []
        for barcode in barcodes:
            if not barcode or barcode == '':
                continue
                
            # Check for range pattern
            range_match = re.match(r'^(\d{6,})\s*-\s*(\d{6,})$', barcode)
            if range_match:
                try:
                    start, end = int(range_match.group(1)), int(range_match.group(2))
                    if start <= end and end - start <= 200000:  # Safety limit
                        for i in range(start, end + 1):
                            expanded_barcodes.append(str(i))
                        continue
                except (ValueError, TypeError):
                    pass
            
            # Regular barcode
            expanded_barcodes.append(barcode)
        
        # Remove duplicates while preserving order
        seen = set()
        final_barcodes = []
        for bc in expanded_barcodes:
            if bc not in seen:
                seen.add(bc)
                final_barcodes.append(bc)
                
        return final_barcodes if final_barcodes else ['']
    
    def find_header_row_in_file(self, file_path, sheet_name, keywords):
        """
        Helper method untuk mencari header row dalam file Excel tertentu.
        
        Args:
            file_path (str): Path ke file Excel
            sheet_name (str): Nama sheet
            keywords (list): List keyword untuk dicari
            
        Returns:
            int: Index baris header (default 0 jika tidak ditemukan)
        """
        try:
            df_preview = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=25)
            header_row = self.find_header_row(df_preview, keywords)
            return header_row if header_row is not None else 0
        except Exception as e:
            self.log_queue.put(f"        ⚠️ Error finding header in {sheet_name}: {str(e)}")
            return 0
    
    def find_column_by_keywords(self, columns, keywords, exclude=None):
        """
        Mencari kolom berdasarkan keyword dengan exclude pattern.
        
        Args:
            columns: List kolom untuk dicari
            keywords (list): List keyword untuk dicari
            exclude (list): List keyword yang harus dihindari
            
        Returns:
            str: Nama kolom yang ditemukan, atau None
        """
        exclude = exclude or []
        
        for col in columns:
            col_upper = str(col).upper()
            
            # Check exclude patterns first
            if any(exc.upper() in col_upper for exc in exclude):
                continue
                
            # Check if any keyword matches
            for keyword in keywords:
                if keyword.upper() in col_upper:
                    return col
        
        return None
    
    def apply_bat_filter(self, df, selected_bats):
        """
        Apply BAT filtering dengan algoritma yang ditingkatkan.
        
        Args:
            df (pd.DataFrame): DataFrame untuk difilter
            selected_bats (set): Set kategori BAT yang dipilih
            
        Returns:
            pd.DataFrame: DataFrame yang sudah difilter
        """
        if 'BAT' not in df.columns or not selected_bats:
            return df
            
        bat_series = df['BAT'].astype(str).str.strip().str.upper()
        
        # Build filter mask
        masks = []
        if 'ICT' in selected_bats: 
            masks.append(bat_series.str.contains('ICT', case=False, na=False))
        if 'ENG' in selected_bats: 
            masks.append(bat_series.str.contains('ENG', case=False, na=False))
        if 'HRGA' in selected_bats: 
            masks.append(bat_series.str.contains('HRGA', case=False, na=False))
        if 'BAT' in selected_bats: 
            masks.append(bat_series.eq('BAT'))
        if 'BLANKS' in selected_bats: 
            masks.append(bat_series.eq('') | bat_series.eq('NAN') | bat_series.isna())
        
        if masks:
            combined_mask = masks[0]
            for mask in masks[1:]:
                combined_mask = combined_mask | mask
            return df[combined_mask]
        
        return df
    
    def process_date_columns(self, year_series, month_series):
        """
        Process year dan month columns menjadi datetime.
        
        Args:
            year_series: Series berisi tahun
            month_series: Series berisi bulan
            
        Returns:
            pd.Series: Series datetime
        """
        try:
            # Convert to numeric first
            years = pd.to_numeric(year_series, errors='coerce')
            months = pd.to_numeric(month_series, errors='coerce')
            
            # Create datetime
            dates = []
            for year, month in zip(years, months):
                try:
                    if pd.notna(year) and pd.notna(month) and 1 <= month <= 12:
                        date_str = f"{int(year)}-{int(month)}-01"
                        dates.append(pd.to_datetime(date_str, errors='coerce'))
                    else:
                        dates.append(pd.NaT)
                except (ValueError, TypeError):
                    dates.append(pd.NaT)
            
            return pd.Series(dates)
            
        except Exception:
            return pd.Series([pd.NaT] * len(year_series))
    
    def extract_year_from_column(self, df, column_letter, sheet_name=""):
        """
        Enhanced extraction tahun perolehan dari kolom tertentu dengan format fleksibel.
        Untuk EXA: kolom K, untuk INV: kolom P
        
        Args:
            df: DataFrame yang akan diproses
            column_letter: Huruf kolom (K untuk EXA, P untuk INV)
            sheet_name: Nama sheet untuk logging
            
        Returns:
            pd.Series: Series dengan tahun perolehan
        """
        try:
            # Convert column letter to index (K=10, P=15)
            col_index = ord(column_letter.upper()) - ord('A')
            
            if col_index >= len(df.columns):
                self.log_queue.put(f"      ⚠️ Kolom {column_letter} tidak ditemukan di sheet {sheet_name}")
                return pd.Series([pd.NaT] * len(df))
            
            # Get column data
            col_data = df.iloc[:, col_index]
            
            # Process each value
            years = []
            for value in col_data:
                year_value = self.parse_year_value(value)
                years.append(year_value)
            
            self.log_queue.put(f"      ✅ Berhasil ekstrak tahun dari kolom {column_letter} di sheet {sheet_name}")
            return pd.Series(years)
            
        except Exception as e:
            self.log_queue.put(f"      ❌ Error ekstrak tahun dari kolom {column_letter}: {str(e)}")
            return pd.Series([pd.NaT] * len(df))
    
    def parse_year_value(self, value):
        """
        Parse berbagai format tahun dari cell value.
        Support format: 2024, "2024", "Jan-2024", "2024-01", dll.
        
        Args:
            value: Cell value yang akan di-parse
            
        Returns:
            datetime atau NaT
        """
        if pd.isna(value):
            return pd.NaT
        
        value_str = str(value).strip()
        if not value_str or value_str.upper() in ['NAN', 'NULL', 'NONE', '']:
            return pd.NaT
        
        # Try different year patterns
        year_patterns = [
            r'\b(20\d{2})\b',  # 2024
            r'\b(19\d{2})\b',  # 1999
            r'(20\d{2})-\d{1,2}',  # 2024-01
            r'\d{1,2}-(20\d{2})',  # 01-2024
        ]
        
        for pattern in year_patterns:
            match = re.search(pattern, value_str)
            if match:
                try:
                    year = int(match.group(1))
                    if 1900 <= year <= 2100:  # Reasonable year range
                        return pd.to_datetime(f"{year}-01-01")
                except (ValueError, IndexError):
                    continue
        
        # Try direct conversion
        try:
            year_int = int(float(value_str))
            if 1900 <= year_int <= 2100:
                return pd.to_datetime(f"{year_int}-01-01")
        except (ValueError, TypeError):
            pass
        
        return pd.NaT
    
    def simple_copy_column_by_index(self, file_path, sheet_name, col_index, target_length):
        """
        SUPER SIMPLE: Copy column by index langsung dari Excel file.
        Tidak peduli header, BAT filter, atau apapun - langsung copy aja.
        
        Args:
            file_path: Path ke file Excel
            sheet_name: Nama sheet
            col_index: Index kolom (10 untuk K, 15 untuk P)
            target_length: Panjang yang diharapkan
            
        Returns:
            list: List data dari kolom
        """
        try:
            # Baca Excel tanpa header apapun
            df_raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            
            if col_index >= len(df_raw.columns):
                return [''] * target_length
            
            # Ambil kolom langsung
            col_data = df_raw.iloc[:, col_index]
            
            # Convert ke string dan clean
            cleaned = []
            for val in col_data:
                if pd.isna(val) or str(val).strip() in ['nan', 'NaN', 'None', 'NULL', '']:
                    cleaned.append('')
                else:
                    cleaned.append(str(val).strip())
            
            # Trim atau pad sesuai target_length
            if len(cleaned) >= target_length:
                return cleaned[:target_length]
            else:
                return cleaned + [''] * (target_length - len(cleaned))
                
        except Exception as e:
            return [''] * target_length
    
    def extract_column_data_from_file(self, file_path, sheet_name, column_letter, target_length):
        """
        Extract data langsung dari file Excel asli berdasarkan kolom letter.
        Untuk EXA: kolom K, untuk INV: kolom P
        
        Args:
            file_path: Path ke file Excel
            sheet_name: Nama sheet
            column_letter: Huruf kolom (K untuk EXA, P untuk INV)
            target_length: Panjang data yang diharapkan
            
        Returns:
            pd.Series: Series dengan data mentah dari kolom
        """
        try:
            # Read Excel file langsung tanpa header processing
            df_raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            
            # Convert column letter to index (K=10, P=15)
            col_index = ord(column_letter.upper()) - ord('A')
            
            if col_index >= len(df_raw.columns):
                self.log_queue.put(f"      ⚠️ Kolom {column_letter} tidak ditemukan di sheet {sheet_name}")
                return pd.Series([''] * target_length)
            
            # Get column data as-is (copy paste)
            col_data = df_raw.iloc[:, col_index]
            
            # Convert to string and clean basic formatting
            cleaned_data = col_data.astype(str).str.strip()
            
            # Replace common null values with empty string
            cleaned_data = cleaned_data.replace(['nan', 'NaN', 'None', 'NULL'], '')
            
            # Trim or pad to match target length
            if len(cleaned_data) > target_length:
                cleaned_data = cleaned_data.iloc[:target_length]
            elif len(cleaned_data) < target_length:
                # Pad with empty strings
                padding = pd.Series([''] * (target_length - len(cleaned_data)))
                cleaned_data = pd.concat([cleaned_data, padding], ignore_index=True)
            
            self.log_queue.put(f"      ✅ Copy-paste data dari kolom {column_letter} di sheet {sheet_name} ({len(cleaned_data)} baris)")
            return cleaned_data
            
        except Exception as e:
            self.log_queue.put(f"      ❌ Error copy-paste dari kolom {column_letter}: {str(e)}")
            return pd.Series([''] * target_length)
    
    def apply_bat_filter_enhanced(self, df, selected_bats, sheet_name=""):
        """
        Enhanced BAT filtering yang benar-benar berfungsi.
        Implementasi filter yang lebih robust dan akurat.
        
        Args:
            df: DataFrame yang akan difilter
            selected_bats: Set kategori BAT yang dipilih
            sheet_name: Nama sheet untuk logging
            
        Returns:
            pd.DataFrame: DataFrame yang sudah difilter
        """
        if 'BAT' not in df.columns or not selected_bats:
            self.log_queue.put(f"      ⚠️ Kolom BAT tidak ditemukan atau tidak ada filter yang dipilih di sheet {sheet_name}")
            return df
        
        original_count = len(df)
        
        # Convert BAT column to string and normalize
        bat_series = df['BAT'].astype(str).str.strip().str.upper()
        
        # Build comprehensive filter mask
        filter_masks = []
        
        for bat_category in selected_bats:
            if bat_category == 'ICT':
                mask = bat_series.str.contains('ICT', case=False, na=False)
                filter_masks.append(mask)
                
            elif bat_category == 'ENG':
                mask = bat_series.str.contains('ENG', case=False, na=False)
                filter_masks.append(mask)
                
            elif bat_category == 'HRGA':
                mask = (bat_series.str.contains('HRGA', case=False, na=False) | 
                       bat_series.str.contains('HR', case=False, na=False) |
                       bat_series.str.contains('GA', case=False, na=False))
                filter_masks.append(mask)
                
            elif bat_category == 'BAT':
                mask = (bat_series.eq('BAT') | 
                       bat_series.str.contains('BATTERY', case=False, na=False))
                filter_masks.append(mask)
                
            elif bat_category == 'BLANKS':
                mask = (bat_series.eq('') | 
                       bat_series.eq('NAN') | 
                       bat_series.isna() | 
                       bat_series.eq('NULL') |
                       bat_series.eq('NONE'))
                filter_masks.append(mask)
        
        # Combine all masks with OR operation
        if filter_masks:
            combined_mask = filter_masks[0]
            for mask in filter_masks[1:]:
                combined_mask = combined_mask | mask
            
            filtered_df = df[combined_mask].copy()
            filtered_count = len(filtered_df)
            
            self.log_queue.put(f"      ✅ BAT Filter applied di sheet {sheet_name}: {original_count} → {filtered_count} baris")
            self.log_queue.put(f"        Filter: {', '.join(selected_bats)}")
            
            return filtered_df
        else:
            self.log_queue.put(f"      ⚠️ Tidak ada filter BAT yang valid di sheet {sheet_name}")
            return df

    def process_exa_file(self, file_path, sheets_to_process, selected_bats):
        """
        Proses file EXA (asset existing).
        
        Input:
            file_path: Path file EXA
            sheets_to_process: Daftar sheet yang dipilih
            selected_bats: Filter BAT yang dipilih
            
        Output:
            DataFrame dengan data yang sudah diproses
        """
        self.log_queue.put(f"    📊 Memproses file EXA: {os.path.basename(file_path)}")
        
        try:
            processed_dfs = []
            
            for sheet_name in sheets_to_process:
                self.log_queue.put(f"      📋 Processing sheet: {sheet_name}")
                
                # Cari baris header
                header_row = self.find_header_row_in_file(file_path, sheet_name, ['NO BARCODE', 'LOKASI', 'BAT'])
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
                
                # Enhanced column mapping
                column_mapping = {
                    'NO BARCODE': 'NO BARCODE',
                    'NO PO': 'ASSET ORACLE',
                    'ASSET ORACLE': 'ASSET ORACLE',
                    'NO ERS': 'ASSET ORACLE', 
                    'NO ASSET': 'ASSET ORACLE',
                    'LOKASI': 'LOKASI',
                    'JENIS HARTA': 'JENIS HARTA',
                    'KONDISI': 'KONDISI',
                    'BAT': 'BAT'
                }
                
                # Smart column detection untuk tanggal
                year_col = self.find_column_by_keywords(df.columns, ['TAHUN', 'PERO/AN.1', 'YEAR'])
                month_col = self.find_column_by_keywords(df.columns, ['BULAN', 'PERO/AN', 'MONTH'], exclude=['TAHUN', '.1'])

                if year_col:
                    column_mapping[year_col] = 'Year'
                if month_col:
                    column_mapping[month_col] = 'Month'
                    
                # Apply column mapping
                df = df.rename(columns=column_mapping)
                
                # Select required columns
                required_cols = list(set(column_mapping.values()))
                df = df[[col for col in required_cols if col in df.columns]]

                # Terapkan filter BAT
                if 'BAT' in df.columns and selected_bats:
                    df = self.apply_bat_filter_enhanced(df, selected_bats, sheet_name)

                # Copy data tahun (kolom K) dan asset oracle (kolom E)
                self.log_queue.put(f"      📊 Copy data dari kolom K & E di sheet {sheet_name}")
                
                try:
                    # Baca file dengan header yang sama seperti df utama
                    df_temp = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
                    
                    # Ambil kolom K (index 10) langsung
                    if len(df_temp.columns) > 10:
                        col_k_raw = df_temp.iloc[:, 10]  # Kolom K
                        
                        # Apply BAT filter yang sama jika ada
                        if 'BAT' in df_temp.columns and selected_bats:
                            # Create same filter mask
                            bat_series = df_temp['BAT'].astype(str).str.strip().str.upper()
                            filter_masks = []
                            
                            for bat_category in selected_bats:
                                if bat_category == 'ICT':
                                    mask = bat_series.str.contains('ICT', case=False, na=False)
                                elif bat_category == 'ENG':
                                    mask = bat_series.str.contains('ENG', case=False, na=False)
                                elif bat_category == 'HRGA':
                                    mask = (bat_series.str.contains('HRGA', case=False, na=False) | 
                                           bat_series.str.contains('HR', case=False, na=False) |
                                           bat_series.str.contains('GA', case=False, na=False))
                                elif bat_category == 'BAT':
                                    mask = (bat_series.eq('BAT') | 
                                           bat_series.str.contains('BATTERY', case=False, na=False))
                                elif bat_category == 'BLANKS':
                                    mask = (bat_series.eq('') | bat_series.eq('NAN') | 
                                           bat_series.isna() | bat_series.eq('NULL') |
                                           bat_series.eq('NONE'))
                                filter_masks.append(mask)
                            
                            if filter_masks:
                                combined_mask = filter_masks[0]
                                for mask in filter_masks[1:]:
                                    combined_mask = combined_mask | mask
                                col_k_filtered = col_k_raw[combined_mask]
                            else:
                                col_k_filtered = col_k_raw
                        else:
                            col_k_filtered = col_k_raw
                        
                        # Clean data
                        tahun_data = col_k_filtered.astype(str).str.strip().replace(['nan', 'NaN', 'None', 'NULL'], '')
                        tahun_data = tahun_data.reset_index(drop=True)
                        
                        # Ensure same length as df
                        if len(tahun_data) >= len(df):
                            df['Tahun Perolehan'] = tahun_data.iloc[:len(df)].values
                        else:
                            # Pad with empty strings
                            padded_data = list(tahun_data) + [''] * (len(df) - len(tahun_data))
                            df['Tahun Perolehan'] = padded_data
                        
                        self.log_queue.put(f"      ✅ SIMPLE: Berhasil copy kolom K - {len(df)} baris")
                        
                        # BONUS: Copy kolom E untuk ASSET ORACLE juga
                        if len(df_temp.columns) > 4:  # Kolom E = index 4
                            col_e_raw = df_temp.iloc[:, 4]  # Kolom E
                            
                            # Apply filter yang sama
                            if 'BAT' in df_temp.columns and selected_bats:
                                col_e_filtered = col_e_raw[combined_mask]
                            else:
                                col_e_filtered = col_e_raw
                            
                            # Clean dan assign
                            asset_data = col_e_filtered.astype(str).str.strip().replace(['nan', 'NaN', 'None', 'NULL'], '')
                            asset_data = asset_data.reset_index(drop=True)
                            
                            if len(asset_data) >= len(df):
                                df['ASSET ORACLE'] = asset_data.iloc[:len(df)].values
                            else:
                                padded_asset = list(asset_data) + [''] * (len(df) - len(asset_data))
                                df['ASSET ORACLE'] = padded_asset
                            
                            self.log_queue.put(f"      ✅ BONUS: Copy kolom E (ASSET ORACLE) - {len(df)} baris")
                    else:
                        df['Tahun Perolehan'] = ''
                        self.log_queue.put(f"      ⚠️ SIMPLE: Kolom K tidak ditemukan")
                        
                except Exception as e:
                    self.log_queue.put(f"      ❌ SIMPLE: Error copy kolom K: {str(e)}")
                    # FALLBACK: Super simple copy
                    self.log_queue.put(f"      🔄 FALLBACK: Super simple copy kolom K")
                    simple_data = self.simple_copy_column_by_index(file_path, sheet_name, 10, len(df))
                    df['Tahun Perolehan'] = simple_data
                    self.log_queue.put(f"      ✅ FALLBACK: Copy {len(simple_data)} data dari kolom K")
                    
                    # FALLBACK BONUS: Copy kolom E untuk ASSET ORACLE
                    self.log_queue.put(f"      🔄 FALLBACK BONUS: Copy kolom E (ASSET ORACLE)")
                    asset_simple = self.simple_copy_column_by_index(file_path, sheet_name, 4, len(df))
                    df['ASSET ORACLE'] = asset_simple
                    self.log_queue.put(f"      ✅ FALLBACK BONUS: Copy {len(asset_simple)} ASSET ORACLE dari kolom E")
                
                # Proses barcode
                if 'NO BARCODE' in df.columns:
                    df['NO BARCODE'] = df['NO BARCODE'].apply(self.parse_barcodes)
                    df = df.explode('NO BARCODE').reset_index(drop=True)
                
                # Final column selection
                final_cols = ['NO BARCODE', 'ASSET ORACLE', 'LOKASI', 'JENIS HARTA', 'KONDISI', 'Tahun Perolehan']
                df_final = df.reindex(columns=final_cols).fillna('')
                
                if not df_final.empty:
                    processed_dfs.append(df_final)
                    self.log_queue.put(f"        ✅ {len(df_final)} baris diproses dari sheet {sheet_name}")
                else:
                    self.log_queue.put(f"        ⚠️ Tidak ada data valid di sheet {sheet_name}")
            
            if processed_dfs:
                result_df = pd.concat(processed_dfs, ignore_index=True)
                self.log_queue.put(f"    ✅ Total EXA: {len(result_df)} baris")
                return result_df
            else:
                self.log_queue.put(f"    ❌ Tidak ada data EXA yang berhasil diproses")
                return pd.DataFrame(columns=final_cols)
                
        except Exception as e:
            self.log_queue.put(f"    ❌ Error processing EXA file: {str(e)}")
            return pd.DataFrame(columns=['NO BARCODE', 'ASSET ORACLE', 'LOKASI', 'JENIS HARTA', 'KONDISI', 'Tahun Perolehan'])
    
    def process_inv_file(self, file_path, sheets_to_process, selected_bats):
        """
        Proses file INV (asset baru).
        
        Input:
            file_path: Path file INV
            sheets_to_process: Daftar sheet yang dipilih
            selected_bats: Filter BAT yang dipilih
            
        Output:
            DataFrame dengan data yang sudah diproses
        """
        self.log_queue.put(f"    📊 Memproses file INV: {os.path.basename(file_path)}")
        
        try:
            processed_dfs = []
            
            for sheet_name in sheets_to_process:
                self.log_queue.put(f"      📋 Processing sheet: {sheet_name}")
                
                # Cari baris header
                header_row = self.find_header_row_in_file(file_path, sheet_name, ['NO BARCODE', 'LOKASI', 'BAT'])
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)

                # Enhanced column mapping untuk INV
                column_mapping = {
                    'NO BARCODE': 'NO BARCODE',
                    'ASSET ORACLE': 'ASSET ORACLE',
                    'NO ERS': 'ASSET ORACLE',
                    'NO ASSET': 'ASSET ORACLE',
                    'LOKASI': 'LOKASI',
                    'JENIS HARTA': 'JENIS HARTA',
                    'KONDISI': 'KONDISI',
                    'BAT': 'BAT'
                }
                
                # Smart column detection untuk tanggal INV
                year_col = self.find_column_by_keywords(df.columns, ['TAHUN', 'THAN', 'YEAR'])
                month_col = self.find_column_by_keywords(df.columns, ['BULAN', 'PERO/AN', 'MONTH'], exclude=['TAHUN'])

                if year_col:
                    column_mapping[year_col] = 'Year'
                if month_col:
                    column_mapping[month_col] = 'Month'
                    
                # Apply column mapping
                df = df.rename(columns=column_mapping)
                
                # Select required columns
                required_cols = list(set(column_mapping.values()))
                df = df[[col for col in required_cols if col in df.columns]]

                # Terapkan filter BAT
                if 'BAT' in df.columns and selected_bats:
                    df = self.apply_bat_filter_enhanced(df, selected_bats, sheet_name)

                # Copy data tahun (kolom P) dan asset oracle (kolom E)
                self.log_queue.put(f"      📊 Copy data dari kolom P & E di sheet {sheet_name}")
                
                try:
                    # Baca file dengan header yang sama seperti df utama
                    df_temp = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
                    
                    # Ambil kolom P (index 15) langsung
                    if len(df_temp.columns) > 15:
                        col_p_raw = df_temp.iloc[:, 15]  # Kolom P
                        
                        # Apply BAT filter yang sama jika ada
                        if 'BAT' in df_temp.columns and selected_bats:
                            # Create same filter mask
                            bat_series = df_temp['BAT'].astype(str).str.strip().str.upper()
                            filter_masks = []
                            
                            for bat_category in selected_bats:
                                if bat_category == 'ICT':
                                    mask = bat_series.str.contains('ICT', case=False, na=False)
                                elif bat_category == 'ENG':
                                    mask = bat_series.str.contains('ENG', case=False, na=False)
                                elif bat_category == 'HRGA':
                                    mask = (bat_series.str.contains('HRGA', case=False, na=False) | 
                                           bat_series.str.contains('HR', case=False, na=False) |
                                           bat_series.str.contains('GA', case=False, na=False))
                                elif bat_category == 'BAT':
                                    mask = (bat_series.eq('BAT') | 
                                           bat_series.str.contains('BATTERY', case=False, na=False))
                                elif bat_category == 'BLANKS':
                                    mask = (bat_series.eq('') | bat_series.eq('NAN') | 
                                           bat_series.isna() | bat_series.eq('NULL') |
                                           bat_series.eq('NONE'))
                                filter_masks.append(mask)
                            
                            if filter_masks:
                                combined_mask = filter_masks[0]
                                for mask in filter_masks[1:]:
                                    combined_mask = combined_mask | mask
                                col_p_filtered = col_p_raw[combined_mask]
                            else:
                                col_p_filtered = col_p_raw
                        else:
                            col_p_filtered = col_p_raw
                        
                        # Clean data
                        tahun_data = col_p_filtered.astype(str).str.strip().replace(['nan', 'NaN', 'None', 'NULL'], '')
                        tahun_data = tahun_data.reset_index(drop=True)
                        
                        # Ensure same length as df
                        if len(tahun_data) >= len(df):
                            df['Tahun Perolehan'] = tahun_data.iloc[:len(df)].values
                        else:
                            # Pad with empty strings
                            padded_data = list(tahun_data) + [''] * (len(df) - len(tahun_data))
                            df['Tahun Perolehan'] = padded_data
                        
                        self.log_queue.put(f"      ✅ SIMPLE: Berhasil copy kolom P - {len(df)} baris")
                        
                        # BONUS: Copy kolom E untuk ASSET ORACLE juga
                        if len(df_temp.columns) > 4:  # Kolom E = index 4
                            col_e_raw = df_temp.iloc[:, 4]  # Kolom E
                            
                            # Apply filter yang sama
                            if 'BAT' in df_temp.columns and selected_bats:
                                col_e_filtered = col_e_raw[combined_mask]
                            else:
                                col_e_filtered = col_e_raw
                            
                            # Clean dan assign
                            asset_data = col_e_filtered.astype(str).str.strip().replace(['nan', 'NaN', 'None', 'NULL'], '')
                            asset_data = asset_data.reset_index(drop=True)
                            
                            if len(asset_data) >= len(df):
                                df['ASSET ORACLE'] = asset_data.iloc[:len(df)].values
                            else:
                                padded_asset = list(asset_data) + [''] * (len(df) - len(asset_data))
                                df['ASSET ORACLE'] = padded_asset
                            
                            self.log_queue.put(f"      ✅ BONUS: Copy kolom E (ASSET ORACLE) - {len(df)} baris")
                    else:
                        df['Tahun Perolehan'] = ''
                        self.log_queue.put(f"      ⚠️ SIMPLE: Kolom P tidak ditemukan")
                        
                except Exception as e:
                    self.log_queue.put(f"      ❌ SIMPLE: Error copy kolom P: {str(e)}")
                    # FALLBACK: Super simple copy
                    self.log_queue.put(f"      🔄 FALLBACK: Super simple copy kolom P")
                    simple_data = self.simple_copy_column_by_index(file_path, sheet_name, 15, len(df))
                    df['Tahun Perolehan'] = simple_data
                    self.log_queue.put(f"      ✅ FALLBACK: Copy {len(simple_data)} data dari kolom P")
                    
                    # FALLBACK BONUS: Copy kolom E untuk ASSET ORACLE
                    self.log_queue.put(f"      🔄 FALLBACK BONUS: Copy kolom E (ASSET ORACLE)")
                    asset_simple = self.simple_copy_column_by_index(file_path, sheet_name, 4, len(df))
                    df['ASSET ORACLE'] = asset_simple
                    self.log_queue.put(f"      ✅ FALLBACK BONUS: Copy {len(asset_simple)} ASSET ORACLE dari kolom E")
                    
                # Proses barcode
                if 'NO BARCODE' in df.columns:
                    df['NO BARCODE'] = df['NO BARCODE'].apply(self.parse_barcodes)
                    df = df.explode('NO BARCODE').reset_index(drop=True)

                # Final column selection
                final_cols = ['NO BARCODE', 'ASSET ORACLE', 'LOKASI', 'JENIS HARTA', 'KONDISI', 'Tahun Perolehan']
                df_final = df.reindex(columns=final_cols).fillna('')
                
                if not df_final.empty:
                    processed_dfs.append(df_final)
                    self.log_queue.put(f"        ✅ {len(df_final)} baris diproses dari sheet {sheet_name}")
                else:
                    self.log_queue.put(f"        ⚠️ Tidak ada data valid di sheet {sheet_name}")
            
            if processed_dfs:
                result_df = pd.concat(processed_dfs, ignore_index=True)
                self.log_queue.put(f"    ✅ Total INV: {len(result_df)} baris")
                return result_df
            else:
                self.log_queue.put(f"    ❌ Tidak ada data INV yang berhasil diproses")
                return pd.DataFrame(columns=final_cols)
                
        except Exception as e:
            self.log_queue.put(f"    ❌ Error processing INV file: {str(e)}")
            return pd.DataFrame(columns=['NO BARCODE', 'ASSET ORACLE', 'LOKASI', 'JENIS HARTA', 'KONDISI', 'Tahun Perolehan'])
    
    def process_source_file(self, filename, sheets_to_process, selected_bats):
        """
        Legacy method untuk backward compatibility.
        Menggunakan enhanced processing methods.
        
        Args:
            filename (str): Path ke file Excel yang akan diproses
            sheets_to_process (list): List nama sheet yang dipilih user untuk diproses
            selected_bats (set): Set kategori BAT yang dipilih (ICT/ENG/HRGA/BAT/BLANKS)
        
        Returns:
            pd.DataFrame: DataFrame hasil processing dengan kolom standar
        """
        # Detect file type berdasarkan nama atau path
        file_basename = os.path.basename(filename).upper()
        
        if 'EXA' in file_basename:
            return self.process_exa_file(filename, sheets_to_process, selected_bats)
        elif 'INV' in file_basename:
            return self.process_inv_file(filename, sheets_to_process, selected_bats)
        else:
            # Fallback ke EXA processing
            self.log_queue.put(f"    ⚠️ File type tidak terdeteksi, menggunakan EXA processing")
            return self.process_exa_file(filename, sheets_to_process, selected_bats)

    def finalize_dataframe_vertical(self, df_combined):
        final_rows = []
        for _, row in df_combined.iterrows():
            barcode_str = str(row.get('NO BARCODE', '')).strip()
            barcodes = self.expand_and_clean_barcodes(barcode_str)
            if not barcodes:
                if barcode_str == '' or barcode_str.lower() == 'nan':
                    final_rows.append(row)
                continue
            for bc in barcodes:
                new_row = row.copy()
                new_row['NO BARCODE'] = self.normalize_single_barcode(bc)
                final_rows.append(new_row)

        df_final = pd.DataFrame(final_rows)

        if 'Tahun Perolehan' in df_final.columns:
            df_final['Tahun Perolehan'] = df_final['Tahun Perolehan'].apply(self._safe_datetime_conversion)

        if 'NO BARCODE' in df_final.columns:
            df_final['NO BARCODE'] = df_final['NO BARCODE'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
            df_with = df_final[df_final['NO BARCODE'].str.lower().ne('nan') & df_final['NO BARCODE'].ne('')].copy()
            df_wo  = df_final[df_final['NO BARCODE'].str.lower().eq('nan') | df_final['NO BARCODE'].eq('')].copy()
            
            # Duplicate removal: simpan data yang paling lengkap
            # Hitung jumlah kolom terisi untuk setiap row
            df_with['_completeness'] = df_with.apply(
                lambda row: sum(1 for val in row if pd.notna(val) and str(val).strip() not in ['', 'nan', 'NaN']), 
                axis=1
            )
            
            # Sort by NO BARCODE dan completeness (descending)
            df_with = df_with.sort_values(['NO BARCODE', '_completeness'], ascending=[True, False])
            
            # Keep first (yang paling lengkap) untuk setiap barcode
            df_with = df_with.drop_duplicates(subset=['NO BARCODE'], keep='first')
            
            # Drop kolom helper
            df_with = df_with.drop('_completeness', axis=1)
            
            df_final = pd.concat([df_with, df_wo], ignore_index=True)

        final_columns = ['NO BARCODE','ASSET ORACLE','LOKASI','JENIS HARTA','KONDISI','Tahun Perolehan']
        for col in final_columns:
            if col not in df_final.columns: df_final[col] = ''
        df_final = df_final[final_columns].fillna('')
        return df_final
    
    def _build_bat_filter_mask(self, df, selected_bats):
        """
        Build BAT filter mask once per sheet for performance.
        
        Args:
            df (pd.DataFrame): DataFrame to filter
            selected_bats (set): Set of selected BAT categories
            
        Returns:
            pd.Series or None: Boolean mask for filtering, None if no BAT column
        """
        if 'BAT' not in df.columns or not selected_bats:
            return None
            
        bat_series = df['BAT'].astype(str).str.strip().str.upper()
        masks = []
        
        if 'ICT' in selected_bats: 
            masks.append(bat_series.eq('ICT'))
        if 'ENG' in selected_bats: 
            masks.append(bat_series.eq('ENG'))
        if 'HRGA' in selected_bats: 
            masks.append(bat_series.eq('HRGA'))
        if 'BAT' in selected_bats: 
            masks.append(bat_series.eq('BAT'))
        if 'BLANKS' in selected_bats: 
            masks.append(bat_series.eq('') | bat_series.eq('NAN') | bat_series.isna())
            
        if not masks:
            return None
            
        # Combine all masks with OR operation
        combined_mask = masks[0]
        for mask in masks[1:]:
            combined_mask = combined_mask | mask
            
        return combined_mask
    
    def validate_bat_filter_enhanced(self, df, selected_bats, source_name):
        """
        Enhanced BAT filter dengan validation dan comprehensive reporting.
        
        Improvements:
        - Detailed logging of filter results
        - BAT distribution analysis
        - Validation of filter effectiveness
        """
        if 'BAT' not in df.columns:
            self.log_queue.put(f"⚠️ {source_name}: No BAT column found - processing all rows")
            return df, {'no_bat_column': True, 'rows_processed': len(df)}
        
        original_count = len(df)
        bat_distribution = df['BAT'].fillna('BLANK').value_counts().to_dict()
        
        # Apply enhanced filter mask
        mask = self._build_bat_filter_mask(df, selected_bats)
        if mask is not None:
            df_filtered = df[mask].copy()
        else:
            df_filtered = df.copy()
        
        filtered_count = len(df_filtered)
        filter_efficiency = (filtered_count / original_count * 100) if original_count > 0 else 0
        
        # Enhanced logging
        self.log_queue.put(f"📊 {source_name} BAT Filter Analysis:")
        self.log_queue.put(f"    • Original rows: {original_count:,}")
        self.log_queue.put(f"    • Filtered rows: {filtered_count:,} ({filter_efficiency:.1f}%)")
        self.log_queue.put(f"    • Selected BATs: {list(selected_bats)}")
        
        # Log distribution details
        for bat_type, count in sorted(bat_distribution.items()):
            included = "✅" if bat_type in selected_bats or (bat_type == 'BLANK' and 'BLANKS' in selected_bats) else "❌"
            self.log_queue.put(f"    • {bat_type}: {count:,} rows {included}")
        
        return df_filtered, {
            'original_count': original_count,
            'filtered_count': filtered_count,
            'filter_efficiency': filter_efficiency,
            'bat_distribution': bat_distribution,
            'selected_bats': list(selected_bats)
        }
    
    def _extract_tahun_perolehan(self, row):
        """
        Enhanced Tahun Perolehan extraction with Month+Year combination.
        
        Strategy:
        1. If both Month (PERO/AN) and Year (PERO/AN.1) are available, combine them
        2. If only Year is available, use Year only
        3. Return consistent datetime format
        
        Args:
            row (pd.Series): Row data from DataFrame
            
        Returns:
            pd.Timestamp or pd.NaT: Parsed date or NaT if invalid
        """
        tahun_perolehan = pd.NaT
        
        # Try to get both month and year
        year_val = None
        month_val = None
        
        # Check for year column variations
        year_columns = ['PERO/AN.1', 'TAHUN', 'YEAR', 'Tahun Perolehan']
        for col in year_columns:
            if col in row.index:
                try:
                    year_val = pd.to_numeric(row[col], errors='coerce')
                    if pd.notna(year_val) and 1900 <= year_val <= 2100:
                        break
                except (ValueError, TypeError):
                    continue
                    
        # Check for month column variations  
        month_columns = ['PERO/AN', 'BULAN', 'MONTH']
        for col in month_columns:
            if col in row.index:
                try:
                    month_val = pd.to_numeric(row[col], errors='coerce')
                    if pd.notna(month_val) and 1 <= month_val <= 12:
                        break
                except (ValueError, TypeError):
                    continue
        
        # Combine month and year if both available
        if pd.notna(year_val) and pd.notna(month_val):
            try:
                tahun_perolehan = pd.to_datetime(f"{int(year_val)}-{int(month_val)}-01", errors='coerce')
                return tahun_perolehan
            except (ValueError, TypeError):
                pass
                
        # Use year only if available
        if pd.notna(year_val):
            try:
                tahun_perolehan = pd.to_datetime(f"{int(year_val)}-01-01", errors='coerce')
                return tahun_perolehan
            except (ValueError, TypeError):
                pass
                
        return pd.NaT
    
    def _extract_year_value(self, row):
        """Extract year value dengan comprehensive validation."""
        year_columns = ['PERO/AN.1', 'TAHUN', 'YEAR', 'Tahun Perolehan']
        
        for col in year_columns:
            if col in row.index and pd.notna(row[col]):
                try:
                    # Handle various year formats
                    year_str = str(row[col]).strip()
                    
                    # Extract 4-digit year from string
                    year_match = re.search(r'\b(19|20)\d{2}\b', year_str)
                    if year_match:
                        year_val = int(year_match.group())
                    else:
                        year_val = pd.to_numeric(row[col], errors='coerce')
                    
                    # Validate year range
                    if pd.notna(year_val) and 1900 <= year_val <= 2100:
                        return year_val
                except (ValueError, TypeError):
                    continue
        
        return None

    def _extract_month_value(self, row):
        """Extract month value dengan validation."""
        month_columns = ['PERO/AN', 'BULAN', 'MONTH']
        
        for col in month_columns:
            if col in row.index and pd.notna(row[col]):
                try:
                    month_val = pd.to_numeric(row[col], errors='coerce')
                    if pd.notna(month_val) and 1 <= month_val <= 12:
                        return month_val
                except (ValueError, TypeError):
                    continue
        
        return None
    
    def _safe_datetime_conversion(self, value):
        """
        Safe datetime conversion dengan format specification untuk menghindari warnings.
        """
        if pd.isna(value) or value == '':
            return pd.NaT
        
        # If already datetime, return as is
        if isinstance(value, (pd.Timestamp, datetime)):
            return value
        
        # Try common date formats to avoid parsing warnings
        date_formats = [
            '%Y-%m-%d',           # 2024-03-01
            '%Y-%m-%d %H:%M:%S',  # 2024-03-01 00:00:00
            '%d/%m/%Y',           # 01/03/2024
            '%m/%d/%Y',           # 03/01/2024
            '%Y',                 # 2024
            '%Y-%m',              # 2024-03
        ]
        
        value_str = str(value).strip()
        
        # Try each format
        for fmt in date_formats:
            try:
                return pd.to_datetime(value_str, format=fmt, errors='raise')
            except (ValueError, TypeError):
                continue
        
        # Fallback to general parsing (may produce warning but will work)
        try:
            return pd.to_datetime(value_str, errors='coerce')
        except:
            return pd.NaT
    
    def finalize_dataframe_vertical_enhanced(self, df_combined):
        """
        Enhanced finalization with source priority deduplication and QA reporting.
        
        Features:
        1. Barcode expansion using enhanced parser
        2. Source priority deduplication (INV > EXA > MASTER)
        3. Completeness scoring for tie-breaking
        4. Comprehensive QA statistics
        
        Args:
            df_combined (pd.DataFrame): Combined data from all sources with __source column
            
        Returns:
            tuple: (final_dataframe, qa_statistics_dict)
        """
        qa_stats = {
            'Total rows input': len(df_combined),
            'Rows expanded': 0,
            'Duplicates resolved': 0,
            'Rows without barcode kept': 0,
            'Final rows output': 0
        }
        
        final_rows = []
        
        # Process each row for barcode expansion
        for _, row in df_combined.iterrows():
            barcode_str = str(row.get('NO BARCODE', '')).strip()
            barcodes = self.normalize_and_expand_barcodes(barcode_str)
            
            if not barcodes:
                # Keep rows without barcodes if other columns have data
                if any(pd.notna(row[col]) and str(row[col]).strip() not in ('', 'nan', 'NaN') 
                       for col in ['ASSET ORACLE', 'LOKASI', 'JENIS HARTA', 'KONDISI']):
                    final_rows.append(row)
                    qa_stats['Rows without barcode kept'] += 1
                continue
                
            # Expand barcodes to individual rows
            for bc in barcodes:
                new_row = row.copy()
                new_row['NO BARCODE'] = bc
                final_rows.append(new_row)
                qa_stats['Rows expanded'] += 1
        
        if not final_rows:
            return pd.DataFrame(columns=['NO BARCODE','ASSET ORACLE','LOKASI','JENIS HARTA','KONDISI','Tahun Perolehan']), qa_stats
            
        df_final = pd.DataFrame(final_rows)
        
        # Enhanced deduplication with source priority and completeness scoring
        if 'NO BARCODE' in df_final.columns:
            df_final['NO BARCODE'] = df_final['NO BARCODE'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
            
            # Separate rows with and without barcodes
            df_with_bc = df_final[
                df_final['NO BARCODE'].str.lower().ne('nan') & 
                df_final['NO BARCODE'].ne('') & 
                df_final['NO BARCODE'].notna()
            ].copy()
            
            df_without_bc = df_final[
                df_final['NO BARCODE'].str.lower().eq('nan') | 
                df_final['NO BARCODE'].eq('') | 
                df_final['NO BARCODE'].isna()
            ].copy()
            
            if len(df_with_bc) > 0:
                # Calculate completeness score for each row
                df_with_bc['__completeness_score'] = df_with_bc.apply(
                    lambda r: sum(1 for col in ['ASSET ORACLE', 'LOKASI', 'JENIS HARTA', 'KONDISI', 'Tahun Perolehan'] 
                                if pd.notna(r[col]) and str(r[col]).strip() not in ('', 'nan', 'NaN')),
                    axis=1
                )
                
                # Add source priority weight (INV > EXA > MASTER)
                source_weights = {'INV': 0.3, 'EXA': 0.2, 'MASTER': 0.1}
                df_with_bc['__source_weight'] = df_with_bc['__source'].map(source_weights).fillna(0.0)
                
                # Combined score: completeness + source priority
                df_with_bc['__total_score'] = df_with_bc['__completeness_score'] + df_with_bc['__source_weight']
                
                # Count duplicates before deduplication
                duplicates_count = df_with_bc.duplicated(subset=['NO BARCODE']).sum()
                qa_stats['Duplicates resolved'] = duplicates_count
                
                # Keep highest scoring row for each barcode
                df_with_bc = (df_with_bc
                            .sort_values(['NO BARCODE', '__total_score'], ascending=[True, False])
                            .drop_duplicates(subset=['NO BARCODE'], keep='first'))
                
                # Clean up helper columns
                df_with_bc = df_with_bc.drop(columns=['__completeness_score', '__source_weight', '__total_score'])
            
            # Combine back
            df_final = pd.concat([df_with_bc, df_without_bc], ignore_index=True)
        
        # Clean up source column
        if '__source' in df_final.columns:
            df_final = df_final.drop(columns=['__source'])
            
        # Ensure proper datetime format with specific format to avoid warnings
        if 'Tahun Perolehan' in df_final.columns:
            # Try common date formats first to avoid parsing warnings
            df_final['Tahun Perolehan'] = df_final['Tahun Perolehan'].apply(self._safe_datetime_conversion)
        
        # Enforce final schema
        final_columns = ['NO BARCODE','ASSET ORACLE','LOKASI','JENIS HARTA','KONDISI','Tahun Perolehan']
        for col in final_columns:
            if col not in df_final.columns: 
                df_final[col] = ''
        df_final = df_final[final_columns].fillna('')
        
        qa_stats['Final rows output'] = len(df_final)
        
        return df_final, qa_stats

    # ---------- Enhanced Saving & Excel Formatting ----------
    def save_with_formatting(self, df, output_path):
        """
        Enhanced Excel formatting with baseline comparison, highlighting, and multiple sheets.
        
        New Features:
        1. Baseline comparison with diff highlighting (NEW=Yellow, CHANGED=Blue)
        2. Summary & QA sheet with comprehensive statistics
        3. Error/Warning sheet for audit trail
        4. Promote baseline functionality with clean copy
        5. Enhanced Excel formatting and styling
        """
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Define professional styling
        LIGHT_GREEN = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
        LIGHTER_GREEN = PatternFill(start_color="F2F8EE", end_color="F2F8EE", fill_type="solid")
        HEADER_FILL = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        
        # Highlight colors for diff
        HIGHLIGHT_NEW_FILL = PatternFill(start_color=HIGHLIGHT_NEW, end_color=HIGHLIGHT_NEW, fill_type="solid")
        HIGHLIGHT_CHANGED_FILL = PatternFill(start_color=HIGHLIGHT_CHANGED, end_color=HIGHLIGHT_CHANGED, fill_type="solid")
        
        # Professional fonts and alignment
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="Calibri", size=10)
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Professional borders
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        # Clean empty rows before processing
        df_clean = df.copy()
        df_clean = df_clean.replace('', pd.NA)
        df_clean = df_clean.dropna(how='all')
        
        # Keep rows with at least one important column filled
        important_columns = ['NO BARCODE', 'ASSET ORACLE', 'LOKASI', 'JENIS HARTA', 'KONDISI', 'Tahun Perolehan']
        existing_important_cols = [col for col in important_columns if col in df_clean.columns]
        
        if existing_important_cols:
            mask = df_clean[existing_important_cols].notna().any(axis=1)
            df_clean = df_clean[mask]
        
        df_clean = df_clean.fillna('')
        
        self.log_queue.put(f"    • Bersihkan baris kosong: {len(df)} → {len(df_clean)} baris")

        # STEP 1: Load baseline and compute diff flags
        self.log_queue.put("🔍 Computing baseline comparison...")
        baseline_df = self.load_baseline_df()
        diff_flags = self.compute_diff_flags(df_clean, baseline_df)
        
        # Calculate diff statistics
        diff_stats = diff_flags.value_counts().to_dict()
        self.log_queue.put(f"    • NEW: {diff_stats.get('NEW', 0):,} | CHANGED: {diff_stats.get('CHANGED', 0):,} | UNCHANGED: {diff_stats.get('UNCHANGED', 0):,}")
        
        # Add diff flags to DataFrame for highlighting
        df_with_flags = df_clean.copy()
        df_with_flags['__diff_flag'] = diff_flags

        # STEP 2: Save main Excel file with highlighting
        with pd.ExcelWriter(output_path, engine='openpyxl', datetime_format="DD/MM/YYYY") as writer:
            df_clean.to_excel(writer, index=False, sheet_name='Recouncil')
            ws = writer.sheets['Recouncil']

            # Freeze panes at A2 (below header)
            ws.freeze_panes = 'A2'
            
            # Enable AutoFilter on header row
            if ws.max_row > 1:
                ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

            # Format header row
            for cell in ws[1]:
                cell.fill = HEADER_FILL
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = thin_border

            # Format data rows with diff highlighting
            for row_num in range(2, ws.max_row + 1):
                # Get diff flag for this row
                df_row_idx = row_num - 2  # Convert to DataFrame index
                diff_flag = diff_flags.iloc[df_row_idx] if df_row_idx < len(diff_flags) else "UNCHANGED"
                
                # Determine fill color based on diff flag
                if diff_flag == "NEW":
                    fill = HIGHLIGHT_NEW_FILL
                elif diff_flag == "CHANGED":
                    fill = HIGHLIGHT_CHANGED_FILL
                else:
                    # Zebra banding for unchanged rows
                    fill = LIGHT_GREEN if row_num % 2 == 0 else LIGHTER_GREEN
                
                for col_num in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.fill = fill
                    cell.font = data_font
                    cell.alignment = center_alignment
                    cell.border = thin_border

            # Auto-size columns with professional limits
            for col in ws.columns:
                max_length = 0
                column_letter = get_column_letter(col[0].column)
                
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
                
                # Set column width with professional limits
                # Minimum 10, maximum 30 characters
                adjusted_width = min(max(max_length + 2, 10), 30)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Set professional row height
            for row_num in range(1, ws.max_row + 1):
                ws.row_dimensions[row_num].height = 22 if row_num == 1 else 18

            # Format Tahun Perolehan column as date
            tp_col_idx = None
            for j, cell in enumerate(ws[1], start=1):
                if str(cell.value).strip().upper() == "TAHUN PEROLEHAN":
                    tp_col_idx = j
                    break
                    
            if tp_col_idx:
                for r in range(2, ws.max_row + 1):
                    cell = ws.cell(row=r, column=tp_col_idx)
                    if cell.value and pd.notna(cell.value):
                        cell.number_format = "dd/mm/yyyy"
            
            self.log_queue.put(f"    • Recouncil sheet: {len(df_clean):,} rows with diff highlighting")
            
            # STEP 3: Add Summary & QA sheet
            processing_stats = getattr(self, 'processing_stats', {})
            self.write_summary_qa(writer, df_clean, diff_stats, processing_stats)
            
            # STEP 4: Add Error/Warning sheet
            self.write_error_warning(writer)
            
            self.log_queue.put(f"    • Excel file saved with 3 sheets: Recouncil, Summary & QA, Error/Warning")
        
        # STEP 5: Handle Promote Baseline functionality
        if hasattr(self, 'promote_baseline') and self.promote_baseline.get():
            self.log_queue.put("🔄 Promote Baseline enabled - creating clean baseline copy...")
            
            # Create clean baseline filename
            base_name = os.path.splitext(output_path)[0]
            clean_baseline_path = f"{base_name} {CLEAN_BASELINE_SUFFIX}.xlsx"
            
            # Save clean copy without highlighting
            with pd.ExcelWriter(clean_baseline_path, engine='openpyxl', datetime_format="DD/MM/YYYY") as clean_writer:
                df_clean.to_excel(clean_writer, index=False, sheet_name=SHEET_RECOUNCIL)
                
                # Apply basic formatting (no highlighting)
                ws_clean = clean_writer.sheets[SHEET_RECOUNCIL]
                ws_clean.freeze_panes = 'A2'
                
                if ws_clean.max_row > 1:
                    ws_clean.auto_filter.ref = f"A1:{get_column_letter(ws_clean.max_column)}{ws_clean.max_row}"
                
                # Basic formatting without highlights
                for cell in ws_clean[1]:
                    cell.fill = HEADER_FILL
                    cell.font = header_font
                    cell.alignment = center_alignment
                    cell.border = thin_border
                
                for row_num in range(2, ws_clean.max_row + 1):
                    fill = LIGHT_GREEN if row_num % 2 == 0 else LIGHTER_GREEN
                    for col_num in range(1, ws_clean.max_column + 1):
                        cell = ws_clean.cell(row=row_num, column=col_num)
                        cell.fill = fill
                        cell.font = data_font
                        cell.alignment = center_alignment
                        cell.border = thin_border
            
            # Update baseline configuration
            self.save_baseline_config(clean_baseline_path)
            self.log_queue.put(f"    • Clean baseline saved: {os.path.basename(clean_baseline_path)}")
        
        self.log_queue.put(f"✅ Excel formatting completed with enhanced features")

    def save_processing_report(self):
        """Save comprehensive processing report to JSON file."""
        try:
            if not hasattr(self, 'processing_stats'):
                return
                
            report_file = os.path.splitext(self.output_file)[0] + "_processing_report.json"
            
            with open(report_file, 'w', encoding='utf-8') as f:
                json.dump(self.processing_stats, f, indent=2, ensure_ascii=False, default=str)
            
            self.log_queue.put(f"📋 Processing report saved: {os.path.basename(report_file)}")
            
        except Exception as e:
            self.log_queue.put(f"⚠️ Failed to save processing report: {e}")

    # ---------- Baseline & Diff Management ----------
    def load_baseline_df(self, baseline_path=None):
        """
        Load baseline DataFrame from previous output file.
        
        Args:
            baseline_path (str, optional): Path to baseline file. If None, loads from config.
            
        Returns:
            pd.DataFrame or None: Baseline DataFrame or None if not found
        """
        try:
            if baseline_path is None:
                # Load from config
                config_path = os.path.join(CONFIG_DIR, BASELINE_CONFIG_FILE)
                if os.path.exists(config_path):
                    with open(config_path, 'r', encoding='utf-8') as f:
                        config = json.load(f)
                        baseline_path = config.get('baseline_path')
                
                if not baseline_path or not os.path.exists(baseline_path):
                    self.log_queue.put("ℹ️ No baseline found - all records will be marked as NEW")
                    return None
            
            # Read baseline file
            baseline_df = pd.read_excel(baseline_path, sheet_name=SHEET_RECOUNCIL)
            self.log_queue.put(f"📋 Baseline loaded: {os.path.basename(baseline_path)} ({len(baseline_df):,} rows)")
            return baseline_df
            
        except Exception as e:
            self.log_queue.put(f"⚠️ Failed to load baseline: {e}")
            return None
    
    def compute_diff_flags(self, new_df, baseline_df):
        """
        Pure function to compute diff flags between new and baseline DataFrames.
        
        Args:
            new_df (pd.DataFrame): New data
            baseline_df (pd.DataFrame): Baseline data for comparison
            
        Returns:
            pd.Series: Series with values "NEW", "CHANGED", or "UNCHANGED"
        """
        if baseline_df is None or baseline_df.empty:
            # No baseline - all records are NEW
            return pd.Series(["NEW"] * len(new_df), index=new_df.index)
        
        # Ensure JOIN_KEY exists in both DataFrames
        if JOIN_KEY not in new_df.columns or JOIN_KEY not in baseline_df.columns:
            self.log_queue.put(f"⚠️ Join key '{JOIN_KEY}' not found - marking all as NEW")
            return pd.Series(["NEW"] * len(new_df), index=new_df.index)
        
        # Clean and prepare data for comparison
        new_clean = new_df.copy()
        baseline_clean = baseline_df.copy()
        
        # Normalize join key
        new_clean[JOIN_KEY] = new_clean[JOIN_KEY].astype(str).str.strip()
        baseline_clean[JOIN_KEY] = baseline_clean[JOIN_KEY].astype(str).str.strip()
        
        # Create baseline lookup with only comparison columns
        comparison_cols = [col for col in DIFF_COLUMNS if col in baseline_clean.columns]
        baseline_lookup = baseline_clean.set_index(JOIN_KEY)[comparison_cols]
        
        diff_flags = []
        
        for idx, row in new_clean.iterrows():
            barcode = row[JOIN_KEY]
            
            if pd.isna(barcode) or barcode == '' or barcode == 'nan':
                diff_flags.append("NEW")
                continue
            
            if barcode not in baseline_lookup.index:
                # Barcode not in baseline
                diff_flags.append("NEW")
            else:
                # Compare values in comparison columns
                baseline_row = baseline_lookup.loc[barcode]
                is_changed = False
                
                for col in comparison_cols:
                    if col in row.index:
                        new_val = str(row[col]).strip() if pd.notna(row[col]) else ""
                        baseline_val = str(baseline_row[col]).strip() if pd.notna(baseline_row[col]) else ""
                        
                        if new_val != baseline_val:
                            is_changed = True
                            break
                
                diff_flags.append("CHANGED" if is_changed else "UNCHANGED")
        
        return pd.Series(diff_flags, index=new_df.index)
    
    def save_baseline_config(self, baseline_path):
        """Save baseline configuration to JSON file."""
        try:
            os.makedirs(CONFIG_DIR, exist_ok=True)
            config = {
                'baseline_path': baseline_path,
                'last_updated': datetime.now().isoformat(),
                'version': '1.0'
            }
            
            config_path = os.path.join(CONFIG_DIR, BASELINE_CONFIG_FILE)
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=2, ensure_ascii=False)
            
            self.log_queue.put(f"📋 Baseline config saved: {os.path.basename(config_path)}")
            
        except Exception as e:
            self.log_queue.put(f"⚠️ Failed to save baseline config: {e}")
    
    def write_summary_qa(self, writer, df_final, diff_stats, processing_stats):
        """
        Write Summary & QA sheet with comprehensive process information.
        
        Args:
            writer: Excel writer object
            df_final: Final processed DataFrame
            diff_stats: Dictionary with diff statistics
            processing_stats: Dictionary with processing statistics
        """
        try:
            # Prepare summary data
            summary_data = []
            
            # Process Information
            summary_data.extend([
                ["📊 PROCESS INFORMATION", ""],
                ["Execution Time", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                ["Duration", f"{processing_stats.get('total_processing_time', 0):.1f} seconds"],
                ["Application Version", "ICT OPNAME PROCESSOR V4.1"],
                ["Username", os.getenv('USERNAME', 'Unknown')],
                ["Python Version", f"{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}"],
                ["", ""],
            ])
            
            # Source Data Information
            summary_data.extend([
                ["📁 SOURCE DATA", ""],
                ["Master File", os.path.basename(self.master_file) if hasattr(self, 'master_file') else "N/A"],
                ["EXA File", os.path.basename(self.exa_file) if hasattr(self, 'exa_file') else "N/A"],
                ["INV File", os.path.basename(self.inv_file) if hasattr(self, 'inv_file') else "N/A"],
                ["EXA Sheets", ", ".join(self.exa_sheets) if hasattr(self, 'exa_sheets') else "N/A"],
                ["INV Sheets", ", ".join(self.inv_sheets) if hasattr(self, 'inv_sheets') else "N/A"],
                ["BAT Filter", "ICT + BLANKS"],
                ["", ""],
            ])
            
            # Row Counts
            summary_data.extend([
                ["📈 ROW COUNTS", ""],
                ["Master Input", f"{processing_stats.get('master_input_rows', 0):,}"],
                ["EXA Input", f"{processing_stats.get('exa_input_rows', 0):,}"],
                ["INV Input", f"{processing_stats.get('inv_input_rows', 0):,}"],
                ["Total Input", f"{processing_stats.get('total_input_rows', 0):,}"],
                ["Final Output", f"{len(df_final):,}"],
                ["", ""],
            ])
            
            # Diff Breakdown
            summary_data.extend([
                ["🔍 CHANGE ANALYSIS", ""],
                ["New Records (Yellow)", f"{diff_stats.get('NEW', 0):,}"],
                ["Changed Records (Blue)", f"{diff_stats.get('CHANGED', 0):,}"],
                ["Unchanged Records", f"{diff_stats.get('UNCHANGED', 0):,}"],
                ["Total Records", f"{sum(diff_stats.values()):,}"],
                ["", ""],
            ])
            
            # Data Quality
            no_barcode_count = len(df_final[df_final[JOIN_KEY].isna() | (df_final[JOIN_KEY] == '')])
            duplicate_barcodes = len(df_final) - len(df_final[JOIN_KEY].drop_duplicates())
            invalid_dates = len(df_final[df_final['Tahun Perolehan'].isna()]) if 'Tahun Perolehan' in df_final.columns else 0
            
            summary_data.extend([
                ["⚠️ DATA QUALITY", ""],
                ["Missing Barcodes", f"{no_barcode_count:,}"],
                ["Duplicate Barcodes", f"{duplicate_barcodes:,}"],
                ["Invalid Dates", f"{invalid_dates:,}"],
                ["Data Completeness", f"{((len(df_final) - no_barcode_count) / len(df_final) * 100):.1f}%" if len(df_final) > 0 else "0%"],
            ])
            
            # Create DataFrame and write to Excel
            summary_df = pd.DataFrame(summary_data, columns=["Metric", "Value"])
            summary_df.to_excel(writer, sheet_name=SHEET_SUMMARY_QA, index=False)
            
            # Format the sheet
            ws = writer.sheets[SHEET_SUMMARY_QA]
            
            # AutoFit columns
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max(max_length + 2, 15), 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Format headers and sections
            for row_num in range(1, ws.max_row + 1):
                cell_a = ws.cell(row=row_num, column=1)
                cell_b = ws.cell(row=row_num, column=2)
                
                # Section headers (with emojis)
                if cell_a.value and str(cell_a.value).startswith(('📊', '📁', '📈', '🔍', '⚠️')):
                    cell_a.font = Font(bold=True, size=12)
                    cell_a.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                    cell_b.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                
                # Add borders
                cell_a.border = THIN_BORDER
                cell_b.border = THIN_BORDER
                
                # Center alignment for values
                cell_b.alignment = CENTER_ALIGNMENT
            
            self.log_queue.put(f"📊 Summary & QA sheet created with {len(summary_data)} metrics")
            
        except Exception as e:
            self.log_queue.put(f"⚠️ Failed to create Summary & QA sheet: {e}")
    
    def write_error_warning(self, writer):
        """
        Write Error/Warning sheet with all issues encountered during processing.
        
        Args:
            writer: Excel writer object
        """
        try:
            # Prepare error data
            if not hasattr(self, 'error_log') or not self.error_log:
                # Create empty error sheet
                error_data = [
                    ["Severity", "Component", "File", "Sheet", "Row Index", "Message", "Suggestion"],
                    ["INFO", "SYSTEM", "N/A", "N/A", "N/A", "No errors or warnings encountered", "Continue normal operation"]
                ]
            else:
                error_data = [["Severity", "Component", "File", "Sheet", "Row Index", "Message", "Suggestion"]]
                error_data.extend(self.error_log)
            
            # Create DataFrame and write to Excel
            error_df = pd.DataFrame(error_data[1:], columns=error_data[0])
            error_df.to_excel(writer, sheet_name=SHEET_ERROR_WARNING, index=False)
            
            # Format the sheet
            ws = writer.sheets[SHEET_ERROR_WARNING]
            
            # AutoFit columns
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max(max_length + 2, 10), 40)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Format header row
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFEB3B", end_color="FFEB3B", fill_type="solid")
                cell.alignment = CENTER_ALIGNMENT
                cell.border = THIN_BORDER
            
            # Format data rows with conditional coloring
            for row_num in range(2, ws.max_row + 1):
                severity_cell = ws.cell(row=row_num, column=1)
                severity = str(severity_cell.value).upper()
                
                # Color code by severity
                if severity == SEVERITY_ERROR:
                    fill_color = "FFCDD2"  # Light red
                elif severity == SEVERITY_WARN:
                    fill_color = "FFF3E0"  # Light orange
                else:
                    fill_color = "E8F5E8"  # Light green
                
                for col_num in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    cell.border = THIN_BORDER
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            # Freeze panes
            ws.freeze_panes = 'A2'
            
            # AutoFilter
            if ws.max_row > 1:
                ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            
            self.log_queue.put(f"⚠️ Error/Warning sheet created with {len(self.error_log) if hasattr(self, 'error_log') else 0} entries")
            
        except Exception as e:
            self.log_queue.put(f"⚠️ Failed to create Error/Warning sheet: {e}")
    
    def add_error_log(self, severity, component, file_name, sheet_name, row_index, message, suggestion=""):
        """Add error/warning to the error log."""
        if not hasattr(self, 'error_log'):
            self.error_log = []
        
        self.error_log.append([
            severity, component, file_name, sheet_name, row_index, message, suggestion
        ])

    # ---------- Enhanced Features Methods ----------
    def show_rule_builder(self):
        """Tampilkan Rule Builder GUI untuk mapping kolom."""
        rule_window = ttk.Toplevel(self.root)
        rule_window.title("⚙️ Column Mapping Rules")
        rule_window.geometry("700x500")
        
        # Main frame
        main_frame = ttk.Frame(rule_window)
        main_frame.pack(fill=BOTH, expand=True, padx=10, pady=10)
        
        # Title
        title_label = ttk.Label(main_frame, text="Column Mapping Rules Configuration", 
                               font=("Segoe UI", 14, "bold"))
        title_label.pack(pady=(0, 10))
        
        # Current rules display
        rules_frame = ttk.LabelFrame(main_frame, text="Current Rules", padding=10)
        rules_frame.pack(fill=BOTH, expand=True, pady=(0, 10))
        
        # Scrollable text for rules
        rules_text = ScrolledText(rules_frame, height=15)
        rules_text.pack(fill=BOTH, expand=True)
        
        # Display current rules
        rules_content = "Current Column Mapping Rules:\n\n"
        for source, target in self.rule_builder.rules["column_mappings"].items():
            rules_content += f"'{source}' → '{target}'\n"
        
        rules_content += "\nDate Pattern Rules:\n\n"
        for pattern in self.rule_builder.rules["date_patterns"]:
            rules_content += f"Pattern: {pattern}\n"
        
        rules_text.insert('1.0', rules_content)
        rules_text.config(state='disabled')
        
        # Add new rule section
        add_frame = ttk.LabelFrame(main_frame, text="Add New Rule", padding=10)
        add_frame.pack(fill=X, pady=(0, 10))
        
        # Source pattern entry
        ttk.Label(add_frame, text="Source Pattern:").pack(anchor=W)
        source_entry = ttk.Entry(add_frame, width=50)
        source_entry.pack(fill=X, pady=(0, 5))
        
        # Target column entry
        ttk.Label(add_frame, text="Target Column:").pack(anchor=W)
        target_entry = ttk.Entry(add_frame, width=50)
        target_entry.pack(fill=X, pady=(0, 5))
        
        def add_rule():
            source = source_entry.get().strip()
            target = target_entry.get().strip()
            if source and target:
                self.rule_builder.add_mapping_rule(source, target)
                messagebox.showinfo("Success", f"Rule added: '{source}' → '{target}'")
                rule_window.destroy()
                self.show_rule_builder()  # Refresh window
        
        # Buttons
        button_frame = ttk.Frame(add_frame)
        button_frame.pack(fill=X, pady=(10, 0))
        
        ttk.Button(button_frame, text="Add Rule", command=add_rule).pack(side=LEFT)
        ttk.Button(button_frame, text="Close", command=rule_window.destroy).pack(side=RIGHT)
        
        # Center window
        rule_window.transient(self.root)
        rule_window.grab_set()
    
    def show_cache_management(self):
        """Tampilkan Cache Management dialog."""
        cache_window = ttk.Toplevel(self.root)
        cache_window.title("💾 Cache Management")
        cache_window.geometry("500x400")
        
        # Main frame
        main_frame = ttk.Frame(cache_window)
        main_frame.pack(fill=BOTH, expand=True, padx=10, pady=10)
        
        # Title
        title_label = ttk.Label(main_frame, text="Parquet Cache Management", 
                               font=("Segoe UI", 14, "bold"))
        title_label.pack(pady=(0, 10))
        
        # Cache info
        info_frame = ttk.LabelFrame(main_frame, text="Cache Information", padding=10)
        info_frame.pack(fill=X, pady=(0, 10))
        
        # Count cache files
        cache_files = []
        if os.path.exists(CACHE_DIR):
            cache_files = [f for f in os.listdir(CACHE_DIR) if f.endswith('.parquet')]
        
        cache_info = f"Cache Directory: {CACHE_DIR}\n"
        cache_info += f"Cached Files: {len(cache_files)}\n"
        cache_info += f"Parquet Available: {'Yes' if PARQUET_AVAILABLE else 'No'}"
        
        ttk.Label(info_frame, text=cache_info, font=("Consolas", 9)).pack(anchor=W)
        
        # Cache files list
        files_frame = ttk.LabelFrame(main_frame, text="Cached Files", padding=10)
        files_frame.pack(fill=BOTH, expand=True, pady=(0, 10))
        
        files_listbox = tk.Listbox(files_frame, height=8)
        files_listbox.pack(fill=BOTH, expand=True)
        
        for cache_file in cache_files:
            files_listbox.insert(tk.END, cache_file)
        
        # Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=X)
        
        def clear_cache():
            if messagebox.askyesno("Confirm", "Clear all cache files?"):
                success = CacheManager.clear_cache()
                if success:
                    messagebox.showinfo("Success", "Cache cleared successfully")
                    cache_window.destroy()
                else:
                    messagebox.showerror("Error", "Failed to clear cache")
        
        ttk.Button(button_frame, text="Clear All Cache", command=clear_cache).pack(side=LEFT)
        ttk.Button(button_frame, text="Refresh", 
                  command=lambda: [cache_window.destroy(), self.show_cache_management()]).pack(side=LEFT, padx=(5, 0))
        ttk.Button(button_frame, text="Close", command=cache_window.destroy).pack(side=RIGHT)
        
        # Center window
        cache_window.transient(self.root)
        cache_window.grab_set()
    
    def show_recovery_options(self):
        """Tampilkan Recovery Options dialog."""
        recovery_window = ttk.Toplevel(self.root)
        recovery_window.title("🔄 Recovery Options")
        recovery_window.geometry("600x400")
        
        # Main frame
        main_frame = ttk.Frame(recovery_window)
        main_frame.pack(fill=BOTH, expand=True, padx=10, pady=10)
        
        # Title
        title_label = ttk.Label(main_frame, text="Recovery & Checkpoint Management", 
                               font=("Segoe UI", 14, "bold"))
        title_label.pack(pady=(0, 10))
        
        # Checkpoint info
        checkpoint_frame = ttk.LabelFrame(main_frame, text="Checkpoint Information", padding=10)
        checkpoint_frame.pack(fill=X, pady=(0, 10))
        
        checkpoint = CheckpointManager.load_checkpoint()
        if checkpoint:
            checkpoint_info = f"Last Checkpoint: {checkpoint.get('timestamp', 'Unknown')}\n"
            checkpoint_info += f"Stage: {checkpoint.get('stage', 'Unknown')}\n"
            checkpoint_info += f"Status: Available for recovery"
        else:
            checkpoint_info = "No checkpoint available"
        
        ttk.Label(checkpoint_frame, text=checkpoint_info, font=("Consolas", 9)).pack(anchor=W)
        
        # Recovery options
        options_frame = ttk.LabelFrame(main_frame, text="Recovery Actions", padding=10)
        options_frame.pack(fill=BOTH, expand=True, pady=(0, 10))
        
        def recover_from_checkpoint():
            if checkpoint:
                result = messagebox.askyesno("Confirm Recovery", 
                    f"Recover from checkpoint at stage '{checkpoint.get('stage', 'Unknown')}'?")
                if result:
                    # Implement recovery logic here
                    messagebox.showinfo("Recovery", "Recovery feature will be implemented in processing logic")
            else:
                messagebox.showwarning("No Checkpoint", "No checkpoint available for recovery")
        
        def clear_checkpoint():
            if messagebox.askyesno("Confirm", "Clear current checkpoint?"):
                success = CheckpointManager.clear_checkpoint()
                if success:
                    messagebox.showinfo("Success", "Checkpoint cleared successfully")
                    recovery_window.destroy()
                else:
                    messagebox.showerror("Error", "Failed to clear checkpoint")
        
        ttk.Button(options_frame, text="Recover from Checkpoint", 
                  command=recover_from_checkpoint, state=NORMAL if checkpoint else DISABLED).pack(pady=5)
        ttk.Button(options_frame, text="Clear Checkpoint", 
                  command=clear_checkpoint).pack(pady=5)
        
        # Auto-save settings
        settings_frame = ttk.LabelFrame(main_frame, text="Auto-Save Settings", padding=10)
        settings_frame.pack(fill=X, pady=(0, 10))
        
        self.auto_save_enabled = tk.BooleanVar(value=True)
        ttk.Checkbutton(settings_frame, text="Enable automatic checkpoints during processing", 
                       variable=self.auto_save_enabled).pack(anchor=W)
        
        # Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=X)
        
        ttk.Button(button_frame, text="Close", command=recovery_window.destroy).pack(side=RIGHT)
        
        # Center window
        recovery_window.transient(self.root)
        recovery_window.grab_set()
    
    def process_sheet_parallel(self, file_path, sheets_to_process, selected_bats, source_type):
        """
        Process multiple sheets in parallel menggunakan ThreadPoolExecutor.
        
        Args:
            file_path (str): Path ke file Excel
            sheets_to_process (list): List sheet names untuk diproses
            selected_bats (set): BAT filters yang dipilih
            source_type (str): 'EXA' atau 'INV'
            
        Returns:
            pd.DataFrame: Combined hasil dari semua sheets
        """
        if not sheets_to_process:
            return pd.DataFrame()
        
        self.log_queue.put(f"    🚀 Processing {len(sheets_to_process)} sheets in parallel...")
        
        def process_single_sheet(sheet_name):
            """Process single sheet dengan caching."""
            try:
                # Try to load from cache first
                cached_df = CacheManager.load_from_cache(file_path, sheet_name)
                if cached_df is not None:
                    self.log_queue.put(f"      📦 Loaded {sheet_name} from cache")
                    return sheet_name, cached_df, True  # True indicates cache hit
                
                # Process sheet normally
                self.log_queue.put(f"      📊 Processing {sheet_name}...")
                
                # Smart header detection
                header_idx, column_mapping = self.smart_header_detection_file(file_path, sheet_name)
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_idx)
                
                # Apply rule-based column mapping
                if not column_mapping:
                    column_mapping = self.rule_builder.get_column_mapping(df.iloc[0] if len(df) > 0 else pd.Series())
                
                # Rename columns based on mapping
                df = df.rename(columns=column_mapping)
                
                # Apply BAT filtering with cached masks
                combined_mask = self._build_bat_filter_mask(df, selected_bats)
                if combined_mask is not None:
                    df = df[combined_mask]
                
                # Normalize data using dictionaries
                if 'LOKASI' in df.columns:
                    df['LOKASI'] = df['LOKASI'].apply(self.data_dictionary.normalize_location)
                
                if 'JENIS HARTA' in df.columns:
                    df['JENIS HARTA'] = df['JENIS HARTA'].apply(self.data_dictionary.normalize_jenis_harta)
                
                # Enhanced date processing
                if 'Tahun Perolehan' not in df.columns:
                    # Try to extract from available date columns
                    for _, row in df.iterrows():
                        tahun_perolehan = self._extract_tahun_perolehan_enhanced(row)
                        if pd.notna(tahun_perolehan):
                            df.loc[row.name, 'Tahun Perolehan'] = tahun_perolehan
                
                # Save to cache
                CacheManager.save_to_cache(df, file_path, sheet_name)
                
                return sheet_name, df, False  # False indicates fresh processing
                
            except Exception as e:
                self.log_queue.put(f"      ❌ Error processing {sheet_name}: {e}")
                return sheet_name, pd.DataFrame(), False
        
        # Process sheets in parallel
        processed_dfs = []
        cache_hits = 0
        
        with ThreadPoolExecutor(max_workers=min(4, len(sheets_to_process))) as executor:
            # Submit all tasks
            future_to_sheet = {
                executor.submit(process_single_sheet, sheet): sheet 
                for sheet in sheets_to_process
            }
            
            # Collect results
            for future in as_completed(future_to_sheet):
                sheet_name, df, from_cache = future.result()
                if not df.empty:
                    processed_dfs.append(df)
                    if from_cache:
                        cache_hits += 1
        
        # Log performance stats
        self.log_queue.put(f"    ✅ Parallel processing complete: {len(processed_dfs)} sheets, {cache_hits} cache hits")
        
        # Combine all DataFrames
        if processed_dfs:
            combined_df = pd.concat(processed_dfs, ignore_index=True)
            self.log_queue.put(f"    📊 Combined {source_type} data: {len(combined_df)} rows")
            return combined_df
        else:
            return pd.DataFrame()
    
    def smart_header_detection_file(self, file_path, sheet_name):
        """Smart header detection untuk file tertentu."""
        try:
            # Read first few rows to detect header
            df_sample = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=25)
            return self.smart_header_detection(df_sample)
        except Exception as e:
            self.log_queue.put(f"    ⚠️ Header detection failed for {sheet_name}: {e}")
            return 0, {}
    
    def _extract_tahun_perolehan_enhanced(self, row):
        """Enhanced Tahun Perolehan extraction dengan DateNormalizer."""
        # Try existing method first
        result = self._extract_tahun_perolehan(row)
        if pd.notna(result):
            return result
        
        # Try enhanced date normalization
        date_columns = ['TAHUN PEROLEHAN', 'TANGGAL', 'DATE', 'ACQUISITION DATE']
        for col in date_columns:
            if col in row.index and pd.notna(row[col]):
                normalized_date = DateNormalizer.normalize_date(row[col])
                if pd.notna(normalized_date):
                    return normalized_date
        
        return pd.NaT
    
    def run_processing_enhanced(self):
        """
        Enhanced processing dengan parallel processing, caching, dan progress tracking.
        """
        try:
            self.start_time = time.time()
            
            # Save initial checkpoint
            if hasattr(self, 'auto_save_enabled') and self.auto_save_enabled.get():
                CheckpointManager.save_checkpoint('start', {
                    'master_file': self.master_file,
                    'exa_file': self.exa_file,
                    'inv_file': self.inv_file,
                    'timestamp': datetime.now().isoformat()
                })
            
            # Stage 1: Master file processing
            self.current_stage = "master"
            stage_start = time.time()
            self.log_queue.put("[1] 📊 Reading Kamus Data Master...")
            
            df_master = self.read_master_file(self.master_file)
            df_master = self.map_master_to_template(df_master)
            
            self.stage_times['master'] = time.time() - stage_start
            self.log_queue.put(f"    ✅ Master data: {len(df_master)} rows ({self.stage_times['master']:.1f}s)")
            
            # Stage 2: EXA processing with parallel processing
            self.current_stage = "exa"
            stage_start = time.time()
            self.log_queue.put("[2] 📊 Processing EXA files...")
            
            selected_bats_exa = self.get_selected_bats(self.bat_opts_exa)
            df_exa = self.process_sheet_parallel(self.exa_file, self.exa_sheets, selected_bats_exa, 'EXA')
            
            self.stage_times['exa'] = time.time() - stage_start
            self.log_queue.put(f"    ✅ EXA data: {len(df_exa)} rows ({self.stage_times['exa']:.1f}s)")
            
            # Save checkpoint after EXA
            if hasattr(self, 'auto_save_enabled') and self.auto_save_enabled.get():
                CheckpointManager.save_checkpoint('exa_complete', {
                    'master_rows': len(df_master),
                    'exa_rows': len(df_exa)
                })
            
            # Stage 3: INV processing with parallel processing
            self.current_stage = "inv"
            stage_start = time.time()
            self.log_queue.put("[3] 📊 Processing INV files...")
            
            selected_bats_inv = self.get_selected_bats(self.bat_opts_inv)
            df_inv = self.process_sheet_parallel(self.inv_file, self.inv_sheets, selected_bats_inv, 'INV')
            
            self.stage_times['inv'] = time.time() - stage_start
            self.log_queue.put(f"    ✅ INV data: {len(df_inv)} rows ({self.stage_times['inv']:.1f}s)")
            
            # Stage 4: Merge and conflict resolution
            self.current_stage = "merge"
            stage_start = time.time()
            self.log_queue.put("[4] 🔄 Merging data with conflict resolution...")
            
            # Add source tracking
            df_master['__source'] = 'MASTER'
            df_exa['__source'] = 'EXA'
            df_inv['__source'] = 'INV'
            
            df_combined = pd.concat([df_master, df_exa, df_inv], ignore_index=True)
            
            # Conflict resolution if enabled
            if hasattr(self, 'show_conflicts') and self.show_conflicts.get():
                conflicts_count = self.conflict_resolver.detect_conflicts(df_combined)
                if conflicts_count > 0:
                    self.log_queue.put(f"    🔍 Found {conflicts_count} conflicts - showing resolver...")
                    resolutions = self.conflict_resolver.show_conflict_dialog()
                    # Apply resolutions to df_combined
                    # Implementation depends on resolution format
            
            # Enhanced finalization
            df_final, qa_stats = self.finalize_dataframe_vertical_enhanced(df_combined)
            
            self.stage_times['merge'] = time.time() - stage_start
            self.log_queue.put(f"    ✅ Merge complete: {len(df_final)} rows ({self.stage_times['merge']:.1f}s)")
            
            # Log QA statistics
            self.log_queue.put("[QA] 📊 Processing Statistics:")
            for key, value in qa_stats.items():
                self.log_queue.put(f"    • {key}: {value:,}")
            
            # Stage 5: Save with enhanced formatting
            self.current_stage = "save"
            stage_start = time.time()
            self.log_queue.put("[5] 💾 Saving with enhanced formatting...")
            
            self.save_with_formatting(df_final, self.output_file)
            
            self.stage_times['save'] = time.time() - stage_start
            total_time = time.time() - self.start_time
            
            # Performance summary
            self.log_queue.put(f"    ✅ File saved: {os.path.basename(self.output_file)} ({self.stage_times['save']:.1f}s)")
            self.log_queue.put(f"🎉 Processing complete! Total time: {total_time:.1f}s")
            
            # Stage time breakdown
            self.log_queue.put("⏱️ Stage Performance:")
            for stage, duration in self.stage_times.items():
                percentage = (duration / total_time) * 100
                self.log_queue.put(f"    • {stage.upper()}: {duration:.1f}s ({percentage:.1f}%)")
            
            # Clear checkpoint on success
            CheckpointManager.clear_checkpoint()
            
            self.processing_result = {
                'status': 'success',
                'message': f"✅ Processing berhasil!\n\nFile: {os.path.basename(self.output_file)}\nTotal rows: {len(df_final):,}\nTotal time: {total_time:.1f}s\n\nBreakdown:\n- Master: {len(df_master):,} rows\n- EXA: {len(df_exa):,} rows\n- INV: {len(df_inv):,} rows\n\nPerformance:\n- Cache hits: Available\n- Parallel processing: Enabled\n- Conflict resolution: {'Enabled' if hasattr(self, 'show_conflicts') and self.show_conflicts.get() else 'Disabled'}"
            }
            
        except Exception as e:
            # Save error checkpoint
            CheckpointManager.save_checkpoint('error', {
                'error': str(e),
                'stage': self.current_stage,
                'timestamp': datetime.now().isoformat()
            })
            
            self.processing_result = {
                'status': 'error', 
                'message': f"❌ Error during {self.current_stage} stage:\n\n{e}\n\n{traceback.format_exc()}"
            }


#########################################################
# APLIKASI 2: EXTRACT OPNAME
#########################################################
class ExtractOpnameApp(ModernBaseWindow):
    def __init__(self, root, on_close):
        super().__init__(root, "📈 EXTRACT OPNAME", "800x700")
        self.add_back_button(on_close)
        
        # File paths dan sheet selections
        self.target_files = []
        self.target_sheets = {}  # file_path -> sheet_name mapping
        self.scanned_file = ""
        self.scanned_sheet = ""
        self.unscanned_file = ""
        self.unscanned_sheet = ""
        self.processing_result = None
        self.log_queue = Queue()

        # Create modern UI
        self.create_modern_ui()

    def create_modern_ui(self):
        # Description card
        desc_card = self.create_card_frame(
            self.main_frame, 
            "📋 Process Description",
            "Extract and process opname data with intelligent column mapping"
        )
        desc_card.pack(fill=X, pady=(0, 20))
        
        # File selection section
        file_section = self.create_card_frame(self.main_frame, "📁 File Selection")
        file_section.pack(fill=X, pady=(0, 20))
        
        # Target files selection
        target_frame = ttk.Frame(file_section)
        target_frame.pack(fill=X, pady=(0, 15))
        
        btn_target = ttk.Button(
            target_frame, 
            text="📊 1. Pilih File Opname Target",
            bootstyle="secondary",
            command=self.select_target_files
        )
        btn_target.pack(fill=X, pady=(0, 5))
        
        self.lbl_target_status = ttk.Label(
            target_frame, 
            text="❌ Belum ada file yang dipilih",
            font=FONT_CONFIG["text"],
            bootstyle="danger"
        )
        self.lbl_target_status.pack(anchor=W)
        
        # Scanned file selection
        scanned_frame = ttk.Frame(file_section)
        scanned_frame.pack(fill=X, pady=(0, 15))
        
        btn_scanned = ttk.Button(
            scanned_frame, 
            text="✅ 2. Pilih File DATA TERSCAN",
            bootstyle="secondary",
            command=self.select_scanned_file
        )
        btn_scanned.pack(fill=X, pady=(0, 5))
        
        self.lbl_scanned_status = ttk.Label(
            scanned_frame, 
            text="❌ Belum ada file yang dipilih",
            font=FONT_CONFIG["text"],
            bootstyle="danger"
        )
        self.lbl_scanned_status.pack(anchor=W)
        
        # Unscanned file selection
        unscanned_frame = ttk.Frame(file_section)
        unscanned_frame.pack(fill=X, pady=(0, 0))
        
        btn_unscanned = ttk.Button(
            unscanned_frame, 
            text="❌ 3. Pilih File DATA TIDAK TERSCAN",
            bootstyle="secondary",
            command=self.select_unscanned_file
        )
        btn_unscanned.pack(fill=X, pady=(0, 5))
        
        self.lbl_unscanned_status = ttk.Label(
            unscanned_frame, 
            text="❌ Belum ada file yang dipilih",
            font=FONT_CONFIG["text"],
            bootstyle="danger"
        )
        self.lbl_unscanned_status.pack(anchor=W)
        
        # Process button section with enhanced features
        process_section = ttk.Frame(self.main_frame)
        process_section.pack(fill=X, pady=(20, 0))
        
        # Main process button
        self.btn_process = ttk.Button(
            process_section, 
            text="🚀 PROSES FILE",
            bootstyle="success",
            command=self.start_processing,
            state=DISABLED
        )
        self.btn_process.pack(fill=X, pady=(0, 10), ipady=10)
        
        # Additional feature buttons
        feature_frame = ttk.Frame(process_section)
        feature_frame.pack(fill=X, pady=(5, 15))
        
        # Data comparison button
        self.btn_compare = ttk.Button(
            feature_frame,
            text="🔍 Compare Data",
            bootstyle="info-outline",
            command=self.compare_data,
            state=DISABLED,
            width=15
        )
        self.btn_compare.pack(side=LEFT, padx=(0, 5))
        
        # Generate report button
        self.btn_report = ttk.Button(
            feature_frame,
            text="📊 Generate Report",
            bootstyle="warning-outline",
            command=self.generate_report,
            state=DISABLED,
            width=18
        )
        self.btn_report.pack(side=LEFT, padx=(0, 5))
        
        # Quick scan button
        self.btn_quick_scan = ttk.Button(
            feature_frame,
            text="⚡ Quick Scan",
            bootstyle="secondary-outline",
            command=self.quick_scan,
            state=DISABLED,
            width=15
        )
        self.btn_quick_scan.pack(side=LEFT, padx=(0, 5))
        
        # Export summary button
        self.btn_export_summary = ttk.Button(
            feature_frame,
            text="📤 Export Summary",
            bootstyle="primary-outline",
            command=self.export_summary,
            width=18
        )
        self.btn_export_summary.pack(side=RIGHT)
        
        # Add tooltips
        ToolTip(self.btn_process, text="Process opname files with data extraction")
        ToolTip(self.btn_compare, text="Compare scanned vs unscanned data")
        ToolTip(self.btn_report, text="Generate detailed processing report")
        ToolTip(self.btn_quick_scan, text="Quick scan of file contents")
        ToolTip(self.btn_export_summary, text="Export processing summary")

    def button_style(self): return {"bg": STYLE["accent"], "fg": STYLE["text"], "font": STYLE["font_button"], "relief": tk.FLAT, "pady": 8}
    def label_style(self, type="default"):
        color = STYLE["text"] if type == "default" else (STYLE["success"] if type == "success" else STYLE["error"])
        return {"bg": STYLE["background"], "fg": color}
    def check_all_files_selected(self):
        if self.target_files and self.scanned_file and self.unscanned_file: 
            self.btn_process.config(state=NORMAL)
            self.btn_process.config(bootstyle="success")
            # Enable additional feature buttons
            self.btn_compare.config(state=NORMAL)
            self.btn_report.config(state=NORMAL)
            self.btn_quick_scan.config(state=NORMAL)
        else: 
            self.btn_process.config(state=DISABLED)
            self.btn_process.config(bootstyle="secondary-outline")
            # Disable additional feature buttons
            self.btn_compare.config(state=DISABLED)
            self.btn_report.config(state=DISABLED)
            self.btn_quick_scan.config(state=DISABLED)

    def compare_data(self):
        """Compare scanned vs unscanned data"""
        try:
            # Read scanned data
            scanned_df = pd.read_excel(self.scanned_file, sheet_name=self.scanned_sheet if self.scanned_sheet else 0)
            unscanned_df = pd.read_excel(self.unscanned_file, sheet_name=self.unscanned_sheet if self.unscanned_sheet else 0)
            
            # Create comparison window
            compare_window = tk.Toplevel(self.window)
            compare_window.title("🔍 Data Comparison")
            compare_window.geometry("900x700")
            compare_window.transient(self.window)
            
            # Create notebook for comparison tabs
            notebook = ttk.Notebook(compare_window)
            notebook.pack(fill=BOTH, expand=True, padx=10, pady=10)
            
            # Summary tab
            summary_frame = ttk.Frame(notebook)
            notebook.add(summary_frame, text="📊 Summary")
            
            summary_text = scrolledtext.ScrolledText(summary_frame, wrap=tk.WORD)
            summary_text.pack(fill=BOTH, expand=True, padx=5, pady=5)
            
            summary_text.insert(tk.END, "=== DATA COMPARISON SUMMARY ===\n\n")
            summary_text.insert(tk.END, f"Scanned Data:\n")
            summary_text.insert(tk.END, f"  - File: {os.path.basename(self.scanned_file)}\n")
            summary_text.insert(tk.END, f"  - Rows: {len(scanned_df)}\n")
            summary_text.insert(tk.END, f"  - Columns: {len(scanned_df.columns)}\n\n")
            
            summary_text.insert(tk.END, f"Unscanned Data:\n")
            summary_text.insert(tk.END, f"  - File: {os.path.basename(self.unscanned_file)}\n")
            summary_text.insert(tk.END, f"  - Rows: {len(unscanned_df)}\n")
            summary_text.insert(tk.END, f"  - Columns: {len(unscanned_df.columns)}\n\n")
            
            messagebox.showinfo("Comparison", "Data comparison window opened!")
            
        except Exception as e:
            messagebox.showerror("Comparison Error", f"Error comparing data:\n{str(e)}")

    def generate_report(self):
        """Generate detailed processing report"""
        try:
            report_file = filedialog.asksaveasfilename(
                title="Save Report",
                defaultextension=".txt",
                filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
            )
            
            if report_file:
                with open(report_file, 'w', encoding='utf-8') as f:
                    f.write("=== ICT OPNAME PROCESSING REPORT ===\n\n")
                    f.write(f"Generated: {pd.Timestamp.now()}\n\n")
                    
                    f.write("FILES SELECTED:\n")
                    f.write(f"Target Files: {len(self.target_files)} files\n")
                    for i, file in enumerate(self.target_files, 1):
                        f.write(f"  {i}. {os.path.basename(file)}\n")
                    
                    f.write(f"\nScanned File: {os.path.basename(self.scanned_file) if self.scanned_file else 'None'}\n")
                    f.write(f"Unscanned File: {os.path.basename(self.unscanned_file) if self.unscanned_file else 'None'}\n")
                    
                    f.write("\nPROCESSING CONFIGURATION:\n")
                    f.write("Default sheet processing enabled\n")
                    f.write("Target sheets: Recouncil, FORM TEMUAN HASIL OPNAME\n")
                
                messagebox.showinfo("Report Generated", f"Report saved to:\n{report_file}")
        except Exception as e:
            messagebox.showerror("Report Error", f"Error generating report:\n{str(e)}")

    def quick_scan(self):
        """Quick scan of file contents"""
        try:
            scan_results = []
            
            # Scan target files
            for file in self.target_files:
                try:
                    xl = pd.ExcelFile(file)
                    sheets = xl.sheet_names
                    scan_results.append(f"✅ {os.path.basename(file)}: {len(sheets)} sheets")
                except Exception as e:
                    scan_results.append(f"❌ {os.path.basename(file)}: Error - {str(e)}")
            
            # Scan scanned file
            if self.scanned_file:
                try:
                    df = pd.read_excel(self.scanned_file, sheet_name=0)
                    scan_results.append(f"✅ Scanned file: {len(df)} rows")
                except Exception as e:
                    scan_results.append(f"❌ Scanned file: Error - {str(e)}")
            
            # Scan unscanned file
            if self.unscanned_file:
                try:
                    df = pd.read_excel(self.unscanned_file, sheet_name=0)
                    scan_results.append(f"✅ Unscanned file: {len(df)} rows")
                except Exception as e:
                    scan_results.append(f"❌ Unscanned file: Error - {str(e)}")
            
            result_text = "\n".join(scan_results)
            messagebox.showinfo("Quick Scan Results", result_text)
            
        except Exception as e:
            messagebox.showerror("Scan Error", f"Error during quick scan:\n{str(e)}")

    def export_summary(self):
        """Export processing summary"""
        try:
            summary_file = filedialog.asksaveasfilename(
                title="Export Summary",
                defaultextension=".json",
                filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
            )
            
            if summary_file:
                import json
                summary = {
                    "app_name": "EXTRACT OPNAME",
                    "target_files": [os.path.basename(f) for f in self.target_files],
                    "scanned_file": os.path.basename(self.scanned_file) if self.scanned_file else None,
                    "unscanned_file": os.path.basename(self.unscanned_file) if self.unscanned_file else None,
                    "default_sheets": ["Recouncil", "FORM TEMUAN HASIL OPNAME"],
                    "export_timestamp": str(pd.Timestamp.now())
                }
                
                with open(summary_file, 'w') as f:
                    json.dump(summary, f, indent=2)
                
                messagebox.showinfo("Export Complete", f"Summary exported to:\n{summary_file}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Error exporting summary:\n{str(e)}")
    
    def select_target_files(self):
        """
        Pilih file-file opname target. Secara default akan memproses sheet 'Recouncil' dan 'FORM TEMUAN HASIL OPNAME'.
        """
        files = filedialog.askopenfilenames(title="Pilih file-file opname target", filetypes=[("Excel files", "*.xlsx")])
        if files:
            self.target_files = list(files)
            self.target_sheets = {}
            
            # Set default sheets untuk setiap file
            for file_path in self.target_files:
                # Default menggunakan sheet 'Recouncil' dan 'FORM TEMUAN HASIL OPNAME'
                self.target_sheets[file_path] = ['Recouncil', 'FORM TEMUAN HASIL OPNAME']
            
            self.lbl_target_status.config(
                text=f"✅ {len(self.target_files)} file dipilih (Default: Recouncil & Form Temuan)", 
                bootstyle="success"
            )
        self.check_all_files_selected()
    
    def select_scanned_file(self):
        """
        Pilih file DATA TERSCAN. Secara default akan menggunakan sheet pertama.
        """
        file = filedialog.askopenfilename(title="Pilih file DATA TERSCAN", filetypes=[("Excel files", "*.xlsx")])
        if file:
            self.scanned_file = file
            self.scanned_sheet = None  # Akan menggunakan sheet default (pertama)
            self.lbl_scanned_status.config(
                text=f"✅ {os.path.basename(file)} (Default Sheet)", 
                bootstyle="success"
            )
        self.check_all_files_selected()
    
    def select_unscanned_file(self):
        """
        Pilih file DATA TIDAK TERSCAN. Secara default akan menggunakan sheet pertama.
        """
        file = filedialog.askopenfilename(title="Pilih file DATA TIDAK TERSCAN", filetypes=[("Excel files", "*.xlsx")])
        if file:
            self.unscanned_file = file
            self.unscanned_sheet = None  # Akan menggunakan sheet default (pertama)
            self.lbl_unscanned_status.config(
                text=f"✅ {os.path.basename(file)} (Default Sheet)", 
                bootstyle="success"
            )
        self.check_all_files_selected()
    
    def start_processing(self):
        self.show_loading_screen()
        self.processing_result = None
        self.processing_thread = threading.Thread(target=self.run_processing)
        self.processing_thread.daemon = True
        self.processing_thread.start()
        self.check_thread()

    def check_thread(self):
        if self.processing_thread.is_alive():
            self.root.after(100, self.check_thread)
        else:
            self.close_loading_screen()
            if self.processing_result:
                self.show_summary(self.processing_result)

    def show_loading_screen(self):
        # Create modern loading window with proper visibility
        self.loading_window = ttk.Toplevel(self.window)
        self.loading_window.title("⏳ Processing - Extract Opname")
        self.loading_window.geometry("450x250")
        self.loading_window.transient(self.window)
        self.loading_window.grab_set()
        
        # Center the window and bring to front
        self.loading_window.update_idletasks()
        x = (self.loading_window.winfo_screenwidth() // 2) - (self.loading_window.winfo_width() // 2)
        y = (self.loading_window.winfo_screenheight() // 2) - (self.loading_window.winfo_height() // 2)
        self.loading_window.geometry(f"+{x}+{y}")
        
        # Force window to front
        self.loading_window.lift()
        self.loading_window.focus_force()
        self.loading_window.attributes('-topmost', True)
        
        # Main frame
        main_frame = ttk.Frame(self.loading_window)
        main_frame.pack(fill=BOTH, expand=True, padx=30, pady=30)
        
        # Loading icon and text
        loading_label = ttk.Label(
            main_frame,
            text="⏳ Sedang Memproses...",
            font=("Segoe UI", 14, "bold"),
            bootstyle="primary"
        )
        loading_label.pack(pady=(10, 15))
        
        # Progress bar
        self.progress_bar = ttk.Progressbar(
            main_frame,
            mode='indeterminate',
            bootstyle="success-striped"
        )
        self.progress_bar.pack(fill=X, pady=(0, 15))
        self.progress_bar.start(10)
        
        # Status text
        status_label = ttk.Label(
            main_frame,
            text="Harap tunggu, sedang memproses file opname Anda...",
            font=("Segoe UI", 10),
            bootstyle="secondary"
        )
        status_label.pack(pady=(0, 10))
        
        # Make window non-resizable
        self.loading_window.resizable(False, False)
    
    def close_loading_screen(self):
        if hasattr(self, 'loading_window') and self.loading_window.winfo_exists(): 
            if hasattr(self, 'progress_bar'):
                self.progress_bar.stop()
            self.loading_window.destroy()

    def run_processing(self):
        results = {"success": [], "no_data": [], "error": []}
        try:
            # APP 2 FIX: Baca file scanned dengan mapping kolom yang benar
            scanned_df = pd.read_excel(self.scanned_file)
            
            # APP 2 FIX: Mapping kolom yang benar dari file scanned
            # Kolom B1 = Oracle Asset, Kolom D1 = No PO
            column_mapping = {}
            if 'B' not in scanned_df.columns and len(scanned_df.columns) > 1:
                # Jika kolom tidak bernama B, gunakan index kolom
                scanned_df.columns = [f'Col_{i}' for i in range(len(scanned_df.columns))]
                column_mapping = {
                    'Oracle Asset': 'Col_1',  # Kolom B (index 1)
                    'No PO': 'Col_3',         # Kolom D (index 3)
                    'Barcode': 'Col_2',       # Kolom C (index 2)
                    'Nama_Asset': 'Col_4',    # Kolom E (index 4)
                    'Ruangan_Opname': 'Col_5', # Kolom F (index 5)
                    'Ruangan_Barcode': 'Col_6', # Kolom G (index 6)
                    'Kondisi': 'Col_7',       # Kolom H (index 7)
                    'Keterangan': 'Col_8'     # Kolom I (index 8)
                }
                # Rename kolom sesuai mapping
                for new_name, old_name in column_mapping.items():
                    if old_name in scanned_df.columns:
                        scanned_df = scanned_df.rename(columns={old_name: new_name})
            
            unscanned_data = self.parse_unscanned_file(self.unscanned_file)
            due_date = self.calculate_due_date()
            for target_file in self.target_files:
                try:
                    wb = openpyxl.load_workbook(target_file)
                    if 'Recouncil' not in wb.sheetnames: raise ValueError("Sheet 'Recouncil' tidak ditemukan.")
                    ws = wb['Recouncil']
                    if ws.max_row >= 8: ws.delete_rows(8, ws.max_row - 7)
                    prep_sheet = wb['PERSIAPAN']
                    room_code_raw = str(prep_sheet['A2'].value)
                    match = re.search(r'(\d{3}[-–—].*?SJA\s*\d+)', room_code_raw)
                    if not match: raise ValueError(f"Format kode ruangan tidak ditemukan di sel A2: '{room_code_raw}'")
                    room_code = match.group(1).strip()
                    filtered_scanned = scanned_df[scanned_df['Ruangan_Opname'] == room_code].copy()
                    filtered_unscanned = unscanned_data.get(room_code, pd.DataFrame())
                    if filtered_scanned.empty and filtered_unscanned.empty:
                        results["no_data"].append(f"{os.path.basename(target_file)} (Kode: {room_code})")
                        continue
                    temuan_data = []
                    current_row = 8
                    for _, row in filtered_scanned.iterrows():
                        keterangan_lower = str(row.get('Keterangan', '')).lower()
                        action_text = ''
                        if "salah ruangan" in keterangan_lower:
                            ruangan_asal = row.get('Ruangan_Barcode', 'TIDAK DIKETAHUI')
                            action_text = f"Kordinasikan Dengan PIC Ruangan ({ruangan_asal}) untuk membuat MAT ke ruangan ({room_code})"
                        elif "double barcode" in keterangan_lower or "cetak ulang" in keterangan_lower:
                            action_text = "CETAK ULANG BARCODE"
                        data_to_append = [current_row - 7, row.get('Oracle Asset', ''), row.get('Barcode', ''), row.get('No PO', ''), row.get('Nama_Asset', ''), row.get('Ruangan_Opname', ''), row.get('Ruangan_Barcode', ''), row.get('Kondisi', ''), row.get('Keterangan', ''), action_text]
                        ws.append(data_to_append)
                        self.apply_row_formatting(ws, current_row, row.get('Keterangan', ''))
                        if "salah ruangan" in keterangan_lower or "double barcode" in keterangan_lower or "cetak ulang" in keterangan_lower:
                            temuan_data.append({"barcode": row.get('Barcode', ''), "nama_asset": row.get('Nama_Asset', ''), "keterangan": row.get('Keterangan', ''), "tipe": keterangan_lower})
                        current_row += 1
                    for _, row in filtered_unscanned.iterrows():
                        data_to_append = [current_row - 7, row.get('ORACLE_ASSET_ID', ''), row.get('BARCODE_ASSET', ''), row.get('NO_PO', ''), row.get('NAMA_ASSET', ''), room_code, room_code, '', 'DATA TIDAK TERSCAN', 'EDIT SESUAI KONDISI SEKARANG DARI ICT']
                        ws.append(data_to_append)
                        self.apply_row_formatting(ws, current_row, 'DATA TIDAK TERSCAN', is_unscanned=True)
                        temuan_data.append({"barcode": row.get('BARCODE_ASSET', ''), "nama_asset": row.get('NAMA_ASSET', ''), "keterangan": 'DATA TIDAK TERSCAN', "tipe": 'data tidak terscan'})
                        current_row += 1
                    self.finalize_sheet_formatting(ws, start_row=8)
                    if temuan_data: self.process_form_temuan(wb, temuan_data, due_date)
                    wb.save(target_file)
                    results["success"].append(os.path.basename(target_file))
                except Exception as e:
                    results["error"].append(f"{os.path.basename(target_file)}:\n{e}\n{traceback.format_exc()}")
            self.processing_result = results
        except Exception as e:
             self.processing_result = {"success": [], "no_data": [], "error": [f"Error fatal saat memuat file sumber:\n{e}\n{traceback.format_exc()}"]}

    def parse_unscanned_file(self, filepath):
        df = pd.read_excel(filepath, header=None).fillna('')
        data_dict, current_room, room_data = {}, None, []
        for _, row in df.iterrows():
            row_str = ' '.join(str(c) for c in row)
            if 'NAMA_RUANGAN:' in row_str.upper():
                if current_room and room_data: data_dict[current_room] = pd.DataFrame(room_data, columns=['BARCODE_ASSET', 'ORACLE_ASSET_ID', 'NO_PO', 'NAMA_ASSET'])
                current_room = row_str.split(':', 1)[1].strip()
                room_data = []
            elif current_room:
                barcode = str(row.iloc[1]) if len(row) > 1 else ''
                if barcode.strip(): room_data.append([barcode, str(row.iloc[2]), str(row.iloc[3]), str(row.iloc[4])])
        if current_room and room_data: data_dict[current_room] = pd.DataFrame(room_data, columns=['BARCODE_ASSET', 'ORACLE_ASSET_ID', 'NO_PO', 'NAMA_ASSET'])
        return data_dict
    def calculate_due_date(self):
        today = date.today()
        work_days_to_add = 7
        current_date = today
        while work_days_to_add > 0:
            current_date += timedelta(days=1)
            if current_date.weekday() < 5: work_days_to_add -= 1
        return current_date.strftime('%d %B %Y')
    def process_form_temuan(self, wb, temuan_data, due_date):
        if 'FORM TEMUAN HASIL OPNAME' not in wb.sheetnames: return
        ws = wb['FORM TEMUAN HASIL OPNAME']
        if ws.max_row >= 7: ws.delete_rows(7, ws.max_row - 6)
        current_temuan_row = 7
        for index, temuan in enumerate(temuan_data, start=1):
            ws.cell(row=current_temuan_row, column=1).value = index
            ws.cell(row=current_temuan_row, column=2).value = temuan["barcode"]
            ws.cell(row=current_temuan_row, column=3).value = temuan["nama_asset"]
            ws.cell(row=current_temuan_row, column=9).value = temuan["keterangan"]
            ws.cell(row=current_temuan_row, column=10).value = due_date
            tipe = temuan["tipe"]
            fill_color = None
            if "data tidak terscan" in tipe: fill_color = YELLOW_FILL
            elif "salah ruangan" in tipe: fill_color = BLUE_FILL
            elif "double barcode" in tipe or "cetak ulang" in tipe: fill_color = PURPLE_FILL
            if fill_color:
                for col_idx in range(1, 12): ws.cell(row=current_temuan_row, column=col_idx).fill = fill_color
            current_temuan_row += 1
        self.finalize_form_temuan_formatting(ws, start_row=7)
    def finalize_form_temuan_formatting(self, ws, start_row):
        if ws.max_row < start_row: return
        for row in ws.iter_rows(min_row=start_row, max_col=11, max_row=ws.max_row):
            for cell in row:
                cell.font = DEFAULT_FONT_EXCEL
                cell.border = THIN_BORDER
                cell.alignment = CENTER_ALIGNMENT
    def apply_row_formatting(self, ws, row_index, keterangan_text, is_unscanned=False):
        keterangan_lower = str(keterangan_text).lower()
        fill_color = None
        if is_unscanned: fill_color = YELLOW_FILL
        elif "salah ruangan" in keterangan_lower: fill_color = BLUE_FILL
        elif "double barcode" in keterangan_lower or "cetak ulang" in keterangan_lower: fill_color = PURPLE_FILL
        if fill_color:
            for col_idx in range(1, 11): ws.cell(row=row_index, column=col_idx).fill = fill_color
    def finalize_sheet_formatting(self, ws, start_row):
        for row in ws.iter_rows(min_row=start_row, max_col=10):
            for cell in row:
                cell.font = DEFAULT_FONT_EXCEL
                cell.border = THIN_BORDER
                cell.alignment = CENTER_ALIGNMENT
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            max_length = 0
            for cell in ws[col_letter]:
                try: max_length = max(max_length, len(str(cell.value)))
                except: pass
            ws.column_dimensions[col_letter].width = max_length + 2
        ws.column_dimensions['J'].width = 50.40
    def show_summary(self, results):
        # Create modern summary window with proper visibility
        summary_window = ttk.Toplevel(self.window)
        summary_window.title("📊 Ringkasan Pemrosesan - Extract Opname")
        summary_window.geometry("800x600")
        summary_window.transient(self.window)
        summary_window.grab_set()
        
        # Center the window and bring to front
        summary_window.update_idletasks()
        x = (summary_window.winfo_screenwidth() // 2) - (summary_window.winfo_width() // 2)
        y = (summary_window.winfo_screenheight() // 2) - (summary_window.winfo_height() // 2)
        summary_window.geometry(f"+{x}+{y}")
        
        # Force window to front
        summary_window.lift()
        summary_window.focus_force()
        summary_window.attributes('-topmost', True)
        summary_window.after(100, lambda: summary_window.attributes('-topmost', False))
        
        # Main frame
        main_frame = ttk.Frame(summary_window)
        main_frame.pack(fill=BOTH, expand=True, padx=20, pady=20)
        
        # Title
        title_label = ttk.Label(
            main_frame, 
            text="📊 Ringkasan Pemrosesan", 
            font=("Segoe UI", 16, "bold"),
            bootstyle="primary"
        )
        title_label.pack(pady=(0, 20))
        
        # Results frame with scrollable text
        results_frame = ttk.LabelFrame(main_frame, text="Hasil Processing", padding=15)
        results_frame.pack(fill=BOTH, expand=True, pady=(0, 20))
        
        # Scrollable text area
        text_area = scrolledtext.ScrolledText(
            results_frame, 
            wrap=tk.WORD, 
            font=("Consolas", 10),
            height=20
        )
        text_area.pack(fill=BOTH, expand=True)
        
        # Format results
        summary_text = "=== RINGKASAN PEMROSESAN EXTRACT OPNAME ===\n\n"
        
        if results["success"]: 
            summary_text += "✅ BERHASIL DIPROSES:\n"
            for item in results["success"]:
                summary_text += f"  • {item}\n"
            summary_text += "\n"
            
        if results["no_data"]: 
            summary_text += "⚠️ DATA TIDAK DITEMUKAN:\n"
            for item in results["no_data"]:
                summary_text += f"  • {item}\n"
            summary_text += "\n"
            
        if results["error"]: 
            summary_text += "❌ GAGAL / ERROR:\n"
            for item in results["error"]:
                summary_text += f"  • {item}\n"
            summary_text += "\n"
        
        # Add statistics
        total_success = len(results.get("success", []))
        total_no_data = len(results.get("no_data", []))
        total_error = len(results.get("error", []))
        total_files = total_success + total_no_data + total_error
        
        summary_text += f"=== STATISTIK ===\n"
        summary_text += f"Total File: {total_files}\n"
        summary_text += f"Berhasil: {total_success}\n"
        summary_text += f"Tidak Ada Data: {total_no_data}\n"
        summary_text += f"Error: {total_error}\n"
        
        text_area.insert(tk.END, summary_text)
        text_area.config(state=tk.DISABLED)
        
        # Button frame
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=X, pady=(10, 0))
        
        # Export button
        export_btn = ttk.Button(
            button_frame,
            text="📤 Export Report",
            bootstyle="info-outline",
            command=lambda: self.export_processing_report(results)
        )
        export_btn.pack(side=LEFT)
        
        # Close button - LARGE AND VISIBLE
        close_btn = ttk.Button(
            button_frame,
            text="✅ OK - Tutup",
            bootstyle="success",
            command=summary_window.destroy
        )
        close_btn.pack(side=RIGHT, ipadx=20, ipady=5)
        
        # Focus on close button
        close_btn.focus_set()
        
        # Bind Enter key to close
        summary_window.bind('<Return>', lambda e: summary_window.destroy())
        summary_window.bind('<Escape>', lambda e: summary_window.destroy())

    def export_processing_report(self, results):
        """Export processing report to file"""
        try:
            from datetime import datetime
            report_file = filedialog.asksaveasfilename(
                title="Export Processing Report",
                defaultextension=".txt",
                filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
            )
            
            if report_file:
                with open(report_file, 'w', encoding='utf-8') as f:
                    f.write("=== ICT OPNAME PROCESSING REPORT ===\n")
                    f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                    f.write(f"Application: EXTRACT OPNAME\n\n")
                    
                    if results["success"]:
                        f.write("BERHASIL DIPROSES:\n")
                        for item in results["success"]:
                            f.write(f"  • {item}\n")
                        f.write("\n")
                    
                    if results["no_data"]:
                        f.write("DATA TIDAK DITEMUKAN:\n")
                        for item in results["no_data"]:
                            f.write(f"  • {item}\n")
                        f.write("\n")
                    
                    if results["error"]:
                        f.write("GAGAL / ERROR:\n")
                        for item in results["error"]:
                            f.write(f"  • {item}\n")
                        f.write("\n")
                
                messagebox.showinfo("Export Complete", f"Report exported to:\n{report_file}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Error exporting report:\n{str(e)}")



#########################################################
# APLIKASI 3: EXTRACT MAT
#########################################################
class ExtractMATApp(ModernBaseWindow):
    def __init__(self, root, on_close):
        super().__init__(root, "📋 EXTRACT MAT", "800x750")
        self.add_back_button(on_close)

        self.pic_db_path = ""
        self.mat_template_path = self._find_default_mat_template()
        self.target_files = []
        self.ruangan_asal_map = {}  # barcode -> ruangan asal
        self.pic_map = {}  # barcode -> PIC

        # Create modern UI
        self.create_modern_ui()
        
        self.log_queue = Queue()

    def create_modern_ui(self):
        # Description card
        desc_card = self.create_card_frame(
            self.main_frame, 
            "📋 Process Description",
            "Extract MAT data from opname files with PIC database integration and template generation"
        )
        desc_card.pack(fill=X, pady=(0, 20))
        
        # File selection section
        file_section = self.create_card_frame(self.main_frame, "📁 File Selection")
        file_section.pack(fill=X, pady=(0, 20))
        
        # Database PIC selection
        db_frame = ttk.Frame(file_section)
        db_frame.pack(fill=X, pady=(0, 15))
        
        btn_db = ttk.Button(
            db_frame, 
            text="🗃️ 1. Pilih Database PIC (Asset Management)",
            bootstyle="secondary",
            command=self.select_pic_db
        )
        btn_db.pack(fill=X, pady=(0, 5))
        
        self.lbl_db = ttk.Label(
            db_frame, 
            text="❌ Belum ada file database PIC dipilih",
            font=FONT_CONFIG["text"],
            bootstyle="danger"
        )
        self.lbl_db.pack(anchor=W)
        
        # Target files selection
        targets_frame = ttk.Frame(file_section)
        targets_frame.pack(fill=X, pady=(0, 15))
        
        btn_targets = ttk.Button(
            targets_frame, 
            text="📊 2. Pilih FILE yang ingin dibuatkan MAT (boleh banyak)",
            bootstyle="secondary",
            command=self.select_target_files
        )
        btn_targets.pack(fill=X, pady=(0, 5))
        
        self.lbl_targets = ttk.Label(
            targets_frame, 
            text="❌ Belum ada file ruangan dipilih",
            font=FONT_CONFIG["text"],
            bootstyle="danger"
        )
        self.lbl_targets.pack(anchor=W)
        
        # Template selection
        template_frame = ttk.Frame(file_section)
        template_frame.pack(fill=X, pady=(0, 0))
        
        template_label_frame = ttk.Frame(template_frame)
        template_label_frame.pack(fill=X, pady=(0, 5))
        
        ttk.Label(
            template_label_frame, 
            text="📄 Template MAT:",
            font=FONT_CONFIG["text"]
        ).pack(side=LEFT)
        
        self.lbl_template = ttk.Label(
            template_label_frame, 
            text=f"✅ {os.path.basename(self.mat_template_path)}" if self.mat_template_path else "❌ Tidak ditemukan",
            font=FONT_CONFIG["text"],
            bootstyle="success" if self.mat_template_path else "danger"
        )
        self.lbl_template.pack(side=LEFT, padx=(10, 0))
        
        btn_template = ttk.Button(
            template_label_frame, 
            text="📂 Pilih Template",
            bootstyle="secondary-outline",
            command=self.select_template
        )
        btn_template.pack(side=RIGHT)
        
        # Process button section
        process_section = ttk.Frame(self.main_frame)
        process_section.pack(fill=X, pady=(20, 0))
        
        self.btn_process = ttk.Button(
            process_section, 
            text="🚀 PROSES & BUAT FILE MAT",
            bootstyle="success",
            command=self.start_processing,
            state=DISABLED
        )
        self.btn_process.pack(fill=X, pady=15, ipady=10)
        
        # Add tooltip
        ToolTip(self.btn_process, text="Process files to generate MAT data")
        
        # Log section
        log_section = self.create_card_frame(self.main_frame, "📝 Processing Log")
        log_section.pack(fill=BOTH, expand=True, pady=(0, 0))
        
        # Modern scrolled text
        self.log = ScrolledText(
            log_section, 
            wrap=WORD, 
            height=12,
            bootstyle="secondary"
        )
        self.log.pack(fill=BOTH, expand=True, padx=5, pady=5)

    # --- helpers UI ---
    def button_style(self):
        return {"bg": STYLE["accent"], "fg": STYLE["text"], "font": STYLE["font_button"], "relief": tk.FLAT, "pady": 8}
    def label_style(self, type="default"):
        # Return ttkbootstrap compatible style
        if type == "success":
            return {"bootstyle": "success"}
        elif type == "error":
            return {"bootstyle": "danger"}
        else:
            return {"bootstyle": "secondary"}

    def _find_default_mat_template(self):
        try:
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(".")
            
        candidates = ["MAT.xlsx", os.path.join(base_path, "MAT.xlsx")]
        for c in candidates:
            if os.path.exists(c):
                return c
        return ""

    def select_template(self):
        f = filedialog.askopenfilename(title="Pilih Template MAT", filetypes=[("Excel files", "*.xlsx")])
        if f:
            self.mat_template_path = f
            self.lbl_template.config(text=os.path.basename(f), **self.label_style("success"))
        self._refresh_process_button_state()

    def select_pic_db(self):
        f = filedialog.askopenfilename(title="Pilih Database PIC (Asset Management)", filetypes=[("Excel files", "*.xlsx *.xls")])
        if f:
            # Secara default menggunakan sheet pertama
            self.pic_db_path = f
            self.pic_db_sheet = None  # Akan menggunakan sheet default (pertama)
            self.lbl_db.config(
                text=f"✅ {os.path.basename(f)} (Default Sheet)", 
                bootstyle="success"
            )
            try:
                self.ruangan_asal_map, self.pic_map = self._build_master_data_maps(f, None)
                self.append_log(f"[OK] Master data dimuat. Total {len(self.pic_map)} mapping barcode → ruangan & PIC.")
            except Exception as e:
                self.append_log(f"[ERROR] Gagal memuat Master Data: {e}")
        self._refresh_process_button_state()

    def select_target_files(self):
        files = filedialog.askopenfilenames(title="Pilih file-file ruangan (boleh >10 sekaligus)", filetypes=[("Excel files", "*.xlsx")])
        if files:
            self.target_files = list(files)
            self.lbl_targets.config(text=f"{len(self.target_files)} file dipilih.", **self.label_style("success"))
        self._refresh_process_button_state()

    def _refresh_process_button_state(self):
        if self.pic_db_path and self.target_files and self.mat_template_path:
            self.btn_process.config(state=NORMAL)
            self.btn_process.config(bootstyle="success")
        else:
            self.btn_process.config(state=DISABLED)
            self.btn_process.config(bootstyle="secondary-outline")

    def append_log(self, msg):
        self.log.insert(END, msg + "\n")
        self.log.see(END)
        self.log.update_idletasks()

    # --- processing entry ---
    def start_processing(self):
        if not (self.pic_db_path and self.target_files):
            messagebox.showwarning("Data belum lengkap", "Pilih database PIC dan file-file ruangan terlebih dahulu.")
            return
        save_path = filedialog.asksaveasfilename(title="Simpan Hasil MAT Sebagai...", defaultextension=".xlsx",
                                                 initialfile="HASIL_MAT.xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not save_path:
            return
        try:
            rows = self.extract_all()
            if not rows:
                messagebox.showinfo("Tidak Ada Data", "Tidak ditemukan baris 'salah ruangan' pada file-file yang dipilih.")
                return
            self.write_to_mat_template(rows, save_path)
            messagebox.showinfo("Selesai", f"File MAT berhasil dibuat:\n{os.path.basename(save_path)}\nTotal baris: {len(rows)}")
        except Exception as e:
            messagebox.showerror("Error", f"Gagal membuat file MAT.\n\n{e}\n\n{traceback.format_exc()}")

    # --- core logic ---
    def _build_master_data_maps(self, db_path, sheet_name=None):
        """Build mapping dictionaries from master data file for barcode -> ruangan asal and barcode -> PIC"""
        import pandas as pd
        
        try:
            df = pd.read_excel(db_path, sheet_name=sheet_name if sheet_name else 0)
            self.append_log(f"[INFO] Master data berhasil dibaca. Jumlah baris: {len(df)}")
            
            # Clean column names
            df.columns = [col.strip() for col in df.columns]
            self.append_log(f"[INFO] Kolom tersedia: {list(df.columns)}")
            
            # Find required columns
            col_barcode = None
            col_ruang = None
            col_pic = None
            
            for c in df.columns:
                uc = str(c).strip().upper()
                # Cari kolom barcode
                if 'BARCODE' in uc and 'ASSET' in uc:
                    col_barcode = c
                    self.append_log(f"[INFO] Kolom barcode ditemukan: {c}")
                # Cari kolom lokasi/ruangan
                if 'LOKASI' in uc and 'ASSET' in uc:
                    col_ruang = c
                    self.append_log(f"[INFO] Kolom lokasi ditemukan: {c}")
                # Cari kolom PIC
                if 'PIC' in uc and 'RUANGAN' in uc:
                    col_pic = c
                    self.append_log(f"[INFO] Kolom PIC ditemukan: {c}")
            
            # Validate required columns found
            if not all([col_barcode, col_ruang, col_pic]):
                error_msg = f"Kolom yang diperlukan tidak ditemukan dalam Master Data.\n"
                error_msg += f"Kolom Barcode: {col_barcode}\n"
                error_msg += f"Kolom Ruangan: {col_ruang}\n"
                error_msg += f"Kolom PIC: {col_pic}\n"
                error_msg += f"Kolom tersedia: {list(df.columns)}"
                raise ValueError(error_msg)
            
            # Convert barcode to string for reliable matching
            df[col_barcode] = df[col_barcode].astype(str)
            
            # Remove duplicates, keep first occurrence
            df = df.drop_duplicates(subset=[col_barcode], keep='first')
            
            # Build mapping dictionaries: barcode -> ruangan asal, barcode -> PIC
            ruangan_asal_map = pd.Series(df[col_ruang].values, index=df[col_barcode]).to_dict()
            pic_map = pd.Series(df[col_pic].values, index=df[col_barcode]).to_dict()
            
            self.append_log(f"[INFO] Mapping berhasil dibuat. {len(ruangan_asal_map)} entri unik.")
            
            # Log sample mappings for debugging
            sample_barcodes = list(ruangan_asal_map.keys())[:3]
            for barcode in sample_barcodes:
                ruangan = ruangan_asal_map.get(barcode, 'N/A')
                pic = pic_map.get(barcode, 'N/A')
                self.append_log(f"[DEBUG] Barcode: {barcode} -> Ruangan: {ruangan}, PIC: {pic}")
            
            return ruangan_asal_map, pic_map
            
        except Exception as e:
            self.append_log(f"[ERROR] Gagal membangun mapping master data: {str(e)}")
            raise e

    def _norm_room(self, s):
        s = str(s).strip().upper().replace('–','-').replace('—','-')
        return re.sub(r'\s+', ' ', s)

    def _parse_header_pic_from_recouncil(self, ws):
        for r in range(1, 7):
            v = ws.cell(row=r, column=1).value
            if isinstance(v, str) and "PIC" in v.upper():
                txt = v.split(':', 1)[-1].strip()
                return txt
        return ""

    def _find_recouncil_header_row(self, ws):
        for r in range(1, 20):
            vals = [ws.cell(row=r, column=c).value for c in range(1, 11)]
            row_str = ' '.join(str(v).upper() for v in vals if v is not None)
            if ("NO" in row_str and "BARCODE" in row_str and "ORACLE" in row_str and "KETERANGAN" in row_str):
                return r
        return None

    def extract_all(self):
        from openpyxl import load_workbook
        rows = []
        processed_barcodes = set()  # Track barcode yang sudah diproses untuk menghindari duplikasi
        
        for file in self.target_files:
            try:
                wb = load_workbook(file, data_only=True)
            except Exception as e:
                self.append_log(f"[SKIP] {os.path.basename(file)} gagal dibuka: {e}")
                continue

            # HANYA PROSES SHEET RECOUNCIL - sesuai permintaan user
            if 'Recouncil' in wb.sheetnames:
                ws = wb['Recouncil']
                self.append_log(f"[INFO] Memproses sheet Recouncil dari file: {os.path.basename(file)}")
                
                # Cari header row secara dinamis
                header_row = None
                for r in range(1, 20):
                    # Cek apakah baris ini mengandung header yang diharapkan
                    row_values = []
                    for c in range(1, 15):  # Cek kolom A sampai N
                        cell_value = ws.cell(row=r, column=c).value
                        if cell_value:
                            row_values.append(str(cell_value).upper())
                    
                    row_text = ' '.join(row_values)
                    # Cari indikator header yang kuat
                    if ('BARCODE' in row_text and 'NAMA' in row_text and 
                        ('RUANGAN' in row_text or 'ASSET' in row_text) and 'KETERANGAN' in row_text):
                        header_row = r
                        self.append_log(f"[INFO] Header ditemukan pada baris: {header_row}")
                        break
                
                if header_row is None:
                    self.append_log(f"[ERROR] Header tidak ditemukan di sheet Recouncil file: {os.path.basename(file)}")
                    continue
                
                # Process data rows starting from header_row + 1
                source_last_row = ws.max_row
                data_found = 0
                duplicates_skipped = 0
                
                for i in range(header_row + 1, source_last_row + 1):
                    # Check for "Salah Ruangan" in column I (column 9)
                    keterangan = ws.cell(row=i, column=9).value  # Column I (Keterangan)
                    
                    if isinstance(keterangan, str) and 'salah ruangan' in keterangan.lower():
                        # Extract barcode untuk checking duplikasi
                        barcode_val = ws.cell(row=i, column=3).value
                        barcode_str = str(barcode_val).strip() if barcode_val else ""
                        
                        # SKIP jika barcode sudah pernah diproses (menghindari duplikasi)
                        if barcode_str in processed_barcodes:
                            duplicates_skipped += 1
                            self.append_log(f"[SKIP] Barcode {barcode_str} sudah diproses sebelumnya (duplikasi)")
                            continue
                        
                        # Tandai barcode sebagai sudah diproses
                        processed_barcodes.add(barcode_str)
                        data_found += 1
                        
                        # Extract data sesuai spesifikasi user:
                        # Nama Asset dari Column E (index 5) 
                        nama_asset = ws.cell(row=i, column=5).value
                        
                        # Ruangan Tujuan dari Column F (index 6) - "Ruangan Opname"
                        ruang_tujuan = ws.cell(row=i, column=6).value
                        
                        # Kondisi dari Column H (index 8)
                        kondisi = ws.cell(row=i, column=8).value
                        
                        # Keterangan dari Column I (index 9)
                        keterangan_val = ws.cell(row=i, column=9).value
                        
                        # LOGIKA BARU: Gunakan master data untuk mengisi Ruangan Asal dan PIC
                        # Ruangan Asal dari master data berdasarkan barcode
                        ruang_asal = self.ruangan_asal_map.get(barcode_str, "TIDAK ADA DI MASTER")
                        
                        # PIC Logic - PENTING: Diisi SETELAH Ruangan Asal terisi dari master data
                        # PIC berdasarkan barcode dari master data
                        pic = self.pic_map.get(barcode_str, "TIDAK ADA DI MASTER")
                        
                        # Action text template
                        action_text = f"Buat MAT Asset ini dan pindahkan dari ruangan {ruang_asal} ke ruangan {ruang_tujuan} sesuai dengan Hasil Opname"
                        
                        # Log extraction details
                        self.append_log(f"[EXTRACT] Baris {i}: Barcode={barcode_str}, Ruangan Asal={ruang_asal} -> PIC={pic}")
                        
                        rows.append({
                            "Barcode": barcode_val,      # Column A output
                            "PIC": pic,                  # Column B output (diisi setelah Ruangan Asal)
                            "Nama_Asset": nama_asset,    # Column C output
                            "R_Asal": ruang_asal,        # Column D output
                            "R_Tujuan": ruang_tujuan,    # Column E output
                            "Kondisi": kondisi,          # Column F output
                            "Keterangan": keterangan_val, # Column G output
                            "Action": action_text,       # Column H output
                            "Sumber": os.path.basename(file)
                        })
                
                self.append_log(f"[INFO] File: {os.path.basename(file)} - Ditemukan {data_found} baris unik, {duplicates_skipped} duplikasi dilewati")
            else:
                self.append_log(f"[SKIP] Sheet 'Recouncil' tidak ditemukan di file: {os.path.basename(file)}")

            # HAPUS PEMROSESAN FORM TEMUAN HASIL OPNAME untuk menghindari duplikasi
            # Hanya fokus pada sheet Recouncil sesuai permintaan user
        self.append_log(f"[INFO] Total baris 'salah ruangan' terkumpul: {len(rows)}")
        return rows

    def _generate_pic_colors(self, rows):
        """Generate unique colors for each PIC name"""
        from openpyxl.styles import PatternFill
        
        # Daftar warna terang yang tidak menutupi tulisan
        light_colors = [
            "E3F2FD",  # Light Blue
            "E8F5E8",  # Light Green
            "FFF3E0",  # Light Orange
            "F3E5F5",  # Light Purple
            "FFF8E1",  # Light Yellow
            "FCE4EC",  # Light Pink
            "E0F2F1",  # Light Teal
            "F1F8E9",  # Light Lime
            "FFF3E0",  # Light Deep Orange
            "E8EAF6",  # Light Indigo
            "F9FBE7",  # Light Light Green
            "FDE7F3",  # Light Rose
            "E1F5FE",  # Light Cyan
            "F3E5F5",  # Light Lavender
            "FFFDE7",  # Light Amber
            "FCE4EC",  # Light Pink
            "E0F7FA",  # Light Aqua
            "F1F8E9",  # Light Mint
            "FFF8E1",  # Light Cream
            "E8F5E8"   # Light Sage
        ]
        
        # Ambil semua nama PIC yang unik
        unique_pics = list(set(item["PIC"] for item in rows))
        self.append_log(f"[INFO] Ditemukan {len(unique_pics)} PIC unik: {unique_pics}")
        
        # Buat mapping PIC -> warna
        pic_color_map = {}
        for i, pic in enumerate(unique_pics):
            color_hex = light_colors[i % len(light_colors)]  # Cycle through colors if more PICs than colors
            pic_color_map[pic] = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
            self.append_log(f"[COLOR] PIC '{pic}' -> Warna #{color_hex}")
        
        return pic_color_map

    def write_to_mat_template(self, rows, save_path):
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
        
        if not self.mat_template_path or not os.path.exists(self.mat_template_path):
            raise FileNotFoundError("Template MAT tidak ditemukan. Pilih template terlebih dahulu.")
        wb = load_workbook(self.mat_template_path)
        if 'MAT' not in wb.sheetnames:
            raise ValueError("Template tidak memiliki sheet 'MAT'.")
        ws = wb['MAT']

        # Generate color mapping untuk setiap PIC
        pic_color_map = self._generate_pic_colors(rows)
        
        # Define styles
        DEFAULT_FONT_EXCEL = Font(name="Calibri", size=10)
        THIN_BORDER = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
        YELLOW_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

        start_row = 3
        r = start_row
        
        self.append_log(f"[INFO] Mulai menulis {len(rows)} baris data ke template MAT...")
        
        for item in rows:
            # Isi data ke cell
            ws.cell(row=r, column=1, value=item["Barcode"])
            ws.cell(row=r, column=2, value=item["PIC"])
            ws.cell(row=r, column=3, value=item["Nama_Asset"])
            ws.cell(row=r, column=4, value=item["R_Asal"])
            ws.cell(row=r, column=5, value=item["R_Tujuan"])
            ws.cell(row=r, column=6, value=item["Kondisi"])
            ws.cell(row=r, column=7, value=item["Keterangan"])
            c_action = ws.cell(row=r, column=8, value=item["Action"])
            c_action.fill = YELLOW_FILL  # Action column tetap kuning
            
            # Dapatkan warna berdasarkan nama PIC
            pic_name = item["PIC"]
            pic_fill = pic_color_map.get(pic_name, PatternFill())  # Default no fill if PIC not found
            
            # Apply warna PIC ke seluruh baris (kecuali kolom Action)
            for col in range(1, 8):  # Kolom A sampai G
                cell = ws.cell(row=r, column=col)
                cell.fill = pic_fill
                cell.font = DEFAULT_FONT_EXCEL
                cell.border = THIN_BORDER
                cell.alignment = CENTER_ALIGNMENT
            
            # Kolom Action (H) tetap dengan styling khusus
            c_action.font = DEFAULT_FONT_EXCEL
            c_action.border = THIN_BORDER
            c_action.alignment = CENTER_ALIGNMENT
            
            r += 1

        max_row = r - 1
        self.append_log(f"[INFO] Data berhasil ditulis sampai baris {max_row}")
        
        # Auto-adjust column widths
        for col in range(1, 9):
            letter = get_column_letter(col)
            max_len = 0
            for rr in range(2, max_row+1):
                v = ws.cell(row=rr, column=col).value
                if v is None: continue
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[letter].width = min(max_len + 4, 70)

        # Set row heights
        for rr in range(start_row, max_row+1):
            ws.row_dimensions[rr].height = 22

        wb.save(save_path)
        self.append_log(f"[INFO] File MAT berhasil disimpan dengan pewarnaan berdasarkan PIC: {save_path}")



#########################################################
# APLIKASI 4: EXTRACT RECOUNCIL
#########################################################
class ExtractRecouncilApp(ModernBaseWindow):
    def __init__(self, root, on_close):
        super().__init__(root, "🔄 EXTRACT RECOUNCIL", "800x750")
        self.add_back_button(on_close)
        self.master_file = ""
        self.target_files = []
        self.overwrite_file = tk.BooleanVar(value=False)

        # Create modern UI
        self.create_modern_ui()

    def create_modern_ui(self):
        # Description card
        desc_card = self.create_card_frame(
            self.main_frame, 
            "📋 Process Description",
            "Extract recouncil data processing with advanced filtering and reconciliation features"
        )
        desc_card.pack(fill=X, pady=(0, 20))
        
        # File selection section
        file_section = self.create_card_frame(self.main_frame, "📁 File Selection")
        file_section.pack(fill=X, pady=(0, 20))
        
        # Master file selection
        master_frame = ttk.Frame(file_section)
        master_frame.pack(fill=X, pady=(0, 15))
        
        btn_master = ttk.Button(
            master_frame, 
            text="🗃️ 1. Pilih Data Master ICT DAFT",
            bootstyle="secondary",
            command=self.pick_master
        )
        btn_master.pack(fill=X, pady=(0, 5))
        
        self.lbl_master = ttk.Label(
            master_frame, 
            text="❌ Belum ada file master dipilih",
            font=FONT_CONFIG["text"],
            bootstyle="danger"
        )
        self.lbl_master.pack(anchor=W)
        
        # Target files selection
        targets_frame = ttk.Frame(file_section)
        targets_frame.pack(fill=X, pady=(0, 15))
        
        btn_targets = ttk.Button(
            targets_frame, 
            text="📊 2. Pilih FILE Recouncil (boleh banyak)",
            bootstyle="secondary",
            command=self.pick_targets
        )
        btn_targets.pack(fill=X, pady=(0, 5))
        
        self.lbl_targets = ttk.Label(
            targets_frame, 
            text="❌ Belum ada file ruangan dipilih",
            font=FONT_CONFIG["text"],
            bootstyle="danger"
        )
        self.lbl_targets.pack(anchor=W)
        
        # Options section
        options_section = self.create_card_frame(self.main_frame, "⚙️ Processing Options")
        options_section.pack(fill=X, pady=(0, 20))
        
        cb_overwrite = ttk.Checkbutton(
            options_section, 
            text="🔄 Overwrite file asli (timpa file)",
            variable=self.overwrite_file,
            bootstyle="secondary-round-toggle"
        )
        cb_overwrite.pack(anchor=W, pady=5)
        
        # Add tooltip
        ToolTip(cb_overwrite, text="Enable to overwrite original files instead of creating copies")
        
        # Process button section
        process_section = ttk.Frame(self.main_frame)
        process_section.pack(fill=X, pady=(20, 0))
        
        self.btn_process = ttk.Button(
            process_section, 
            text="🚀 PROSES FILE",
            bootstyle="success",
            command=self.process_all,
            state=DISABLED
        )
        self.btn_process.pack(fill=X, pady=15, ipady=10)
        
        # Add tooltip
        ToolTip(self.btn_process, text="Process recouncil files with data reconciliation")
        
        # Log section
        log_section = self.create_card_frame(self.main_frame, "📝 Processing Log")
        log_section.pack(fill=BOTH, expand=True, pady=(0, 0))
        
        # Modern scrolled text
        self.log = ScrolledText(
            log_section, 
            wrap=WORD, 
            height=12,
            bootstyle="secondary"
        )
        self.log.pack(fill=BOTH, expand=True, padx=5, pady=5)

    # ---- helpers ----
    def _btn_style(self): return {"bg": STYLE["accent"], "fg": STYLE["text"], "font": STYLE["font_button"], "relief": tk.FLAT, "pady": 8}
    def _log(self, msg):
        self.log.insert(tk.END, msg + "\n"); self.log.see(tk.END); self.log.update_idletasks()

    def pick_master(self):
        f = filedialog.askopenfilename(title="Pilih Data Master ICT DAFT", filetypes=[("Excel files","*.xlsx *.xls")])
        if f:
            self.master_file = f
            self.lbl_master.config(
                text=f"✅ {os.path.basename(f)}", 
                bootstyle="success"
            )
        self._refresh()

    def pick_targets(self):
        files = filedialog.askopenfilenames(title="Pilih banyak file Recouncil", filetypes=[("Excel files","*.xlsx")])
        if files:
            self.target_files = list(files)
            self.lbl_targets.config(
                text=f"✅ {len(self.target_files)} file dipilih", 
                bootstyle="success"
            )
        self._refresh()

    def _refresh(self):
        if self.master_file and self.target_files: 
            self.btn_process.config(state=NORMAL)
            self.btn_process.config(bootstyle="success")
        else: 
            self.btn_process.config(state=DISABLED)
            self.btn_process.config(bootstyle="secondary-outline")

    # ---- core ----
    def process_all(self):
        try:
            mapping = self._load_master_mapping(self.master_file)
            self._log(f"[INFO] Data Master berhasil dimuat dengan {len(mapping)} barcode.")
        except Exception as e:
            messagebox.showerror("Error Master", f"Gagal memuat master: {e}")
            return

        ok, fail = 0, 0
        for path in self.target_files:
            try:
                self._process_one(path, mapping)
                ok += 1
            except Exception as e:
                fail += 1; self._log(f"[GAGAL] {os.path.basename(path)} → {e}")
        messagebox.showinfo("Selesai", f"Proses Selesai.\n\nBerhasil: {ok}\nGagal: {fail}")

    def _load_master_mapping(self, master_path):
        from openpyxl import load_workbook
        wb = load_workbook(master_path, data_only=True)
        
        sheet_name = ""
        if 'Recouncil' in wb.sheetnames:
            sheet_name = 'Recouncil'
        else:
            sheet_name = wb.sheetnames[0]
            self._log(f"[PERINGATAN] Sheet 'Recouncil' tidak ditemukan. Menggunakan sheet pertama: '{sheet_name}'.")

        self._log(f"[INFO] Membaca sheet '{sheet_name}' dari file master...")
        ws = wb[sheet_name]

        def norm_bar(x):
            s = str(x).strip()
            m = re.search(r'(\d{6,})', s)
            return m.group(1) if m else s

        mapping = {}
        for r in range(2, ws.max_row + 1):
            key = norm_bar(ws.cell(row=r, column=1).value)
            if not key or str(key).strip().lower() == "nan": continue
            vals = [ws.cell(row=r, column=c).value for c in range(1, 7)]
            mapping[str(key)] = vals
        return mapping

    def _cell_rgb(self, cell):
        f = cell.fill
        if f and f.fill_type:
            c = f.start_color
            if getattr(c, "rgb", None):
                return c.rgb.upper()
        return None
    def _is_yellow(self, rgb): return rgb in {"FFFF00","FFFFFF00"}
    def _is_blue(self, rgb):   return rgb in {"FF5C97FA","5C97FA","FFADD8E6","ADD8E6"}

    def _process_one(self, file_path, mapping):
        from openpyxl import load_workbook
        self._log(f"[PROSES] Membaca {os.path.basename(file_path)}...")
        wb = load_workbook(file_path)
        if 'Recouncil' not in wb.sheetnames:
            raise ValueError("Sheet 'Recouncil' tidak ditemukan.")
        ws = wb['Recouncil']

        last_data_row = 0
        for r in range(ws.max_row, 7, -1):
            if ws.cell(row=r, column=2).value not in (None, ""):
                last_data_row = r
                break
        
        if last_data_row == 0:
            self._log(f"   -> Tidak ada data barcode ditemukan di {os.path.basename(file_path)}.")
            return

        # APP 4 FIX: Clear previous results first (like VBA)
        for r in range(8, last_data_row + 1):
            for col in range(11, 18):  # Clear columns K-Q
                ws.cell(row=r, column=col).value = None

        # APP 4 FIX: Process each row with VBA logic
        for r in range(8, last_data_row + 1):
            # Get barcode from column C only (like VBA)
            bc_val_c = ws.cell(row=r, column=3).value
            barcode = str(bc_val_c).strip() if bc_val_c is not None else ""
            
            q_cell = ws.cell(row=r, column=17)
            
            if barcode and barcode in mapping:
                # Data found in master (like VBA: If masterDataDict.Exists(barcodeToFind) Then)
                data_item = mapping[barcode]
                
                # Fill columns K-P with master data (like VBA)
                output_data = []
                is_data_complete = True
                
                for k in range(len(data_item)):
                    if data_item[k] is None or str(data_item[k]).strip() == "":
                        output_data.append("-")
                        is_data_complete = False
                    else:
                        output_data.append(data_item[k])
                
                # Fill columns K-P
                for col_idx, value in enumerate(output_data):
                    if col_idx < 6:  # Only 6 columns K-P
                        ws.cell(row=r, column=11 + col_idx).value = value
                
                # Set column Q based on data completeness (exactly like VBA)
                if is_data_complete:
                    q_cell.value = "Sudah Sesuai"
                else:
                    q_cell.value = "Data Ditemukan Tidak Lengkap"
                    
            else:
                # Data not found in master (like VBA: Else clause)
                # Fill with "-" (like VBA)
                for col in range(11, 17):  # Columns K-P
                    ws.cell(row=r, column=col).value = "-"
                
                # Set message (like VBA)
                q_cell.value = "Barcode Belum Sesuai, Asset di Oracle tidak ada"
            
            # APP 4 FIX: Format kolom P (tanggal) - hanya tanggal tanpa waktu
            p_cell = ws.cell(row=r, column=16)  # Kolom P
            if p_cell.value and hasattr(p_cell.value, 'date'):
                p_cell.value = p_cell.value.date()
                p_cell.number_format = 'dd/mm/yyyy'

            for i in range(11, 18):
                cell = ws.cell(row=r, column=i)
                cell.font = DEFAULT_FONT_EXCEL
                cell.border = THIN_BORDER
                cell.alignment = CENTER_ALIGNMENT

        if self.overwrite_file.get():
            out_path = file_path
            self._log(f"   -> SUKSES: Menimpa file asli: {os.path.basename(out_path)}")
        else:
            base, ext = os.path.splitext(os.path.basename(file_path))
            out_path = os.path.join(os.path.dirname(file_path), f"{base}_output{ext}")
            self._log(f"   -> SUKSES: Menyimpan salinan: {os.path.basename(out_path)}")
        wb.save(out_path)



#########################################################
# SIMPLE MAIN MENU APP (ttkbootstrap)
#########################################################
class MainMenuApp:
    def __init__(self):
        # Initialize ttkbootstrap window with dark theme
        self.root = ttk.Window(themename=BOOTSTRAP_THEME)
        self.root.title("ICT OPNAME PROCESSOR")
        self.root.geometry("900x700")
        
        # Make window resizable but set minimum size
        self.root.minsize(800, 600)
        self.root.maxsize(1000, 800)
        self.root.resizable(True, True)
        
        # Center window
        self.center_window()
        
        # Create simple interface
        self.create_interface()
    
    def center_window(self):
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f"{width}x{height}+{x}+{y}")
    
    def create_menu_bar(self, parent):
        # Menu bar frame with enhanced features
        menu_frame = ttk.Frame(parent)
        menu_frame.pack(fill=X, pady=(0, 15))
        
        # Left side buttons
        left_frame = ttk.Frame(menu_frame)
        left_frame.pack(side=LEFT)
        
        # Help button
        help_btn = ttk.Button(
            left_frame,
            text="❓ Help",
            bootstyle="info-outline",
            command=self.show_help,
            width=10
        )
        help_btn.pack(side=LEFT, padx=(0, 5))
        
        # About button
        about_btn = ttk.Button(
            left_frame,
            text="ℹ️ About",
            bootstyle="secondary-outline",
            command=self.show_about,
            width=10
        )
        about_btn.pack(side=LEFT, padx=(0, 5))
        
        # Settings button
        settings_btn = ttk.Button(
            left_frame,
            text="⚙️ Settings",
            bootstyle="warning-outline",
            command=self.show_settings,
            width=12
        )
        settings_btn.pack(side=LEFT, padx=(0, 5))
        
        # Tools button
        tools_btn = ttk.Button(
            left_frame,
            text="🔧 Tools",
            bootstyle="primary-outline",
            command=self.show_tools,
            width=10
        )
        tools_btn.pack(side=LEFT, padx=(0, 5))
        
        # Right side buttons
        right_frame = ttk.Frame(menu_frame)
        right_frame.pack(side=RIGHT)
        
        # Minimize button
        minimize_btn = ttk.Button(
            right_frame,
            text="🗕 Minimize",
            bootstyle="secondary-outline",
            command=self.root.iconify,
            width=12
        )
        minimize_btn.pack(side=RIGHT, padx=(5, 0))
        
        # Exit button
        exit_btn = ttk.Button(
            right_frame,
            text="❌ Exit",
            bootstyle="danger-outline",
            command=self.root.quit,
            width=10
        )
        exit_btn.pack(side=RIGHT, padx=(5, 0))
    
    def create_interface(self):
        # Main container with proper padding
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=BOTH, expand=True, padx=30, pady=20)
        
        # Menu bar with useful features
        self.create_menu_bar(main_frame)
        
        # Title with proper spacing
        title_label = ttk.Label(
            main_frame,
            text="ICT OPNAME PROCESSOR",
            font=("Segoe UI", 20, "bold"),
            bootstyle="light"
        )
        title_label.pack(pady=(0, 25))
        
        # Apps grid - SIMPLE AND CLEAR
        apps_frame = ttk.Frame(main_frame)
        apps_frame.pack(expand=True, fill=BOTH)
        
        # Configure grid
        apps_frame.columnconfigure(0, weight=1)
        apps_frame.columnconfigure(1, weight=1)
        apps_frame.rowconfigure(0, weight=1)
        apps_frame.rowconfigure(1, weight=1)
        
        # App buttons - SIMPLE APPROACH
        apps = [
            {"title": "DAFT ICT FILE\nPROCESSOR", "icon": "📊", "command": self.open_daft_processor},
            {"title": "EXTRACT\nOPNAME", "icon": "📈", "command": self.open_extract_opname},
            {"title": "EXTRACT\nMAT", "icon": "📋", "command": self.open_extract_mat},
            {"title": "EXTRACT\nRECOUNCIL", "icon": "🔄", "command": self.open_extract_recouncil}
        ]
        
        for i, app in enumerate(apps):
            row = i // 2
            col = i % 2
            
            # Simple container frame
            container = ttk.Frame(apps_frame)
            container.grid(row=row, column=col, padx=15, pady=15, sticky="nsew")
            
            # Large icon
            icon_label = ttk.Label(
                container,
                text=app['icon'],
                font=("Segoe UI", 40)
            )
            icon_label.pack(pady=(10, 5))
            
            # BIG VISIBLE BUTTON
            btn = ttk.Button(
                container,
                text=app['title'],
                bootstyle="success",  # Green for high visibility
                command=app['command']
            )
            btn.pack(pady=(5, 15), ipady=20, ipadx=20, fill=X)
            
            # Configure button font
            btn.configure(style="AppButton.TButton")
        
        # Configure button style
        style = ttk.Style()
        style.configure("AppButton.TButton", 
                       font=("Segoe UI", 12, "bold"),
                       padding=(10, 15))
        
        # Configure grid weights
        for i in range(2):
            apps_frame.rowconfigure(i, weight=1)
            apps_frame.columnconfigure(i, weight=1)
        
        # Footer with reduced spacing
        footer_label = ttk.Label(
            main_frame,
            text="© 2024 ICT Opname Processor V4",
            font=("Segoe UI", 8),
            bootstyle="secondary"
        )
        footer_label.pack(side=BOTTOM, pady=(10, 0))
        
    def open_window(self, app_class):
        self.root.withdraw()
        new_window = ttk.Toplevel(self.root)
        app_class(new_window, on_close=lambda: self.show_main_menu(new_window))

    def open_daft_processor(self):
        try:
            self.root.withdraw()
            # Create new toplevel window immediately to avoid white flash
            new_window = ttk.Toplevel(self.root)
            new_window.withdraw()  # Hide initially
            DaftProcessorApp(new_window, self.show_main_menu)
            new_window.deiconify()  # Show after creation
        except Exception as e:
            self.show_main_menu()
            self.show_app_error("DAFT ICT FILE PROCESSOR", e)

    def open_extract_opname(self):
        try:
            self.root.withdraw()
            # Create new toplevel window immediately to avoid white flash
            new_window = ttk.Toplevel(self.root)
            new_window.withdraw()  # Hide initially
            ExtractOpnameApp(new_window, self.show_main_menu)
            new_window.deiconify()  # Show after creation
        except Exception as e:
            self.show_main_menu()
            self.show_app_error("EXTRACT OPNAME", e)

    def open_extract_mat(self):
        try:
            self.root.withdraw()
            # Create new toplevel window immediately to avoid white flash
            new_window = ttk.Toplevel(self.root)
            new_window.withdraw()  # Hide initially
            ExtractMATApp(new_window, self.show_main_menu)
            new_window.deiconify()  # Show after creation
        except Exception as e:
            self.show_main_menu()
            self.show_app_error("EXTRACT MAT", e)

    def open_extract_recouncil(self):
        try:
            self.root.withdraw()
            # Create new toplevel window immediately to avoid white flash
            new_window = ttk.Toplevel(self.root)
            new_window.withdraw()  # Hide initially
            ExtractRecouncilApp(new_window, self.show_main_menu)
            new_window.deiconify()  # Show after creation
        except Exception as e:
            self.show_main_menu()
            self.show_app_error("EXTRACT RECOUNCIL", e)
        
    def show_main_menu(self, child_window=None):
        # Close child window if provided
        if child_window:
            child_window.destroy()
        
        # Re-show main window when returning from apps
        self.root.deiconify()
        
        # Ensure main menu stays in center with normal size
        self.center_window()
    
    def show_app_error(self, app_name, error):
        error_msg = f"Error membuka {app_name}!\n\n"
        error_msg += f"Error Type: {type(error).__name__}\n"
        error_msg += f"Error Message: {str(error)}\n\n"
        error_msg += f"Detailed Traceback:\n{traceback.format_exc()}"
        
        try:
            import tkinter.messagebox as messagebox
            messagebox.showerror(f"Error - {app_name}", error_msg)
        except:
            try:
                Messagebox.show_error(f"Error - {app_name}", error_msg)
            except:
                print(f"Error opening {app_name}: {error}")
    
    def show_help(self):
        help_text = """🔧 ICT OPNAME PROCESSOR V4 - PANDUAN PENGGUNAAN

📊 APP 1 - DAFT ICT FILE PROCESSOR:
• Memproses file allocation user aktiva
• Input: File Master, File EXA, File INV
• Output: File kamus data yang sudah diupdate
• Fitur: Filter BAT untuk EXA dan INV

📈 APP 2 - EXTRACT OPNAME:
• Ekstrak dan proses data opname
• Input: File opname target, data terscan, data tidak terscan
• Output: File hasil ekstraksi opname
• Fitur: Column mapping otomatis

📋 APP 3 - EXTRACT MAT:
• Ekstrak data MAT dari file opname
• Input: Database PIC, file ruangan, template MAT
• Output: File MAT yang sudah diproses
• Fitur: Integrasi database PIC

🔄 APP 4 - EXTRACT RECOUNCIL:
• Proses data recouncil dengan filtering
• Input: Data master ICT DAFT, file recouncil
• Output: File recouncil yang sudah diproses
• Fitur: Opsi overwrite file asli

💡 TIPS:
• Pastikan format file Excel sesuai
• Backup file sebelum processing
• Periksa log untuk detail proses"""
        
        try:
            import tkinter.messagebox as messagebox
            messagebox.showinfo("Help - ICT Opname Processor", help_text)
        except:
            try:
                Messagebox.show_info("Help - ICT Opname Processor", help_text)
            except:
                print(help_text)
    
    def show_settings(self):
        """Show settings dialog with useful configuration options"""
        settings_window = ttk.Toplevel(self.root)
        settings_window.title("⚙️ Settings - ICT Opname Processor")
        settings_window.geometry("500x400")
        settings_window.transient(self.root)
        settings_window.grab_set()
        
        # Center the window
        settings_window.update_idletasks()
        x = (settings_window.winfo_screenwidth() // 2) - (settings_window.winfo_width() // 2)
        y = (settings_window.winfo_screenheight() // 2) - (settings_window.winfo_height() // 2)
        settings_window.geometry(f"+{x}+{y}")
        
        main_frame = ttk.Frame(settings_window)
        main_frame.pack(fill=BOTH, expand=True, padx=20, pady=20)
        
        # Title
        title_label = ttk.Label(main_frame, text="⚙️ Application Settings", font=("Segoe UI", 14, "bold"))
        title_label.pack(pady=(0, 20))
        
        # Settings options
        settings_frame = ttk.LabelFrame(main_frame, text="General Settings", padding=15)
        settings_frame.pack(fill=X, pady=(0, 10))
        
        # Auto-save settings
        auto_save_var = tk.BooleanVar(value=True)
        auto_save_cb = ttk.Checkbutton(settings_frame, text="Auto-save processed files", variable=auto_save_var)
        auto_save_cb.pack(anchor=W, pady=5)
        
        # Show progress dialog
        show_progress_var = tk.BooleanVar(value=True)
        progress_cb = ttk.Checkbutton(settings_frame, text="Show progress dialog during processing", variable=show_progress_var)
        progress_cb.pack(anchor=W, pady=5)
        
        # Default output folder
        output_frame = ttk.Frame(settings_frame)
        output_frame.pack(fill=X, pady=10)
        ttk.Label(output_frame, text="Default Output Folder:").pack(anchor=W)
        output_entry = ttk.Entry(output_frame, width=50)
        output_entry.pack(side=LEFT, fill=X, expand=True, padx=(0, 5))
        ttk.Button(output_frame, text="Browse", width=8).pack(side=RIGHT)
        
        # Performance settings
        perf_frame = ttk.LabelFrame(main_frame, text="Performance Settings", padding=15)
        perf_frame.pack(fill=X, pady=(0, 10))
        
        # Memory optimization
        memory_var = tk.BooleanVar(value=False)
        memory_cb = ttk.Checkbutton(perf_frame, text="Enable memory optimization for large files", variable=memory_var)
        memory_cb.pack(anchor=W, pady=5)
        
        # Parallel processing
        parallel_var = tk.BooleanVar(value=True)
        parallel_cb = ttk.Checkbutton(perf_frame, text="Enable parallel processing", variable=parallel_var)
        parallel_cb.pack(anchor=W, pady=5)
        
        # Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=X, pady=(20, 0))
        
        ttk.Button(button_frame, text="Save Settings", bootstyle="success").pack(side=RIGHT, padx=(5, 0))
        ttk.Button(button_frame, text="Cancel", command=settings_window.destroy).pack(side=RIGHT)
        ttk.Button(button_frame, text="Reset to Default", bootstyle="warning-outline").pack(side=LEFT)

    def show_tools(self):
        """Show tools dialog with utility functions"""
        tools_window = ttk.Toplevel(self.root)
        tools_window.title("🔧 Tools - ICT Opname Processor")
        tools_window.geometry("600x500")
        tools_window.transient(self.root)
        tools_window.grab_set()
        
        # Center the window
        tools_window.update_idletasks()
        x = (tools_window.winfo_screenwidth() // 2) - (tools_window.winfo_width() // 2)
        y = (tools_window.winfo_screenheight() // 2) - (tools_window.winfo_height() // 2)
        tools_window.geometry(f"+{x}+{y}")
        
        main_frame = ttk.Frame(tools_window)
        main_frame.pack(fill=BOTH, expand=True, padx=20, pady=20)
        
        # Title
        title_label = ttk.Label(main_frame, text="🔧 Utility Tools", font=("Segoe UI", 14, "bold"))
        title_label.pack(pady=(0, 20))
        
        # Tools grid
        tools_frame = ttk.Frame(main_frame)
        tools_frame.pack(fill=BOTH, expand=True)
        
        tools = [
            {"name": "📊 Excel File Analyzer", "desc": "Analyze Excel file structure and sheets", "cmd": self.tool_excel_analyzer},
            {"name": "🔍 Barcode Validator", "desc": "Validate barcode formats and duplicates", "cmd": self.tool_barcode_validator},
            {"name": "📋 Template Generator", "desc": "Generate standard templates for processing", "cmd": self.tool_template_generator},
            {"name": "🗂️ File Organizer", "desc": "Organize and rename processed files", "cmd": self.tool_file_organizer},
            {"name": "📈 Data Statistics", "desc": "Generate statistics from processed data", "cmd": self.tool_data_statistics},
            {"name": "🔄 Batch Converter", "desc": "Convert multiple files between formats", "cmd": self.tool_batch_converter}
        ]
        
        for i, tool in enumerate(tools):
            row = i // 2
            col = i % 2
            
            tool_frame = ttk.LabelFrame(tools_frame, text=tool["name"], padding=10)
            tool_frame.grid(row=row, column=col, padx=10, pady=10, sticky="nsew")
            
            desc_label = ttk.Label(tool_frame, text=tool["desc"], wraplength=200)
            desc_label.pack(pady=(0, 10))
            
            ttk.Button(tool_frame, text="Launch Tool", command=tool["cmd"], bootstyle="primary").pack()
        
        # Configure grid
        for i in range(3):
            tools_frame.rowconfigure(i, weight=1)
        for i in range(2):
            tools_frame.columnconfigure(i, weight=1)
        
        # Close button
        ttk.Button(main_frame, text="Close", command=tools_window.destroy, bootstyle="secondary").pack(pady=(20, 0))

    def tool_excel_analyzer(self):
        messagebox.showinfo("Excel Analyzer", "Excel File Analyzer akan segera tersedia!")
    
    def tool_barcode_validator(self):
        messagebox.showinfo("Barcode Validator", "Barcode Validator akan segera tersedia!")
    
    def tool_template_generator(self):
        messagebox.showinfo("Template Generator", "Template Generator akan segera tersedia!")
    
    def tool_file_organizer(self):
        messagebox.showinfo("File Organizer", "File Organizer akan segera tersedia!")
    
    def tool_data_statistics(self):
        messagebox.showinfo("Data Statistics", "Data Statistics akan segera tersedia!")
    
    def tool_batch_converter(self):
        messagebox.showinfo("Batch Converter", "Batch Converter akan segera tersedia!")

    def show_about(self):
        about_text = """🖥️ ICT OPNAME PROCESSOR


📋 Aplikasi untuk memproses file opname ICT
🔧 ttkbootstrap style
🎨 dark theme
⚡ Enhanced performance & reliability

📅 Version: Enhanced Edition
🏢 Untuk: ICT Asset Management
👨‍💻 Interface: Modern ttkbootstrap GUI

✨ Fitur Utama:
• Modern dark theme dengan UI yang responsif
• Enhanced error handling & comprehensive logging
• Smart default sheet processing
• Multiple file processing dengan progress tracking
• Advanced tools & utilities
• Customizable settings & preferences
• Template-based processing untuk konsistensi

🆕 Fitur Baru:
• Settings panel untuk kustomisasi
• Tools panel dengan utility functions
• Enhanced main menu dengan descriptions
• Improved button visibility & styling
• Better error handling & user feedback

© 2025 ICT Opname Processor"""
        
        try:
            import tkinter.messagebox as messagebox
            messagebox.showinfo("About - ICT Opname Processor V4", about_text)
        except:
            try:
                Messagebox.show_info("About - ICT Opname Processor V4", about_text)
            except:
                print(about_text)
    
    def run(self):
        self.root.mainloop()



#########################################################
# MAIN ENTRY POINT
#########################################################
def main():
    try:
        print("Starting ICT OPNAME PROCESSOR V4...")
        print("Creating MainMenuApp...")
        app = MainMenuApp()
        print("MainMenuApp created successfully")
        print("Starting application...")
        app.run()
    except Exception as e:
        error_msg = f"CRITICAL ERROR - Aplikasi tidak dapat dijalankan!\n\n"
        error_msg += f"Error Type: {type(e).__name__}\n"
        error_msg += f"Error Message: {str(e)}\n\n"
        error_msg += f"Detailed Traceback:\n{traceback.format_exc()}"
        
        print("ERROR OCCURRED:")
        print(error_msg)
        
        # Try to show error in messagebox
        try:
            import tkinter as tk
            import tkinter.messagebox as messagebox
            root = tk.Tk()
            root.withdraw()  # Hide main window
            messagebox.showerror("Error Aplikasi Kritis", error_msg)
            root.destroy()
        except:
            # If messagebox fails, try ttkbootstrap
            try:
                Messagebox.show_error("Error Aplikasi Kritis", error_msg)
            except:
                # If all GUI fails, print to console
                print("GUI Error display failed. Error details printed above.")
                input("Press Enter to exit...")

def _setup_locale():
    """Setup locale for Indonesian number formatting"""
    try:
        locale.setlocale(locale.LC_ALL, 'id_ID.UTF-8')
    except:
        try:
            locale.setlocale(locale.LC_ALL, 'Indonesian_Indonesia.1252')
        except:
            pass  # Use default locale

def _check_modern_interface_files():
    """Check if modern interface files are available"""
    required_files = [
        "integrated_main_menu_final.py", 
        "theme.py", 
        "components.py", 
        "simple_modern_menu.py"
    ]
    
    print("Checking modern interface files...")
    files_status = {}
    for file in required_files:
        exists = os.path.exists(file)
        files_status[file] = exists
        print(f"  {file}: {'OK' if exists else 'MISSING'}")
    
    files_exist = all(files_status.values())
    
    if files_exist:
        print("[SUCCESS] All modern interface files detected!")
        return True
    else:
        missing_files = [f for f, exists in files_status.items() if not exists]
        print(f"[ERROR] Missing files: {', '.join(missing_files)}")
        return False

def _launch_advanced_modern_interface():
    """Launch advanced modern interface (CustomTkinter)"""
    try:
        import importlib.util
        spec = importlib.util.spec_from_file_location("modern_menu", "integrated_main_menu_final.py")
        modern_menu = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(modern_menu)
        
        print("Loading Advanced Modern Interface (CustomTkinter)...")
        modern_menu.main()
        return True
    except Exception as e:
        print(f"Advanced modern interface failed: {e}")
        return False

def _launch_simple_modern_interface():
    """Launch simple modern interface (Standard Tkinter)"""
    try:
        import importlib.util
        spec = importlib.util.spec_from_file_location("simple_modern", "simple_modern_menu.py")
        simple_modern = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(simple_modern)
        
        print("Loading Simple Modern Interface (Standard Tkinter)...")
        simple_modern.main()
        return True
    except Exception as e:
        print(f"Simple modern interface failed: {e}")
        return False

def _launch_classic_interface():
    """Launch classic interface (ttkbootstrap)"""
    try:
        print("Loading Classic Interface (ttkbootstrap)...")
        app = MainMenuApp()
        app.run()
        return True
    except Exception as e:
        print(f"Classic interface failed: {e}")
        return False

def main():
    """
    Main entry point with robust multi-layer interface system
    
    Interface Priority:
    1. Advanced Modern Interface (CustomTkinter) - Best experience
    2. Simple Modern Interface (Standard Tkinter) - Good experience, universal compatibility  
    3. Classic Interface (ttkbootstrap) - Reliable fallback
    """
    try:
        # Setup locale
        _setup_locale()
        
        # Display header
        print("ICT OPNAME PROCESSOR - Enhanced Edition")
        print("=" * 50)
        
        # Check modern interface availability
        modern_files_available = _check_modern_interface_files()
        
        if modern_files_available:
            print("Starting Modern Interface...")
            print("(Press Ctrl+C during startup to use Classic Interface)")
            
            # Try advanced modern interface first
            if _launch_advanced_modern_interface():
                return
            
            # Fallback to simple modern interface
            print("Trying simple modern interface...")
            if _launch_simple_modern_interface():
                return
            
            # If both modern interfaces fail, use classic
            print("Modern interfaces failed, falling back to classic...")
        else:
            print("Modern interface files not available.")
        
        # Launch classic interface as final fallback
        print("Using Classic Interface...")
        if not _launch_classic_interface():
            raise Exception("All interface options failed")
        
    except KeyboardInterrupt:
        print("\nSwitching to Classic Interface...")
        try:
            _launch_classic_interface()
        except:
            print("Application interrupted by user")
    except Exception as e:
        print(f"Critical error: {e}")
        traceback.print_exc()
        if sys.platform.startswith('win'):
            input("Press Enter to exit...")

if __name__ == "__main__":
    main()
