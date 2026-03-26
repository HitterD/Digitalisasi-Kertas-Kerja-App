import pandas as pd
from openpyxl import load_workbook
import traceback, sys

def clean_string(val):
    if pd.isna(val) or val is None:
        return ""
    if isinstance(val, float):
        if val.is_integer():
            return str(int(val))
        return str(val).rstrip('0').rstrip('.')
    if isinstance(val, int):
        return str(val)
    if isinstance(val, str):
        val = val.strip()
        if val.endswith('.0'):
            val = val[:-2]
        return val
    return str(val).strip()

try:
    opname_file = r"d:\Digitalisasi Kertas Kerja APP\Recouncil\Recouncil - Januari\Recouncil\519-OPEN OFFICE QA LT DASAR GILANG 137-SJA 1.XLSX"
    wb = load_workbook(opname_file, data_only=True)
    ws = wb['Recouncil'] if 'Recouncil' in wb.sheetnames else wb.active
    
    print("Reading rows 8-15")
    for r in range(8, 15):
        val_c = ws.cell(row=r, column=3).value
        val_d = ws.cell(row=r, column=4).value
        val_f = ws.cell(row=r, column=6).value
        print(f"Row {r}: C={val_c} | D={val_d} | F={val_f} | cleanedC={clean_string(val_c)}")
        
except Exception as e:
    traceback.print_exc()
