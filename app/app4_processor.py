import pandas as pd
import numpy as np
import os
import re
import traceback
import json
import zipfile
import shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

DEFAULT_FONT_EXCEL = Font(name='Calibri', size=11)
HEADER_FONT_EXCEL = Font(name='Calibri', size=11, bold=True, color='000000')
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)

MASTER_FILL = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid') # Kuning/orange untuk data Oracle MS
RESULT_FILL = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid') # Biru cerah (hasil recouncil)

def clean_string(val):
    if pd.isna(val) or val is None:
        return ""
    if isinstance(val, float):
        if val.is_integer():
            return str(int(val))
        return str(val).rstrip('0').rstrip('.')
    if isinstance(val, int):
        return str(val)
    if isinstance(val, str):
        val = val.strip()
        if val.endswith('.0'):
            val = val[:-2]
        return val
    return str(val).strip()

def make_unique_cols(cols):
    seen = {}
    new_cols = []
    for c in cols:
        new_c = str(c).strip()
        if new_c in seen:
            seen[new_c] += 1
            new_cols.append(f"{new_c}_{seen[new_c]}")
        else:
            seen[new_c] = 0
            new_cols.append(new_c)
    return new_cols

def style_cell(cell, is_header=False, section="master"):
    cell.font = HEADER_FONT_EXCEL if is_header else DEFAULT_FONT_EXCEL
    cell.border = THIN_BORDER
    cell.alignment = CENTER_ALIGNMENT
    if is_header:
        if section == "master":
            cell.fill = MASTER_FILL
        else:
            cell.fill = RESULT_FILL

def run_recouncil(opname_paths, master_path, output_path):
    """
    Memproses data Opname (App2) dengan Data Master (App3).
    Mengedit file Opname in-place untuk mempertahankan format aslinya (headers, judul, dsb),
    hanya menambahkan kolom K-Q di sebelah kanan tabel aslinya.
    Jika multiple file, mengembalikan bentuk zip.
    """
    if not isinstance(opname_paths, list):
        opname_paths = [opname_paths]
        
    for path in opname_paths:
        if not os.path.exists(path):
            return {"status": "error", "message": f"File sumber tidak ditemukan: {path}"}
            
    if not os.path.exists(master_path):
        return {"status": "error", "message": "File Master Data tidak ditemukan."}

    try:
        # 1. Load Master Data ke Dictionary
        df_master = pd.read_excel(master_path)
        df_master.columns = make_unique_cols(df_master.columns)
        
        master_dict = {}
        for _, row in df_master.iterrows():
            barcode = clean_string(row.get('NO BARCODE', ''))
            if barcode and barcode not in ['nan', 'None', '', '-']:
                master_data = {
                    'NO BARCODE': clean_string(row.get('NO BARCODE', '-')),
                    'ASSET ORACLE': clean_string(row.get('ASSET ORACLE', '-')),
                    'LOKASI': clean_string(row.get('LOKASI', '-')),
                    'JENIS HARTA': clean_string(row.get('JENIS HARTA', '-')),
                    'KONDISI': clean_string(row.get('KONDISI', '-')),
                    'TAHUN Perolehan': clean_string(row.get('Tahun Perolehan', '-'))
                }
                
                # Handle dates for excel writing
                tp = master_data['TAHUN Perolehan']
                if hasattr(tp, 'date'):
                    master_data['TAHUN Perolehan'] = tp.strftime('%Y')
                elif isinstance(tp, str) and ' ' in tp and ':' in tp:
                     master_data['TAHUN Perolehan'] = tp.split(' ')[0][:4]
                     
                master_dict[barcode] = master_data

        processed_files = []
        
        # 2. Proses masing-masing Opname Path secara In-Place
        for opname_path in opname_paths:
            wb = load_workbook(opname_path)
            
            # Cari sheet yang digunakan (prioritas: Recouncil)
            ws = wb['Recouncil'] if 'Recouncil' in wb.sheetnames else wb.active
                
            # Deteksi baris terakhir yang memiliki data di kolom A-J supaya tidak format sampai ribuan ke bawah
            last_data_row = 7
            for r in range(ws.max_row, 7, -1):
                has_data = False
                for c in range(1, 11): # Cek kolom A sampai J
                    val = ws.cell(row=r, column=c).value
                    if val is not None and str(val).strip() != "":
                        has_data = True
                        break
                if has_data:
                    last_data_row = r
                    break
            
            # Jika last_data_row mentok di 7 dan max_row jauh lebih besar, bisa jadi sheet kosong
            if last_data_row < 8:
                last_data_row = max(ws.max_row, 8)
                
            # Kita tulis ulang header di baris 7 agar formatnya rapi
            new_headers = ['NO BARCODE', 'ASSET ORACLE', 'LOKASI', 'JENIS HARTA', 'KONDISI MASTER', 'TAHUN Perolehan', 'HASIL RECOUNCIL']
            for i, h_name in enumerate(new_headers):
                cell = ws.cell(row=7, column=11 + i) # Mulai dari Kolom K (11)
                cell.value = h_name
                style_cell(cell, is_header=True, section="master" if i < 6 else "result")

            # Mulai memproses data tepat di bawah header (baris 8 ke bawah)
            for r in range(8, last_data_row + 1):
                raw_barcode = ws.cell(row=r, column=3).value # Pasti Kolom C (Index 3)
                raw_ruang = ws.cell(row=r, column=6).value   # Pasti Kolom F (Index 6)
                
                barcode_opname = clean_string(raw_barcode)
                ruangan_opname = clean_string(raw_ruang)
                
                status_recouncil = ""
                out_master = {
                    'NO BARCODE': '-', 'ASSET ORACLE': '-', 'LOKASI': '-', 
                    'JENIS HARTA': '-', 'KONDISI MASTER': '-', 'TAHUN Perolehan': '-'
                }
                
                if not barcode_opname or barcode_opname in ['nan', '-', 'None', '']:
                    status_recouncil = "Barcode Kosong di Opname"
                elif barcode_opname not in master_dict:
                    status_recouncil = "Barcode Belum Sesuai, Asset di Oracle tidak ada"
                    out_master['NO BARCODE'] = barcode_opname # Force paste barcode opname here
                else:
                    m_data = master_dict[barcode_opname]
                    out_master = {
                        'NO BARCODE': barcode_opname, # Always use the scanned barcode
                        'ASSET ORACLE': m_data['ASSET ORACLE'],
                        'LOKASI': m_data['LOKASI'],
                        'JENIS HARTA': m_data['JENIS HARTA'],
                        'KONDISI MASTER': m_data['KONDISI'],
                        'TAHUN Perolehan': m_data['TAHUN Perolehan']
                    }
                    
                    if out_master['JENIS HARTA'] in ['', '-', 'nan', 'None'] or out_master['TAHUN Perolehan'] in ['', '-', 'nan', 'None']:
                        status_recouncil = "Data Ditemukan Tidak Lengkap"
                    elif ruangan_opname.upper() != out_master['LOKASI'].upper() and ruangan_opname:
                        status_recouncil = f"Salah Ruangan seharusnya di {out_master['LOKASI']}"
                    else:
                        status_recouncil = "Sudah Sesuai"
                        
                # Tulis Master Data dan Hasil ke Excel Document (Kolom K sampai Q)
                master_vals = [
                    out_master['NO BARCODE'], out_master['ASSET ORACLE'], out_master['LOKASI'], 
                    out_master['JENIS HARTA'], out_master['KONDISI MASTER'], out_master['TAHUN Perolehan']
                ]
                
                for i, m_val in enumerate(master_vals):
                    cell = ws.cell(row=r, column=11 + i)
                    cell.value = m_val
                    style_cell(cell, is_header=False)
                    
                q_cell = ws.cell(row=r, column=17)
                q_cell.value = status_recouncil
                style_cell(q_cell, is_header=False)

            # Auto Column Width for K-Q
            widths = [15, 15, 30, 30, 10, 15, 35]
            for i, w in enumerate(widths):
                col_letter = get_column_letter(11 + i)
                ws.column_dimensions[col_letter].width = w

            # Save modified file to output path
            base_fname = os.path.basename(opname_path)
            if len(opname_paths) == 1:
                # Single file -> Directly to requested .xlsx output path
                wb.save(output_path)
                processed_files.append(output_path)
            else:
                # Multiple files -> Save to temp, later zipped
                temp_out = os.path.join(os.path.dirname(output_path), f"Processed_{base_fname}")
                wb.save(temp_out)
                processed_files.append(temp_out)
                
        # 3. Handle ZIP archiving if multiple files
        if len(opname_paths) > 1:
            with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for tf in processed_files:
                    zipf.write(tf, os.path.basename(tf))
                    # Cleanup the temporary modified file after zipping
                    os.remove(tf)

        return {"status": "success", "message": "Proses berhasil", "output": output_path, "files_processed": len(opname_paths)}

    except Exception as e:
        err_msg = traceback.format_exc()
        print(f"Error in recouncil processing: {err_msg}")
        return {"status": "error", "message": str(e)}

if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print(json.dumps({"status": "error", "message": "Missing arguments"}))
        sys.exit(1)
        
    cmd = sys.argv[1]
    
    if cmd == "process":
        try:
            data_json = sys.argv[2]
            data = json.loads(data_json)
            
            opname_paths = data.get("opnames")
            master_path = data.get("master")
            output_path = data.get("output_path")
            
            result = run_recouncil(opname_paths, master_path, output_path)
            print(json.dumps(result))
            
        except Exception as e:
            print(json.dumps({"status": "error", "message": str(e)}))
