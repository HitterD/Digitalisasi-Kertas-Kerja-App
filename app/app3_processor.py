import sys
import os
import json
import re
import pandas as pd
import numpy as np
import traceback
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def parse_barcodes(barcode_str):
    if pd.isna(barcode_str) or not isinstance(barcode_str, str) or str(barcode_str).strip() == '':
        return ['']
    cleaned_str = str(barcode_str).replace(';', ',').replace('\n', ',').replace('|', ',')
    special_cases = ['N/A', 'NULL', 'NONE', 'TIDAK ADA', 'KOSONG', '-']
    if cleaned_str.strip().upper() in special_cases:
        return ['']

    tokens = re.split(r'[,\s]+', cleaned_str)
    valid_barcodes = []
    seen = set()
    
    for token in tokens:
        token = token.strip().strip('.')
        if not token: continue
        
        # Valid barcode should be >=3 chars
        if len(token) >= 3:
            has_digit = any(c.isdigit() for c in token)
            has_special = any(c in '-_/' for c in token)
            
            # Keep if it contains digit, OR (no digit but uppercase and contains special chars like QC-ROOM)
            is_valid = has_digit or (has_special and token.isupper() and len(token) >= 4)
            
            if is_valid and token not in seen:
                seen.add(token)
                valid_barcodes.append(token)
                
    if not valid_barcodes:
        return [cleaned_str.strip()]
        
    return valid_barcodes

def expand_and_clean_barcodes(raw_barcode_text):
    barcodes = parse_barcodes(raw_barcode_text)
    expanded_barcodes = []
    for barcode in barcodes:
        if not barcode or barcode == '':
            continue
        range_match = re.match(r'^(\d{6,})\s*-\s*(\d{6,})$', barcode)
        if range_match:
            try:
                start, end = int(range_match.group(1)), int(range_match.group(2))
                if start <= end and end - start <= 200000:
                    for i in range(start, end + 1):
                        expanded_barcodes.append(str(i))
                    continue
            except (ValueError, TypeError):
                pass
        expanded_barcodes.append(barcode)
    seen = set()
    final_barcodes = []
    for bc in expanded_barcodes:
        if bc not in seen:
            seen.add(bc)
            final_barcodes.append(bc)
    return final_barcodes if final_barcodes else ['']

def dynamic_extract(df, file_type):
    # Dynamic header search
    cols_map = {'NO BARCODE': -1, 'ASSET ORACLE': -1, 'LOKASI': -1, 'JENIS HARTA': -1, 'KONDISI': -1, 'PERO_1': -1, 'PERO_2': -1, 'BAT': -1}
    header_idx = -1
    
    for idx, row in df.head(20).iterrows():
        row_strs = [str(x).strip().upper() for x in row.values]
        temp_map = {'NO BARCODE': -1, 'ASSET ORACLE': -1, 'LOKASI': -1, 'JENIS HARTA': -1, 'KONDISI': -1, 'PERO_1': -1, 'PERO_2': -1, 'BAT': -1}
        
        for i, val in enumerate(row_strs):
            if 'BARCODE' in val: temp_map['NO BARCODE'] = i
            elif 'ASSET' in val: temp_map['ASSET ORACLE'] = i
            elif val == 'LOKASI': temp_map['LOKASI'] = i
            elif 'HARTA' in val: temp_map['JENIS HARTA'] = i
            elif 'KONDISI' in val: temp_map['KONDISI'] = i
            elif 'BAT' in val: temp_map['BAT'] = i
            elif 'BULAN' in val and file_type == 'inv': temp_map['PERO_1'] = i
            elif 'TAHUN' in val and file_type == 'inv': temp_map['PERO_2'] = i
            elif 'PERO' in val and temp_map['PERO_1'] == -1: temp_map['PERO_1'] = i
            elif 'PERO' in val and temp_map['PERO_1'] != -1 and temp_map['PERO_2'] == -1: temp_map['PERO_2'] = i
            
        # Consider it a header row if it contains BARCODE and HARTA
        if temp_map['NO BARCODE'] != -1 and temp_map['JENIS HARTA'] != -1:
            header_idx = idx
            cols_map = temp_map
            # for INV, BULAN TAHUN might be 1 row below header.
            if file_type == 'inv' and cols_map['PERO_1'] == -1 and idx+1 < len(df):
                next_row = [str(x).strip().upper() for x in df.iloc[idx+1].values]
                for j, nval in enumerate(next_row):
                    if 'BULAN' in nval: cols_map['PERO_1'] = j
                    elif 'TAHUN' in nval: cols_map['PERO_2'] = j
            break
            
    if header_idx == -1:
        return None
        
    data_df = df.iloc[header_idx+1:]
    
    # Pre-allocate dictionary for the resulting dataframe to save memory over sequential inserts
    res_data = {}
    res_data['NO BARCODE'] = data_df.iloc[:, cols_map['NO BARCODE']].copy() if cols_map['NO BARCODE'] != -1 else pd.Series([""] * len(data_df), index=data_df.index)
    res_data['ASSET ORACLE'] = data_df.iloc[:, cols_map['ASSET ORACLE']].copy() if cols_map['ASSET ORACLE'] != -1 else pd.Series([""] * len(data_df), index=data_df.index)
    res_data['LOKASI'] = data_df.iloc[:, cols_map['LOKASI']].copy() if cols_map['LOKASI'] != -1 else pd.Series([""] * len(data_df), index=data_df.index)
    res_data['JENIS HARTA'] = data_df.iloc[:, cols_map['JENIS HARTA']].copy() if cols_map['JENIS HARTA'] != -1 else pd.Series([""] * len(data_df), index=data_df.index)
    res_data['KONDISI'] = data_df.iloc[:, cols_map['KONDISI']].copy() if cols_map['KONDISI'] != -1 else pd.Series([""] * len(data_df), index=data_df.index)
    res_data['BAT'] = data_df.iloc[:, cols_map['BAT']].copy() if cols_map['BAT'] != -1 else pd.Series([""] * len(data_df), index=data_df.index)
    
    res_df = pd.DataFrame(res_data)
    
    if cols_map['PERO_1'] != -1 and cols_map['PERO_2'] != -1:
        res_df['Tahun Perolehan'] = data_df.apply(
            lambda r: format_custom_date(
                r.iloc[cols_map['PERO_1']], 
                r.iloc[cols_map['PERO_2']],
                r.iloc[cols_map['NO BARCODE']] if cols_map['NO BARCODE'] != -1 else ""
            ), axis=1)
    else:
        res_df['Tahun Perolehan'] = ""
        
    expected_order = ['NO BARCODE', 'ASSET ORACLE', 'LOKASI', 'JENIS HARTA', 'KONDISI', 'Tahun Perolehan', 'BAT']
    res_df = res_df[[c for c in expected_order if c in res_df.columns]]
        
    return res_df

def get_bat_filters(files):
    """
    Ekstrak BAT unik dari file EXA, ADD, INV dengan pencarian dinamis.
    """
    b_filters = set()
    errors = []
    
    for ftype in ['exa', 'add', 'inv']:
        path = files.get(ftype)
        if path and os.path.exists(path):
            try:
                dfs = pd.read_excel(path, header=None, sheet_name=None)
                for df in dfs.values():
                    extracted = dynamic_extract(df, ftype)
                    if extracted is not None and not extracted.empty:
                        b_filters.update(extracted['BAT'].dropna().astype(str).str.strip().unique())
            except Exception as e:
                errors.append(f"{ftype.upper()} BAT Error: {e}")

    # clean empty strings
    clean_filters = [b for b in b_filters if b and b.upper() not in ['NAN', 'NONE', 'NULL']]
    return {
        "status": "success",
        "data": sorted(clean_filters),
        "errors": errors
    }

def format_custom_date(month, year, raw_barcode=""):
    if pd.isna(month) and pd.isna(year):
        return ""
    if pd.isna(month):
        month = 1
    if pd.isna(year):
        year = 1
        
    y_val = None
    if pd.notna(raw_barcode):
        bc_str = str(raw_barcode).strip()
        match = re.search(r'\d{6,}', bc_str)
        if match:
            y_val = match.group(0)[:2] # Extracted year prefix from barcode
            
    try:
        m = int(float(month))
        if y_val is not None:
            y = int(y_val)
        else:
            y = int(float(year))
        
        # Format year to 4 digits intelligently
        if y < 100:
            # If 0-25 (e.g., 4 or 24), map to 2000s
            if y <= 50:
                y = 2000 + y
            # If 51-99, map to 1900s
            else:
                y = 1900 + y
                
        return f"01/{m:02d}/{y}"
    except (ValueError, TypeError):
        return f"{month}/{year}"

def clean_string(val):
    if pd.isna(val):
        return ""
    if isinstance(val, str):
        return val.strip()
    return val

def process_file(file_path, file_type, selected_bats):
    if not file_path or not os.path.exists(file_path):
        return pd.DataFrame()
        
    try:
        dfs = pd.read_excel(file_path, header=None, sheet_name=None)
        all_exploded_rows = []

        for df in dfs.values():
            if df.empty: continue
            
            res_df = dynamic_extract(df, file_type)
            if res_df is None or res_df.empty:
                continue

            # Free up the original df early
            del df 

            # Filter by BAT
            res_df['BAT'] = res_df['BAT'].astype(str).str.strip()
            mask = res_df['BAT'].isin(selected_bats) | res_df['BAT'].str.contains('|'.join([re.escape(b) for b in selected_bats]), case=False, na=False)
            df_filtered = res_df[mask].copy()

            if df_filtered.empty:
                continue

            # Explode barcodes
            for _, row in df_filtered.iterrows():
                barcodes = expand_and_clean_barcodes(row['NO BARCODE'])
                for bc in barcodes:
                    new_row = row.copy()
                    new_row['NO BARCODE'] = bc
                    all_exploded_rows.append(new_row)
                    
        return pd.DataFrame(all_exploded_rows)
    except Exception as e:
        print(f"Error processing {file_type}: {e}")
        return pd.DataFrame()

def _is_empty_val(val):
    """Returns True if a value is considered empty/missing."""
    if val is None:
        return True
    s = str(val).strip()
    return s in ('', 'nan', 'None', 'NaN', '-')

def run_consolidation(files, selected_bats, output_path):
    """
    Konsolidasi data master dengan sumber EXA, ADD, INV.
    
    Skema:
    - File master eksisting → tetap, tidak diubah urutan/isian awalnya.
    - Setiap baris dari EXA, ADD, INV diproses satu per satu:
        * Barcode kosong → skip.
        * Barcode SUDAH ADA di master:
            - Cek setiap kolom pada baris master tersebut.
            - Jika kolom di master kosong DAN kolom sumber ada isinya → isi (update).
            - Tidak menambah baris baru.
        * Barcode BELUM ADA di master → tambahkan sebagai baris baru di bawah.
    - Hasil akhir: TIDAK ADA barcode kembar.
    """
    try:
        master_path = files.get('master')

        # ── Kolom kunci yang dikelola ──────────────────────────────────────────
        KEY_COLS = ['NO BARCODE', 'ASSET ORACLE', 'LOKASI', 'JENIS HARTA', 'KONDISI', 'Tahun Perolehan']

        # ── Garbage filter ─────────────────────────────────────────────────────
        garbage_keywords = ['JUMLAH', 'TOTAL ', 'INVENTARIS YANG ']
        def is_garbage(val):
            return any(str(val).strip().upper().startswith(kw) for kw in garbage_keywords)

        # ── 1. Load Master Data eksisting ──────────────────────────────────────
        master_rows = []       # list of dicts, urutan dipertahankan
        barcode_index = {}     # barcode → index in master_rows  (untuk lookup O(1))

        if master_path and os.path.exists(master_path):
            df_master = pd.read_excel(master_path)
            df_master.replace('', np.nan, inplace=True)

            for col in df_master.columns:
                if df_master[col].dtype == 'object':
                    df_master[col] = df_master[col].apply(clean_string)

            df_master.replace(np.nan, '', inplace=True)

            for _, row in df_master.iterrows():
                row_dict = row.to_dict()
                bc = str(row_dict.get('NO BARCODE', '')).strip()

                # Simpan baris master apa adanya (sudah pasti valid dari sumber lama)
                master_rows.append(row_dict)
                if bc and bc not in ('', 'nan', 'None', 'NaN', '-'):
                    if bc not in barcode_index:          # ambil yang pertama jika ada duplikat di master sendiri
                        barcode_index[bc] = len(master_rows) - 1

        # ── 2. Proses tiap sumber (EXA → ADD → INV) ──────────────────────────
        for ftype in ['exa', 'add', 'inv']:
            df_src = process_file(files.get(ftype), ftype, selected_bats)
            if df_src.empty:
                continue

            # Bersihkan nilai string
            for col in df_src.columns:
                if df_src[col].dtype == 'object':
                    df_src[col] = df_src[col].apply(clean_string)
            df_src.replace(np.nan, '', inplace=True)

            for _, src_row in df_src.iterrows():
                src_bc = str(src_row.get('NO BARCODE', '')).strip()

                # a. Barcode kosong → skip
                if _is_empty_val(src_bc):
                    continue

                # b. Filter garbage berdasarkan JENIS HARTA
                if is_garbage(src_row.get('JENIS HARTA', '')):
                    continue

                # c. Barcode sudah ada di master → cek & isi kolom yang kosong
                if src_bc in barcode_index:
                    idx = barcode_index[src_bc]
                    existing = master_rows[idx]
                    updated = False
                    for col in KEY_COLS:
                        master_val = existing.get(col, '')
                        src_val    = src_row.get(col, '')
                        if _is_empty_val(master_val) and not _is_empty_val(src_val):
                            existing[col] = src_val
                            updated = True
                    # master_rows[idx] sudah dict by reference, tidak perlu re-assign
                    # tapi untuk kejelasan:
                    if updated:
                        master_rows[idx] = existing

                # d. Barcode baru → tambahkan sebagai baris baru
                else:
                    new_row = src_row.to_dict()
                    master_rows.append(new_row)
                    barcode_index[src_bc] = len(master_rows) - 1

        del files  # free memory

        if not master_rows:
            return {"status": "error", "message": "Tidak ada data yang diproses."}

        # ── 3. Bangun DataFrame final ──────────────────────────────────────────
        df_final = pd.DataFrame(master_rows)
        df_final.replace('', np.nan, inplace=True)

        # Buang baris yang TIDAK punya barcode DAN TIDAK punya jenis harta
        df_final.dropna(subset=['NO BARCODE', 'JENIS HARTA'], how='all', inplace=True)
        df_final.replace(np.nan, '', inplace=True)

        # Pastikan urutan kolom konsisten
        expected_order = ['NO BARCODE', 'ASSET ORACLE', 'LOKASI', 'JENIS HARTA', 'KONDISI', 'Tahun Perolehan', 'BAT']
        existing_cols = [c for c in expected_order if c in df_final.columns]
        extra_cols    = [c for c in df_final.columns if c not in expected_order]
        df_final = df_final[existing_cols + extra_cols]

        # ── 4. Tulis ke Excel dengan styling ──────────────────────────────────
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df_final.to_excel(writer, index=False, sheet_name='Master Data')
            workbook  = writer.book
            worksheet = writer.sheets['Master Data']

            worksheet.freeze_panes = 'A2'

            header_fill = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            center_align = Alignment(horizontal='center', vertical='center')
            thin_border  = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'),  bottom=Side(style='thin')
            )

            for col_num, value in enumerate(df_final.columns.values):
                cell = worksheet.cell(row=1, column=col_num + 1)
                cell.fill      = header_fill
                cell.font      = header_font
                cell.alignment = center_align
                cell.border    = thin_border

            widths = [20, 20, 35, 40, 25, 20, 15]
            for i, w in enumerate(widths):
                if i < len(df_final.columns):
                    col_letter = get_column_letter(i + 1)
                    worksheet.column_dimensions[col_letter].width = w

            for row in range(2, len(df_final) + 2):
                for col in range(1, len(df_final.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.alignment = center_align
                    cell.border    = thin_border
                    if col == 6 and cell.value and isinstance(cell.value, pd.Timestamp):
                        cell.number_format = 'dd/mm/yyyy'

        return {
            "status": "success",
            "message": "Proses berhasil",
            "output": output_path,
            "rows": len(df_final)
        }

    except Exception as e:
        return {"status": "error", "message": str(e), "traceback": traceback.format_exc()}

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps({"status": "error", "message": "No command provided"}))
        sys.exit(1)
        
    command = sys.argv[1]
    
    # Check if sys.argv[2] is a file path and read from it
    try:
        if len(sys.argv) > 2:
            input_arg = sys.argv[2]
            if os.path.exists(input_arg):
                with open(input_arg, 'r', encoding='utf-8') as f:
                    data = json.load(f)
            else:
                data = json.loads(input_arg)
        else:
            data = {}
    except Exception as e:
        print(json.dumps({"status": "error", "message": f"Invalid JSON input: {str(e)}"}))
        sys.exit(1)
        
    if command == "get_bat":
        result = get_bat_filters(data.get("files", {}))
        print(json.dumps(result))
    elif command == "process":
        files = data.get("files", {})
        selected_bats = data.get("selected_bats", [])
        output_path = data.get("output_path", "Kamus Data Master new 2.xlsx")
        result = run_consolidation(files, selected_bats, output_path)
        print(json.dumps(result))
    else:
        print(json.dumps({"status": "error", "message": "Unknown command"}))
