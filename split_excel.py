"""
Script untuk memisahkan file Excel multi-sheet menjadi file Excel terpisah per sheet.
Nama file output disesuaikan dengan nama ruangan di Row 3 (B3) setiap sheet.

Cara pakai:
  python split_excel.py <nama_file.xlsx>

Contoh:
  python split_excel.py SJA1-03202601-ICT-20260303.xlsx
"""

import sys
import os
import openpyxl
from copy import copy

def split_excel(input_file):
    if not os.path.exists(input_file):
        print(f"Error: File '{input_file}' tidak ditemukan.")
        sys.exit(1)

    wb = openpyxl.load_workbook(input_file)
    output_dir = os.path.join(os.path.dirname(input_file) or '.', 'output_split')
    os.makedirs(output_dir, exist_ok=True)

    print(f"Memproses: {input_file}")
    print(f"Jumlah sheet: {len(wb.sheetnames)}")
    print(f"Output folder: {output_dir}")
    print("-" * 60)

    for i, sheet_name in enumerate(wb.sheetnames, 1):
        ws = wb[sheet_name]

        # Ambil nama ruangan dari B3
        room_name = ws.cell(row=3, column=2).value
        if room_name:
            room_name = str(room_name).strip()
            # Hapus prefix "RUANG " jika ada
            if room_name.upper().startswith("RUANG "):
                room_name = room_name[6:].strip()
        else:
            room_name = sheet_name

        # Bersihkan karakter yang tidak valid untuk nama file
        safe_name = room_name.replace('\\', '-').replace('/', '-').replace(':', '-')
        safe_name = safe_name.replace('*', '').replace('?', '').replace('"', '')
        safe_name = safe_name.replace('<', '').replace('>', '').replace('|', '')
        safe_name = safe_name.strip()

        # Buat workbook baru dengan sheet ini
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        new_ws.title = sheet_name

        # Copy semua cell termasuk value dan style
        for row in ws.iter_rows():
            for cell in row:
                new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)

        # Copy column widths
        for col_letter, col_dim in ws.column_dimensions.items():
            new_ws.column_dimensions[col_letter].width = col_dim.width

        # Copy row heights
        for row_num, row_dim in ws.row_dimensions.items():
            new_ws.row_dimensions[row_num].height = row_dim.height

        # Copy merged cells
        for merged_range in ws.merged_cells.ranges:
            new_ws.merge_cells(str(merged_range))

        # Save file
        output_file = os.path.join(output_dir, f"{safe_name}.xlsx")
        new_wb.save(output_file)
        asset_count = ws.max_row - 8  # Approximate data rows
        print(f"  [{i}/{len(wb.sheetnames)}] {safe_name}.xlsx ({max(0, asset_count)} aset)")

    print("-" * 60)
    print(f"Selesai! {len(wb.sheetnames)} file tersimpan di: {output_dir}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Penggunaan: python split_excel.py <nama_file.xlsx>")
        print("Contoh: python split_excel.py SJA1-03202601-ICT-20260303.xlsx")
        sys.exit(1)

    split_excel(sys.argv[1])
